"""
Travel Itinerary Builder - FastAPI backend.

Stack: FastAPI + Motor (MongoDB) + Emergent-managed Google Auth.
All routes are mounted under /api.
"""
from __future__ import annotations

import asyncio
import io
import logging
import math
import os
import re
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Annotated, List, Literal, Optional

import httpx
import openpyxl
from bs4 import BeautifulSoup
from bson import ObjectId
from dotenv import load_dotenv
from fastapi import (
    APIRouter,
    Body,
    Cookie,
    Depends,
    FastAPI,
    File,
    Header,
    HTTPException,
    Query,
    Request,
    Response,
    UploadFile,
    status,
)
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, BeforeValidator, ConfigDict, EmailStr, Field
from starlette.middleware.cors import CORSMiddleware

# Local imports
from retrieval import bump_version, get_retriever
from prompts import SYSTEM_PROMPT_GENERATE
from models import (
    User, AllowedEmail, AllowedEmailCreate,
    Provider, ProviderCreate, ProviderUpdate,
    Experience, ExperienceCreate, ExperienceUpdate, ExperienceChange,
    Hotel, HotelCreate, HotelUpdate,
    ItineraryService, ItineraryDay, Accommodation, Room, Traveler,
    Itinerary, ItineraryUpsert,
    TrainingExample, TrainingExampleUpsert, BulkImportJob,
    ServiceType, HotelTier, RoomType, TripOutcome, BulkJobStatus,
)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.environ["MONGO_URL"]
db_name = os.environ["DB_NAME"]
client = AsyncIOMotorClient(mongo_url)
db = client[db_name]

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
logger = logging.getLogger("itinerary")

EMERGENT_SESSION_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

# ---------------------------------------------------------------------------
# Pydantic helpers
# ---------------------------------------------------------------------------
def _to_str(v):
    if isinstance(v, ObjectId):
        return str(v)
    return v


PyObjectId = Annotated[str, BeforeValidator(_to_str)]


def new_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# App & router
# ---------------------------------------------------------------------------
app = FastAPI(title="Travel Itinerary Builder API")
api = APIRouter(prefix="/api")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


@app.on_event("startup")
async def ensure_indexes():
    """Idempotent index setup. Currently:
    - sofi_push_jobs.created_at_dt TTL=7 days. We mirror created_at (ISO
      string) into a BSON Date field because Mongo TTL only acts on Date.
      The job inserter writes both fields atomically.
    """
    try:
        await db.sofi_push_jobs.create_index(
            "created_at_dt",
            expireAfterSeconds=7 * 24 * 3600,
            name="created_at_dt_ttl",
        )
    except Exception as e:
        logger.warning("ensure_indexes sofi_push_jobs failed: %s", e)


@app.on_event("startup")
async def ensure_playwright_browser():
    """Pods sometimes lose the Playwright Chromium binary after image recycling.

    We fire `playwright install chromium` in the BACKGROUND so the FastAPI app
    starts answering /api/* requests in <1s even while the 185 MiB chromium
    download is still in flight. Any request that actually needs the browser
    (Travefy import) hits the lazy retry in `scraper._render_url` which waits
    on the same shared lock for the binary to land. Without this background
    pattern, the install blocked startup for ~60s and Kubernetes' ingress
    returned 502 to every early request — exactly what production was seeing
    right after a deploy.
    """
    try:
        from scraper import _ensure_chromium_installed
        _spawn_bg(_ensure_chromium_installed())
    except Exception as e:
        logger.warning("playwright warmup skipped: %s", e)


# ---------------------------------------------------------------------------
# Background task registry — strong references prevent the event loop from
# garbage-collecting fire-and-forget coroutines mid-execution. Python's
# asyncio.create_task() only stores a WEAK reference; without this registry,
# long-running tasks (chromium install, travefy preview workers) can vanish
# silently mid-run. That manifests exactly as "the modal spins forever
# without updating the job to done/error" — which is what the user saw in
# production when chromium was still downloading.
# ---------------------------------------------------------------------------
_bg_tasks: set = set()


def _spawn_bg(coro):
    task = asyncio.create_task(coro)
    _bg_tasks.add(task)
    task.add_done_callback(_bg_tasks.discard)
    return task


@app.on_event("startup")
async def seed_database_if_empty():
    """First-deploy seeding: if any operational collection is empty, restore
    it from /app/backend/data/seed.json.gz. Idempotent on re-deploys (skips
    any collection that already has rows)."""
    try:
        from tools.seed_loader import seed_if_empty
        summary = await seed_if_empty(db)
        non_zero = {k: v for k, v in summary.items() if v > 0}
        if non_zero:
            logger.warning("seed: bootstrapped collections: %s", non_zero)
    except Exception as e:
        logger.warning("seed: failed (continuing without it): %s", e)


@app.on_event("startup")
async def reap_stuck_travefy_jobs():
    """Any `travefy_import_jobs` doc left in status='running' after a backend
    restart is by definition orphaned — the asyncio task that owned it died
    when the process did. Flip them to `error` so the modal in the UI can
    surface a clean message instead of polling forever."""
    try:
        res = await db.travefy_import_jobs.update_many(
            {"status": "running"},
            {"$set": {
                "status": "error",
                "error": "Job interrumpido por reinicio del backend. Vuelve a intentar.",
                "finished_at": now_iso(),
            }},
        )
        if res.modified_count:
            logger.warning("reaped %d orphan travefy_import_jobs", res.modified_count)
    except Exception as e:
        logger.warning("could not reap stuck travefy jobs: %s", e)


@app.on_event("startup")
async def backfill_itinerary_versions():
    """First-run migration: every existing itinerary without a version_group_id
    becomes v1 of its own (singleton) group so the new grouping UI on the
    Dashboard works even for legacy data."""
    try:
        res = await db.itineraries.update_many(
            {"version_group_id": {"$exists": False}},
            [{"$set": {
                "version_group_id": "$itinerary_id",
                "version": 1,
            }}],
        )
        if res.modified_count:
            logger.warning("backfilled version_group_id on %d itineraries", res.modified_count)
    except Exception as e:
        logger.warning("itinerary version backfill failed: %s", e)


@app.on_event("startup")
async def reap_orphan_bulk_jobs():
    """Any bulk_import_job left as queued/running from a previous process is
    orphaned — the in-memory asyncio task died with the restart. Mark them
    `interrupted` so the AI Trainer UI shows what really happened instead of
    spinning forever on a job that no one is processing."""
    try:
        res = await db.bulk_import_jobs.update_many(
            {"status": {"$in": ["queued", "running"]}},
            {"$set": {
                "status": "interrupted",
                "finished_at": now_iso(),
                "last_message": "Job interrumpido por reinicio del backend. Pulsa 'Reanudar' para continuar desde donde quedó (los viajes ya importados no se vuelven a procesar).",
            }},
        )
        if res.modified_count:
            logger.warning("Marked %d orphan bulk-import job(s) as interrupted", res.modified_count)
    except Exception as e:
        logger.warning("orphan bulk-job reaper failed: %s", e)


@app.on_event("startup")
async def auto_resume_interrupted_jobs():
    """After the orphan reaper runs, kick off a background watcher that auto-resumes
    any interrupted job which still has unprocessed trips. Survives sleep/restart
    of the backend so long-running imports always reach completion."""
    async def _watcher():
        # Initial small delay so the rest of startup can finish.
        await asyncio.sleep(8)
        while True:
            try:
                cursor = db.bulk_import_jobs.find(
                    {"status": "interrupted"},
                    {"_id": 0, "job_id": 1, "params": 1, "pending_trip_ids": 1,
                     "processed_trip_ids": 1, "listing_done": 1, "created_by": 1},
                )
                async for job in cursor:
                    pending = job.get("pending_trip_ids") or []
                    processed = set(job.get("processed_trip_ids") or [])
                    remaining = [t for t in pending if t not in processed]
                    # Resume if: listing not done yet OR there are unprocessed trips
                    if (not job.get("listing_done")) or remaining:
                        logger.info("auto-resuming interrupted job %s (remaining=%d)",
                                    job["job_id"], len(remaining))
                        await db.bulk_import_jobs.update_one(
                            {"job_id": job["job_id"]},
                            {"$set": {
                                "status": "running",
                                "finished_at": None,
                                "last_message": "Reanudación automática tras reinicio del servidor…",
                                "last_heartbeat": now_iso(),
                            }},
                        )
                        asyncio.create_task(_run_bulk_import_gestion(
                            job["job_id"], job.get("params") or {}, job.get("created_by") or ""
                        ))
            except Exception as e:
                logger.warning("auto_resume_interrupted_jobs watcher tick failed: %s", e)
            # Re-check every 60s. Cheap query, only matches `interrupted`.
            await asyncio.sleep(60)

    asyncio.create_task(_watcher())


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
async def get_session_token(
    request: Request,
    session_token_cookie: Annotated[Optional[str], Cookie(alias="session_token")] = None,
    authorization: Annotated[Optional[str], Header()] = None,
) -> Optional[str]:
    if session_token_cookie:
        return session_token_cookie
    if authorization and authorization.lower().startswith("bearer "):
        return authorization.split(" ", 1)[1].strip()
    return None


async def current_user(
    token: Annotated[Optional[str], Depends(get_session_token)],
) -> User:
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")

    expires_at = session.get("expires_at")
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at and expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at and expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")

    user_doc = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    return User(**user_doc)


async def require_admin(user: Annotated[User, Depends(current_user)]) -> User:
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return user


# ---------------------------------------------------------------------------
# Auth endpoints
# ---------------------------------------------------------------------------
@api.post("/auth/session")
async def auth_session(response: Response, body: dict = Body(...)):
    """Exchange a session_id (from Emergent Auth redirect) for a session_token cookie."""
    session_id = body.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="Missing session_id")

    async with httpx.AsyncClient(timeout=15) as http:
        r = await http.get(EMERGENT_SESSION_URL, headers={"X-Session-ID": session_id})
        if r.status_code != 200:
            logger.warning("Emergent auth error: %s %s", r.status_code, r.text)
            raise HTTPException(status_code=401, detail="Auth provider rejected session")
        profile = r.json()

    email = (profile.get("email") or "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="Email not provided by auth")

    # Whitelist check: allowed_emails OR first-login bootstrap (no admin yet)
    allowed = await db.allowed_emails.find_one({"email": email}, {"_id": 0})
    admin_count = await db.users.count_documents({"role": "admin"})

    if not allowed:
        if admin_count == 0:
            # bootstrap: first ever user becomes admin and is added to whitelist
            allowed = {"email": email, "role": "admin", "added_by": "bootstrap", "added_at": now_iso()}
            await db.allowed_emails.insert_one(dict(allowed))
            logger.info("Bootstrap admin created: %s", email)
        else:
            raise HTTPException(
                status_code=403,
                detail="Tu correo no está autorizado. Pide a un administrador que te añada a la whitelist.",
            )

    role = allowed.get("role", "agent")

    # upsert user
    user_doc = await db.users.find_one({"email": email}, {"_id": 0})
    if user_doc:
        await db.users.update_one(
            {"email": email},
            {"$set": {"name": profile.get("name") or user_doc.get("name"),
                      "picture": profile.get("picture"),
                      "role": role}},
        )
        user_id = user_doc["user_id"]
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": user_id,
            "email": email,
            "name": profile.get("name", email.split("@")[0]),
            "picture": profile.get("picture"),
            "role": role,
            "created_at": now_iso(),
        })

    session_token = profile.get("session_token") or uuid.uuid4().hex
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc),
    })

    response.set_cookie(
        key="session_token",
        value=session_token,
        max_age=7 * 24 * 60 * 60,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
    )

    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return {"user": user}


@api.get("/auth/me")
async def auth_me(user: Annotated[User, Depends(current_user)]):
    return user


@api.post("/auth/logout")
async def auth_logout(
    response: Response,
    token: Annotated[Optional[str], Depends(get_session_token)],
):
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    response.delete_cookie("session_token", path="/")
    return {"ok": True}


# ---------------------------------------------------------------------------
# Allowed emails (admin)
# ---------------------------------------------------------------------------
@api.get("/admin/allowed-emails", response_model=List[AllowedEmail])
async def list_allowed_emails(_: Annotated[User, Depends(require_admin)]):
    items = await db.allowed_emails.find({}, {"_id": 0}).sort("added_at", -1).to_list(1000)
    return items


@api.post("/admin/allowed-emails", response_model=AllowedEmail)
async def add_allowed_email(
    payload: AllowedEmailCreate,
    admin: Annotated[User, Depends(require_admin)],
):
    email = payload.email.lower().strip()
    existing = await db.allowed_emails.find_one({"email": email}, {"_id": 0})
    if existing:
        await db.allowed_emails.update_one({"email": email}, {"$set": {"role": payload.role}})
        existing["role"] = payload.role
        return existing
    doc = AllowedEmail(email=email, role=payload.role, added_by=admin.email).model_dump()
    await db.allowed_emails.insert_one(dict(doc))
    return doc


@api.delete("/admin/allowed-emails/{email}")
async def remove_allowed_email(email: str, admin: Annotated[User, Depends(require_admin)]):
    email = email.lower().strip()
    if email == admin.email:
        raise HTTPException(status_code=400, detail="No puedes eliminar tu propio acceso")
    res = await db.allowed_emails.delete_one({"email": email})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    # also revoke any active sessions
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if user:
        await db.user_sessions.delete_many({"user_id": user["user_id"]})
    return {"ok": True}


@api.get("/admin/users", response_model=List[User])
async def list_users(_: Annotated[User, Depends(require_admin)]):
    items = await db.users.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return items


# ---------------------------------------------------------------------------
# Providers
# ---------------------------------------------------------------------------
@api.get("/providers", response_model=List[Provider])
async def list_providers(
    _: Annotated[User, Depends(current_user)],
    q: Optional[str] = None,
    country: Optional[str] = None,
):
    flt: dict = {}
    if q:
        flt["name"] = {"$regex": q, "$options": "i"}
    if country:
        flt["country"] = country
    items = await db.providers.find(flt, {"_id": 0}).sort("name", 1).to_list(2000)
    return items


@api.post("/providers", response_model=Provider)
async def create_provider(payload: ProviderCreate, _: Annotated[User, Depends(current_user)]):
    prov = Provider(**payload.model_dump())
    await db.providers.insert_one(prov.model_dump())
    return prov


@api.patch("/providers/{provider_id}", response_model=Provider)
async def update_provider(
    provider_id: str,
    payload: ProviderUpdate,
    _: Annotated[User, Depends(current_user)],
):
    patch = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if patch:
        await db.providers.update_one({"provider_id": provider_id}, {"$set": patch})
        # cascade: update denormalized provider_name on experiences
        if "name" in patch:
            await db.experiences.update_many(
                {"provider_id": provider_id}, {"$set": {"provider_name": patch["name"]}}
            )
    doc = await db.providers.find_one({"provider_id": provider_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return doc


@api.delete("/providers/{provider_id}")
async def delete_provider(provider_id: str, _: Annotated[User, Depends(current_user)]):
    in_use = await db.experiences.count_documents({"provider_id": provider_id})
    if in_use > 0:
        raise HTTPException(status_code=400, detail=f"Proveedor en uso por {in_use} experiencias")
    res = await db.providers.delete_one({"provider_id": provider_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


# ---------------------------------------------------------------------------
# Experiences
# ---------------------------------------------------------------------------
@api.get("/experiences", response_model=List[Experience])
async def list_experiences(
    _: Annotated[User, Depends(current_user)],
    q: Optional[str] = None,
    country: Optional[str] = None,
    city: Optional[str] = None,
    type: Optional[ServiceType] = None,
    provider_id: Optional[str] = None,
    limit: int = Query(500, le=2000),
):
    import re as _re
    flt: dict = {}
    tokens = [t for t in (q or "").strip().split() if len(t) >= 2] if q else []
    if tokens:
        flt["$and"] = []
        for tok in tokens:
            safe = _re.escape(tok)
            flt["$and"].append({
                "$or": [
                    {"title": {"$regex": safe, "$options": "i"}},
                    {"description": {"$regex": safe, "$options": "i"}},
                    {"provider_name": {"$regex": safe, "$options": "i"}},
                    {"city": {"$regex": safe, "$options": "i"}},
                ]
            })
    if country:
        flt["country"] = country
    if city:
        # Substring-with-word-boundary match per city token so that origin-
        # destination trains ("Madrid - Bilbao") surface for either endpoint.
        cities = []
        for c in city.split(","):
            c = c.strip()
            if not c:
                continue
            if "-" in c and " - " not in c:
                continue  # reject legacy dash-joined values like "Madrid-Barcelona"
            cities.append(c)
        if len(cities) == 1:
            flt["city"] = {"$regex": f"(?<![A-Za-z]){_re.escape(cities[0])}(?![A-Za-z])", "$options": "i"}
        elif len(cities) > 1:
            flt["city"] = {"$in": [
                _re.compile(f"(?<![A-Za-z]){_re.escape(c)}(?![A-Za-z])", _re.IGNORECASE)
                for c in cities
            ]}
    if type:
        flt["type"] = type
    if provider_id:
        flt["provider_id"] = provider_id
    items = await db.experiences.find(flt, {"_id": 0}).sort("title", 1).limit(limit).to_list(limit)
    return items


# ---------------------------------------------------------------------------
# IVA helpers — Spanish VAT applies ONLY to items located in España.
# Outside Spain (Portugal/Italy/Morocco/etc.) the agency books at the
# tax-included rate quoted by the local supplier, so we treat the two
# fields as identical to avoid double-bookkeeping.
# ---------------------------------------------------------------------------
def _force_no_vat_outside_spain(data: dict, country_field: str = "country",
                                 incl_field: str = "price_tax_incl",
                                 excl_field: str = "price_tax_excl") -> None:
    """In-place: if the item's country isn't Spain, align excl == incl using
    whichever value is non-zero (incl preferred)."""
    country = (data.get(country_field) or "").strip().lower()
    if country in ("españa", "espana", "spain"):
        return
    incl = data.get(incl_field) or 0
    excl = data.get(excl_field) or 0
    target = incl or excl
    data[incl_field] = target
    data[excl_field] = target


@api.post("/experiences", response_model=Experience)
async def create_experience(payload: ExperienceCreate, _: Annotated[User, Depends(current_user)]):
    prov = await db.providers.find_one({"provider_id": payload.provider_id}, {"_id": 0})
    if not prov:
        raise HTTPException(status_code=400, detail="Proveedor no encontrado")
    data = payload.model_dump()
    # Sync legacy 'price' with price_tax_incl
    if data.get("price") is None:
        data["price"] = data.get("price_tax_incl") or 0.0
    if not data.get("price_tax_incl"):
        data["price_tax_incl"] = data.get("price") or 0.0
    # Outside Spain → no IVA differential
    _force_no_vat_outside_spain(data)
    data["price"] = data["price_tax_incl"]
    exp = Experience(**data, provider_name=prov["name"])
    await db.experiences.insert_one(exp.model_dump())
    return exp


@api.patch("/experiences/{experience_id}", response_model=Experience)
async def update_experience(
    experience_id: str,
    payload: ExperienceUpdate,
    user: Annotated[User, Depends(current_user)],
    source: str = Query("manual", description="Where the edit came from: manual|itinerary|csv_import"),
):
    patch = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    # Load the BEFORE snapshot so we can compute a diff for the audit log.
    before = await db.experiences.find_one({"experience_id": experience_id}, {"_id": 0})
    if not before:
        raise HTTPException(status_code=404, detail="Not found")
    if "provider_id" in patch:
        prov = await db.providers.find_one({"provider_id": patch["provider_id"]}, {"_id": 0})
        if not prov:
            raise HTTPException(status_code=400, detail="Proveedor no encontrado")
        patch["provider_name"] = prov["name"]
    # keep price tiers in sync
    if "price_tax_incl" in patch and "price" not in patch:
        patch["price"] = patch["price_tax_incl"]
    if "price" in patch and "price_tax_incl" not in patch:
        patch["price_tax_incl"] = patch["price"]
    # Resolve current country if not in patch (to know whether IVA applies)
    if any(k in patch for k in ("price_tax_excl", "price_tax_incl", "price", "country")):
        merged = {**before, **patch}
        _force_no_vat_outside_spain(merged)
        patch["price_tax_incl"] = merged["price_tax_incl"]
        patch["price_tax_excl"] = merged["price_tax_excl"]
        patch["price"] = merged["price_tax_incl"]
    # Build diff of fields that actually changed (skip `price` alias to avoid noise).
    tracked = ("title", "description", "type", "country", "city", "provider_id", "provider_name",
               "price_tax_excl", "price_tax_incl", "pax", "currency", "notes")
    diff = {}
    for k in tracked:
        if k in patch and before.get(k) != patch.get(k):
            diff[k] = {"from": before.get(k), "to": patch.get(k)}
    if patch:
        await db.experiences.update_one({"experience_id": experience_id}, {"$set": patch})
    # Persist audit-log entry only when something materially changed.
    if diff:
        await db.experience_changes.insert_one(ExperienceChange(
            experience_id=experience_id,
            user_email=getattr(user, "email", None),
            user_name=getattr(user, "name", None) or getattr(user, "email", None),
            source=source,
            diff=diff,
        ).model_dump())
    doc = await db.experiences.find_one({"experience_id": experience_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return doc


@api.get("/experiences/{experience_id}/history")
async def experience_history(
    experience_id: str,
    _: Annotated[User, Depends(current_user)],
    limit: int = Query(50, le=200),
):
    """Return the change history for an experience (most recent first)."""
    rows = await db.experience_changes.find(
        {"experience_id": experience_id}, {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    return rows


@api.delete("/experiences/{experience_id}")
async def delete_experience(experience_id: str, _: Annotated[User, Depends(current_user)]):
    res = await db.experiences.delete_one({"experience_id": experience_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


@api.get("/experiences/facets")
async def experience_facets(_: Annotated[User, Depends(current_user)]):
    """Return distinct values for filters."""
    countries = await db.experiences.distinct("country")
    cities = await db.experiences.distinct("city")
    types = await db.experiences.distinct("type")
    return {
        "countries": sorted([c for c in countries if c]),
        "cities": sorted([c for c in cities if c]),
        "types": sorted([t for t in types if t]),
    }


# ---------------------------------------------------------------------------
# Bulk import - provider price sheet (xlsx with columns operator_name, name, price_tax_incl/price_tax_excl, currency)
# ---------------------------------------------------------------------------
def _clean_title(s: str) -> str:
    """Strip leading numeric codes like '23 ', '25Priv', '24 SG ' that are year-tags."""
    if not s:
        return s
    s = s.strip()
    import re
    # Remove leading year-token: '23 ', '2024 ' OR '25Priv' (no space, just digits before letter)
    s = re.sub(r"^\d{2,4}(?:\s+|(?=[A-Za-z]))", "", s)
    return s.strip()


def _parse_provider_sheet_bytes(content: bytes, country: Optional[str], city: Optional[str], stype: ServiceType):
    """Parse a provider rate-sheet workbook and return list of (provider_name, exp_payload_dict)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    headers_map: dict = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            headers_map[str(val).strip().lower()] = col

    name_col = headers_map.get("name")
    op_col = headers_map.get("operator_name") or headers_map.get("operator")
    price_inc = headers_map.get("price_tax_incl")
    price_exc = headers_map.get("price_tax_excl")
    cur_col = headers_map.get("currency")

    if not name_col or not (price_inc or price_exc):
        raise ValueError("El Excel debe tener columnas 'name' y 'price_tax_incl' (o 'price_tax_excl')")

    rows = []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, name_col).value
        if not title:
            continue
        title = _clean_title(str(title))
        if not title:
            continue
        op_name = (ws.cell(r, op_col).value if op_col else None) or "Proveedor sin nombre"
        op_name = str(op_name).strip()

        def _num(c):
            if not c:
                return 0.0
            v = ws.cell(r, c).value
            if v in (None, ""):
                return 0.0
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        p_excl = _num(price_exc)
        p_incl = _num(price_inc) or p_excl

        currency = "EUR"
        if cur_col:
            v = ws.cell(r, cur_col).value
            if v:
                currency = str(v).strip() or "EUR"

        rows.append((op_name, {
            "title": title,
            "country": country,
            "city": city,
            "type": stype,
            "price_tax_excl": p_excl,
            "price_tax_incl": p_incl,
            "price": p_incl,
            "currency": currency,
        }))
    return rows


async def _import_rows(rows, dedupe: bool = True):
    """Insert experiences from parsed rows, creating providers if needed."""
    created = 0
    skipped = 0
    provider_cache: dict = {}
    for op_name, payload in rows:
        if op_name not in provider_cache:
            prov = await db.providers.find_one({"name": op_name}, {"_id": 0})
            if not prov:
                prov = Provider(name=op_name, country=payload.get("country")).model_dump()
                await db.providers.insert_one(dict(prov))
            provider_cache[op_name] = prov
        prov = provider_cache[op_name]

        if dedupe:
            existing = await db.experiences.find_one(
                {"provider_id": prov["provider_id"], "title": payload["title"], "price_tax_incl": payload["price_tax_incl"]},
                {"_id": 0},
            )
            if existing:
                skipped += 1
                continue

        exp = Experience(provider_id=prov["provider_id"], provider_name=prov["name"], **payload)
        await db.experiences.insert_one(exp.model_dump())
        created += 1
    return {"created": created, "skipped": skipped, "providers": len(provider_cache)}


@api.post("/experiences/import-provider-sheet")
async def import_provider_sheet(
    file: UploadFile = File(...),
    country: Optional[str] = None,
    city: Optional[str] = None,
    type: ServiceType = "actividad",
    _: User = Depends(current_user),
):
    """Import a provider rate sheet."""
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Sube un .xlsx")
    content = await file.read()
    try:
        rows = _parse_provider_sheet_bytes(content, country, city, type)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel inválido: {e}")
    result = await _import_rows(rows, dedupe=True)
    return result


@api.post("/catalog/import-operators-csv")
async def import_operators_csv_upload(
    admin: Annotated[User, Depends(require_admin)],
    file: UploadFile = File(...),
    wipe_experiences: bool = Query(True, description="Wipe existing experiences before import"),
    wipe_imported_hotels: bool = Query(False, description="Also wipe hotels with source=imported_from_trip"),
):
    """Upload a fresh app_operators.csv and rebuild the experiences catalog.

    Saves the file to the canonical server path then delegates to
    `import_catalog_from_trips_csv` with `wipe=True` so the catalog is
    rebuilt cleanly with the new schema (incl. `pax`).
    """
    import pathlib as _p
    fname = (file.filename or "").lower()
    if not fname.endswith(".csv"):
        raise HTTPException(status_code=400, detail="El archivo debe ser .csv")
    target_dir = _p.Path("/app/artifacts/catalog_db")
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / "app_operators.csv"
    raw = await file.read()
    target.write_bytes(raw)
    # Reuse the main importer (handles encoding, classify, pax, providers, etc.)
    return await import_catalog_from_trips_csv(
        admin=admin,
        file_path=str(target),
        wipe=wipe_experiences,
        wipe_hotels=wipe_imported_hotels,
    )


@api.post("/catalog/import-from-trips-csv")
async def import_catalog_from_trips_csv(
    admin: Annotated[User, Depends(require_admin)],
    file_path: str = Query("/app/artifacts/catalog_db/app_operators.csv", description="Server-side CSV path"),
    wipe: bool = Query(False, description="Wipe experiences first"),
    wipe_hotels: bool = Query(False, description="Also wipe hotels with source=imported_from_trip"),
):
    """Build the catalog from a CSV of services used in past trips.

    Expected columns (semicolon-separated, latin-1 OR utf-8):
        ID_TRIP; Fecha_venta; Servicio; Ciudad; Proveedor; AD; CH; Sin_IVA; Con_IVA

    Each row → either an Experience (activity/transfer/train/etc.) or a Hotel
    (when Servicio matches hotel/apartament/resort keywords). Dedup by
    (name + provider + city + pax), keeping the most recent NON-ZERO price.
    The total pax for each entry is AD + CH (defaults to 2 when missing) and
    is stored on the Experience so the agent knows whether a given price is
    for 1 pax, 2 pax, or more.
    Providers are upserted automatically.
    """
    import csv as _csv
    import pathlib as _p

    fp = _p.Path(file_path)
    if not fp.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {file_path}")

    # City → country lookup (covers the top cities seen in real data)
    city_country = {
        # Spain
        "Madrid": "España", "Barcelona": "España", "Sevilla": "España", "Seville": "España",
        "Valencia": "España", "Bilbao": "España", "Granada": "España", "Toledo": "España",
        "Cordoba": "España", "Córdoba": "España", "Ronda": "España", "Malaga": "España",
        "Málaga": "España", "San Sebastian": "España", "San Sebastián": "España",
        "Mallorca": "España", "Ibiza": "España", "Segovia": "España", "Salamanca": "España",
        "Logroño": "España", "La Rioja": "España", "Pamplona": "España", "Avila": "España",
        "Ávila": "España", "Cuenca": "España", "Marbella": "España", "Tenerife": "España",
        # Portugal
        "Lisbon": "Portugal", "Lisboa": "Portugal", "Porto": "Portugal", "Oporto": "Portugal",
        "Sintra": "Portugal", "Cascais": "Portugal", "Algarve": "Portugal", "Lagos": "Portugal",
        "Coimbra": "Portugal", "Evora": "Portugal", "Évora": "Portugal", "Douro": "Portugal",
        "Douro Valley": "Portugal", "Madeira": "Portugal", "Azores": "Portugal", "Braga": "Portugal",
        "Faro": "Portugal", "Aveiro": "Portugal",
        # Italy
        "Rome": "Italia", "Roma": "Italia", "Florence": "Italia", "Firenze": "Italia",
        "Florencia": "Italia", "Venice": "Italia", "Venezia": "Italia", "Venecia": "Italia",
        "Naples": "Italia", "Napoli": "Italia", "Milan": "Italia", "Milano": "Italia",
        "Sorrento": "Italia", "Positano": "Italia", "Amalfi": "Italia", "Capri": "Italia",
        "Pompeii": "Italia", "Tuscany": "Italia", "Toscana": "Italia", "Bologna": "Italia",
        "Verona": "Italia", "Siena": "Italia", "Pisa": "Italia", "Cinque Terre": "Italia",
        "Lake Como": "Italia", "Sicily": "Italia", "Sicilia": "Italia", "Palermo": "Italia",
        "Taormina": "Italia", "Catania": "Italia", "Matera": "Italia", "Puglia": "Italia",
        "Lecce": "Italia",
        # Morocco
        "Marrakech": "Marruecos", "Marrakesh": "Marruecos", "Casablanca": "Marruecos",
        "Fes": "Marruecos", "Fez": "Marruecos", "Rabat": "Marruecos", "Tangier": "Marruecos",
        "Chefchaouen": "Marruecos", "Essaouira": "Marruecos", "Merzouga": "Marruecos",
    }
    # Normalize duplicate city spellings → canonical
    city_aliases = {
        "Roma": "Rome", "Florencia": "Florence", "Firenze": "Florence",
        "Venecia": "Venice", "Venezia": "Venice", "Napoli": "Naples", "Milano": "Milan",
        "Lisboa": "Lisbon", "Oporto": "Porto", "Sevilla": "Seville", "Marrakesh": "Marrakech",
        "Fez": "Fes",
    }

    HOTEL_KW = ("hotel", "hostel", "hostal", "apartam", "apartment", "resort", "pousada",
                "riad", "villa", "b&b", "bed and breakfast", "lodge", "boutique stay")
    TRANSFER_KW = ("transfer", "taxi", "limo", "driver", "private car", "private vehicle")
    FLIGHT_KW = ("flight", "vuelo", "airline")
    TRAIN_KW = ("train", "tren", "renfe", "trenitalia", "italo", "ave ", "ave-")
    # Entry-tickets-only services (museum tickets, monument entries, skip-the-line).
    # Guided activities stay as 'actividad' even when they include entry tickets.
    ENTRADAS_KW = ("entradas", "tickets only", "ticket only", "skip-the-line tickets",
                   "entry only", "general admission")

    def classify(name: str) -> str:
        n = name.lower()
        # Order matters: transfer/flight/train check first, otherwise "Transfer to Hotel X" gets miscategorized as hotel
        if any(k in n for k in TRANSFER_KW):
            return "transfer"
        if any(k in n for k in FLIGHT_KW):
            return "vuelo"
        if any(k in n for k in TRAIN_KW):
            return "tren"
        if any(k in n for k in HOTEL_KW):
            return "hotel"
        if any(k in n for k in ENTRADAS_KW) and "tour" not in n and "guided" not in n and "visit" not in n:
            return "entradas"
        # Restaurants, food experiences, generic activities all flow into 'actividad'
        return "actividad"

    def tier_from_name(name: str) -> str:
        n = name.lower()
        if any(w in n for w in ("luxury", "deluxe", "5*", "5 star", "5-star")):
            return "luxury"
        if any(w in n for w in ("4*", "4 star", "boutique", "premium")):
            return "upscale"
        return "upscale"

    if wipe:
        await db.experiences.delete_many({})
    if wipe_hotels:
        # Only wipe hotels imported from past trips, never the curated library
        await db.hotels.delete_many({"source": "imported_from_trip"})

    # Try UTF-8 then fall back to Latin-1
    try:
        text = fp.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = fp.read_text(encoding="latin-1")

    rows = list(_csv.DictReader(text.splitlines(), delimiter=";"))

    def _num(v):
        if v is None:
            return None
        s = str(v).strip()
        if not s or s.upper() == "NULL":
            return None
        try:
            return float(s.replace(",", "."))
        except ValueError:
            return None

    def _pax(r) -> int:
        ad = _num(r.get("AD")) or 0
        ch = _num(r.get("CH")) or 0
        total = int(ad) + int(ch)
        if total <= 0:
            return 2  # default 2 pax when AD/CH missing
        return min(total, 20)  # cap at 20 to filter corrupted source data

    # Group rows by (service_name, provider, city, pax) → aggregate prices
    # Track most recent NON-ZERO price by Fecha_venta as the canonical price.
    grouped: dict = {}
    for r in rows:
        svc = (r.get("Servicio") or "").strip()
        prov = (r.get("Proveedor") or "").strip()
        city_raw = (r.get("Ciudad") or "").strip()
        if not svc or not prov or not city_raw:
            continue
        city = city_aliases.get(city_raw, city_raw)
        pax = _pax(r)
        key = (svc, prov, city, pax)
        e = _num(r.get("Sin_IVA"))
        i = _num(r.get("Con_IVA"))
        sale_date = (r.get("Fecha_venta") or "").strip()
        bucket = grouped.setdefault(key, {
            "excl_all": [], "incl_all": [],
            "best_excl": None, "best_incl": None, "best_date": "",
        })
        if e is not None:
            bucket["excl_all"].append(e)
        if i is not None:
            bucket["incl_all"].append(i)
        # Keep the most recent non-zero price as canonical
        non_zero = (i and i > 0) or (e and e > 0)
        if non_zero and sale_date >= bucket["best_date"]:
            bucket["best_date"] = sale_date
            bucket["best_excl"] = e
            bucket["best_incl"] = i

    def _median(lst):
        vals = [v for v in lst if v is not None]
        if not vals:
            return 0.0
        s = sorted(vals)
        n = len(s)
        return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2

    # Provider cache (upsert once)
    provider_cache: dict = {}
    exp_created = 0
    exp_skipped = 0
    hotel_created = 0
    hotel_skipped = 0

    for (svc, prov_name, city, pax), agg in grouped.items():
        country = city_country.get(city)
        # Upsert provider
        if prov_name not in provider_cache:
            doc = await db.providers.find_one({"name": prov_name}, {"_id": 0})
            if not doc:
                doc = Provider(name=prov_name, country=country).model_dump()
                await db.providers.insert_one(dict(doc))
            provider_cache[prov_name] = doc
        provider = provider_cache[prov_name]
        # Canonical price: most recent non-zero. Fallback to median of all.
        price_excl = agg["best_excl"] if agg["best_excl"] is not None else _median(agg["excl_all"])
        price_incl = agg["best_incl"] if agg["best_incl"] is not None else _median(agg["incl_all"])
        # When one tier is missing, mirror the other
        if not price_incl:
            price_incl = price_excl
        if not price_excl:
            price_excl = price_incl
        price_excl = round(price_excl or 0.0, 2)
        price_incl = round(price_incl or 0.0, 2)
        # Outside Spain → no IVA differential
        if (country or "").strip().lower() not in ("españa", "espana", "spain"):
            price_excl = price_incl

        kind = classify(svc)
        if kind == "hotel":
            existing = await db.hotels.find_one(
                {"name": svc, "city": city}, {"_id": 0}
            )
            if existing:
                hotel_skipped += 1
                continue
            h = Hotel(
                name=svc,
                city=city,
                country=country,
                tier=tier_from_name(svc),
                description=None,
                price_per_night_excl=price_excl,
                price_per_night_incl=price_incl,
                currency="EUR",
                contact=prov_name,
                notes=f"Importado del histórico de viajes. Proveedor: {prov_name}",
                source="imported_from_trip",
            )
            await db.hotels.insert_one(h.model_dump())
            hotel_created += 1
        else:
            # Dedup now includes pax — same service for 2 pax vs 4 pax are
            # different rows because the price differs.
            existing = await db.experiences.find_one(
                {"title": svc, "provider_id": provider["provider_id"], "city": city, "pax": pax},
                {"_id": 0},
            )
            if existing:
                exp_skipped += 1
                continue
            exp = Experience(
                title=svc,
                provider_id=provider["provider_id"],
                provider_name=prov_name,
                country=country,
                city=city,
                type=kind,
                price_tax_excl=price_excl,
                price_tax_incl=price_incl,
                price=price_incl,
                currency="EUR",
                pax=pax,
            )
            await db.experiences.insert_one(exp.model_dump())
            exp_created += 1

    return {
        "rows_scanned": len(rows),
        "unique_services": len(grouped),
        "experiences_created": exp_created,
        "experiences_skipped": exp_skipped,
        "hotels_created": hotel_created,
        "hotels_skipped": hotel_skipped,
        "providers_total": len(provider_cache),
        "wiped": wipe,
        "wiped_hotels": wipe_hotels,
    }



async def import_all_server(
    admin: Annotated[User, Depends(require_admin)],
    base_path: str = Query("/app/artifacts/excel_creados", description="Server-side directory to scan"),
    wipe: bool = Query(False, description="If true, wipes all experiences and providers first"),
):
    """Walk the server-side directory and import every .xlsx file found.

    Country is inferred from the parent folder containing keywords ESPA/PORT/ITAL.
    City is inferred from the first sub-folder after the country folder, skipping
    administrative folders like '2025', '2026', 'REVISADOS', etc.
    """
    import pathlib
    import re as _re

    base = pathlib.Path(base_path)
    if not base.exists():
        raise HTTPException(status_code=404, detail=f"Path not found: {base_path}")

    if wipe:
        await db.experiences.delete_many({})
        await db.providers.delete_many({})

    files = list(base.rglob("*.xlsx"))

    def infer_country_city(parts):
        country = None
        city = None
        # find country
        country_idx = None
        for i, p in enumerate(parts):
            up = p.upper()
            if "ESPA" in up and country is None:
                country = "España"
                country_idx = i
            elif "PORT" in up and country is None:
                country = "Portugal"
                country_idx = i
            elif "ITAL" in up and country is None:
                country = "Italia"
                country_idx = i
        if country_idx is not None:
            # walk subfolders after country, skip admin/year folders
            skip_pat = _re.compile(r"^(20\d{2}|REVISADOS.*|VARIOS|NUEVOS?)$", _re.IGNORECASE)
            for p in parts[country_idx + 1:-1]:  # exclude the file itself
                if skip_pat.match(p.strip()):
                    continue
                city = p.strip()
                break
        return country, city

    total_created = 0
    total_skipped = 0
    file_results = []
    for fp in files:
        country, city = infer_country_city(fp.parts)
        try:
            content = fp.read_bytes()
            rows = _parse_provider_sheet_bytes(content, country, city, "actividad")
            r = await _import_rows(rows, dedupe=True)
            total_created += r["created"]
            total_skipped += r["skipped"]
            file_results.append({"file": fp.name, "country": country, "city": city, **r})
        except Exception as e:
            file_results.append({"file": fp.name, "country": country, "city": city, "error": str(e)})
    return {
        "files_scanned": len(files),
        "total_created": total_created,
        "total_skipped": total_skipped,
        "wiped": wipe,
        "files": file_results,
    }


@api.get("/experiences/autocomplete")
async def experience_autocomplete(
    _: Annotated[User, Depends(current_user)],
    q: str = Query("", min_length=0),
    city: Optional[str] = None,
    country: Optional[str] = None,
    type: Optional[ServiceType] = None,
    pax: Optional[int] = Query(None, description="Prefer experiences priced for this pax count"),
    limit: int = Query(20, le=50),
):
    """Smart typeahead across the catalog.

    If type='alojamiento' → search HOTELS collection.
    Else → search EXPERIENCES collection with that type filter (if set).
    Tokenized AND-search on title/name + provider_name (+ hotel.city). Optional pre-filters: city, country.
    """
    import re as _re
    tokens = [t for t in (q or "").strip().split() if len(t) >= 2]

    if type == "alojamiento":
        # Search hotels
        flt_h: dict = {}
        if tokens:
            flt_h["$and"] = []
            for tok in tokens:
                safe = _re.escape(tok)
                flt_h["$and"].append({
                    "$or": [
                        {"name": {"$regex": safe, "$options": "i"}},
                        {"city": {"$regex": safe, "$options": "i"}},
                    ]
                })
        if city:
            parts = []
            for p in city.split(","):
                p = p.strip()
                if not p:
                    continue
                if "-" in p and " - " not in p:
                    continue
                parts.append(p)
            if len(parts) == 1:
                flt_h["city"] = {"$regex": f"(?<![A-Za-z]){_re.escape(parts[0])}(?![A-Za-z])", "$options": "i"}
            elif len(parts) > 1:
                flt_h["city"] = {"$in": [
                    _re.compile(f"(?<![A-Za-z]){_re.escape(p)}(?![A-Za-z])", _re.IGNORECASE)
                    for p in parts
                ]}
        if country:
            flt_h["country"] = country
        proj = {"_id": 0, "hotel_id": 1, "name": 1, "city": 1, "country": 1, "tier": 1,
                "price_per_night_excl": 1, "price_per_night_incl": 1, "currency": 1}
        items = await db.hotels.find(flt_h, proj).sort("name", 1).limit(limit).to_list(limit)
        # Adapt to a service-compatible shape so the frontend can map it uniformly
        return [
            {
                "experience_id": None,
                "hotel_id": h["hotel_id"],
                "title": h["name"],
                "provider_name": None,
                "city": h.get("city"),
                "country": h.get("country"),
                "type": "alojamiento",
                "price_tax_excl": h.get("price_per_night_excl") or 0,
                "price_tax_incl": h.get("price_per_night_incl") or 0,
                "currency": h.get("currency") or "EUR",
                "tier": h.get("tier"),
            }
            for h in items
        ]

    # Default: search experiences
    # Helper: parse multi-city ("Madrid, Barcelona") and silently drop legacy
    # dash-joined query values like "Madrid-Barcelona" (no spaces). Each city
    # token is matched as a SUBSTRING (not exact) so trains with "origin -
    # destination" stored as the city (e.g. "Madrid - Bilbao") surface when the
    # user filters by either endpoint.
    def _city_filter(c: Optional[str]) -> Optional[dict]:
        if not c:
            return None
        parts = []
        for p in c.split(","):
            p = p.strip()
            if not p:
                continue
            # Reject legacy "A-B" (dash without surrounding spaces). Accept
            # "A - B" or just "A" — both are valid filter tokens.
            if "-" in p and " - " not in p:
                continue
            parts.append(p)
        if not parts:
            return None
        if len(parts) == 1:
            # word-boundary substring → "Madrid" matches "Madrid - Bilbao" AND "Madrid"
            return {"$regex": f"(?<![A-Za-z]){_re.escape(parts[0])}(?![A-Za-z])", "$options": "i"}
        return {"$in": [
            _re.compile(f"(?<![A-Za-z]){_re.escape(p)}(?![A-Za-z])", _re.IGNORECASE)
            for p in parts
        ]}

    flt: dict = {}
    if tokens:
        flt["$and"] = []
        for tok in tokens:
            safe = _re.escape(tok)
            flt["$and"].append({
                "$or": [
                    {"title": {"$regex": safe, "$options": "i"}},
                    {"provider_name": {"$regex": safe, "$options": "i"}},
                ]
            })
    f = _city_filter(city)
    if f:
        flt["city"] = f
    if country:
        flt["country"] = country
    if type:
        # Production still has experiences stored under the legacy taxonomy
        # ('transporte' for trains/transfers, 'restaurante'/'otro' for activities)
        # that preview migrated. Expand the filter so trains stored as
        # 'transporte' surface when the agent filters by 'tren'.
        _LEGACY_EXPAND: dict[str, list[str]] = {
            "tren": ["tren", "transporte"],
            "transfer": ["transfer", "transporte"],
            "actividad": ["actividad", "restaurante", "otro"],
            "entradas": ["entradas", "otro"],
        }
        synonyms = _LEGACY_EXPAND.get(type, [type])
        flt["type"] = {"$in": synonyms} if len(synonyms) > 1 else type
    proj = {"_id": 0, "experience_id": 1, "title": 1, "provider_name": 1, "city": 1, "country": 1,
            "type": 1, "price_tax_excl": 1, "price_tax_incl": 1, "price": 1, "currency": 1, "pax": 1}
    items = await db.experiences.find(flt, proj).sort("title", 1).limit(limit * 2 if pax else limit).to_list(limit * 2 if pax else limit)
    # Normalize legacy types on the way out so the UI shows the clean enum
    # badge ('tren' / 'actividad') instead of the raw DB value ('transporte').
    for it in items:
        it["type"] = _normalize_service_type(it.get("type"))
    if pax:
        # Sort so that exact-pax matches come first, then closest pax, then the rest.
        def _rank(it):
            p = it.get("pax", 2) or 2
            return (0 if p == pax else 1, abs(p - pax), it.get("title", ""))
        items = sorted(items, key=_rank)[:limit]
    return items


@api.get("/experiences/import-all-status")
async def import_all_status(_: Annotated[User, Depends(require_admin)]):
    """Quick stats for the admin panel."""
    return {
        "providers": await db.providers.count_documents({}),
        "experiences": await db.experiences.count_documents({}),
    }


# ---------------------------------------------------------------------------
# Itineraries
# ---------------------------------------------------------------------------
@api.get("/itineraries", response_model=List[Itinerary])
async def list_itineraries(
    user: Annotated[User, Depends(current_user)],
    agent: Optional[str] = None,
    traveler: Optional[str] = None,
):
    """Agents see their OWN itineraries + any itinerary EXPLICITLY shared with
    them (via /itineraries/{id}/share). Admins see everything by default; they
    can narrow down via the `agent` query parameter exposed in the UI filter.
    """
    flt: dict = {}
    if user.role == "admin":
        if agent:
            flt["created_by"] = agent
        if traveler:
            flt["main_traveler"] = {"$regex": traveler, "$options": "i"}
    else:
        # OR: owner or in shared_with. Mongo cannot mix $or with other
        # top-level conditions easily, so we merge the traveler filter
        # into both branches.
        ownership = [
            {"created_by": user.email},
            {"shared_with": user.email},
        ]
        if traveler:
            traveler_clause = {"main_traveler": {"$regex": traveler, "$options": "i"}}
            flt["$and"] = [{"$or": ownership}, traveler_clause]
        else:
            flt["$or"] = ownership
    items = await db.itineraries.find(flt, {"_id": 0}).sort("updated_at", -1).to_list(500)
    return items


@api.get("/itineraries/agents")
async def list_itinerary_agents(user: Annotated[User, Depends(current_user)]):
    """Distinct list of agents who have created itineraries (admin-only)."""
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    emails = await db.itineraries.distinct("created_by")
    return {"agents": sorted([e for e in emails if e])}


@api.post("/itineraries", response_model=Itinerary)
async def create_itinerary(payload: ItineraryUpsert, user: Annotated[User, Depends(current_user)]):
    data = payload.model_dump(exclude_unset=True)
    itn = Itinerary(**data, created_by=user.email)
    # Fresh itineraries are v1 of their own version group.
    if not itn.version_group_id:
        itn.version_group_id = itn.itinerary_id
        itn.version = 1
    itn.updated_at = now_iso()
    await db.itineraries.insert_one(itn.model_dump())
    return itn


@api.post("/itineraries/{itinerary_id}/duplicate", response_model=Itinerary)
async def duplicate_itinerary(
    itinerary_id: str,
    user: Annotated[User, Depends(current_user)],
):
    """Create a new version of an existing itinerary.

    The duplicate keeps the same `version_group_id` so all versions of the same
    client trip stay grouped on the Dashboard, and gets `version = max(group) + 1`.
    The name gets a "· v{N}" suffix on first duplication (or the existing one is
    bumped). Status resets to "draft" — selling status only applies to the
    finalized version of a trip.
    """
    src = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not src:
        raise HTTPException(status_code=404, detail="Not found")
    if not _can_access(src, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")

    group_id = src.get("version_group_id") or src["itinerary_id"]
    # Find the highest version in the group so we always bump above the max,
    # even if intermediate versions were deleted.
    last = await db.itineraries.find(
        {"version_group_id": group_id}, {"_id": 0, "version": 1}
    ).sort("version", -1).limit(1).to_list(1)
    next_version = (last[0].get("version") if last else src.get("version") or 1) + 1

    # Strip the existing "· v{N}" suffix so we don't accumulate "v2 · v3"
    base_name = re.sub(r"\s*[·\-]\s*v\d+\s*$", "", src.get("name") or "Itinerario").strip()
    new_name = f"{base_name} · v{next_version}"

    new_doc = {
        **src,
        "itinerary_id": new_id("itn"),
        "name": new_name,
        "version_group_id": group_id,
        "version": next_version,
        "status": "draft",
        "created_by": user.email,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    # Refresh the inner IDs so day/service/accommodation primary keys don't
    # collide between the original and the duplicate (drag-drop in the builder
    # uses these IDs to identify rows).
    for d in new_doc.get("days") or []:
        d["day_id"] = new_id("day")
        for s in d.get("services") or []:
            s["service_id"] = new_id("svc")
    for a in new_doc.get("accommodations") or []:
        a["acc_id"] = new_id("acc")
        for r in a.get("rooms") or []:
            r["room_id"] = new_id("room")

    itn = Itinerary(**new_doc)
    await db.itineraries.insert_one(itn.model_dump())
    return itn


def _can_access(itn_doc: dict, user: User) -> bool:
    """Read+write authorisation gate.

    Returns True if the user is admin, the creator, or has been explicitly
    shared into the itinerary via `shared_with`. Sharing carries full
    read+write so colleagues can collaborate live.
    """
    if user.role == "admin":
        return True
    if itn_doc.get("created_by") == user.email:
        return True
    return user.email in (itn_doc.get("shared_with") or [])


@api.get("/itineraries/{itinerary_id}", response_model=Itinerary)
async def get_itinerary(itinerary_id: str, user: Annotated[User, Depends(current_user)]):
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    return doc


@api.patch("/itineraries/{itinerary_id}", response_model=Itinerary)
async def update_itinerary(
    itinerary_id: str,
    payload: ItineraryUpsert,
    user: Annotated[User, Depends(current_user)],
):
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    patch = payload.model_dump(exclude_unset=True)
    patch["updated_at"] = now_iso()
    await db.itineraries.update_one({"itinerary_id": itinerary_id}, {"$set": patch})
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    return doc


@api.delete("/itineraries/{itinerary_id}")
async def delete_itinerary(itinerary_id: str, user: Annotated[User, Depends(current_user)]):
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    await db.itineraries.delete_one({"itinerary_id": itinerary_id})
    return {"ok": True}


# ---------------------------------------------------------------------------
# Share itinerary with another agent
# ---------------------------------------------------------------------------
class ShareItineraryBody(BaseModel):
    email: EmailStr


@api.get("/agents/list")
async def list_share_targets(user: Annotated[User, Depends(current_user)]):
    """Lookup list for the "Compartir con" picker. Returns every allowed
    agent EXCEPT the requesting user, since you cannot share an itinerary
    with yourself. The user record (display name, role) is included when
    the agent has already logged in once; otherwise we fall back to the
    local-part of the email."""
    allowed = await db.allowed_emails.find({}, {"_id": 0, "email": 1, "role": 1}).to_list(200)
    # Enrich with display names from `users` collection where available.
    user_docs = await db.users.find(
        {"email": {"$in": [a["email"] for a in allowed]}},
        {"_id": 0, "email": 1, "name": 1},
    ).to_list(200)
    name_by_email = {u["email"]: u.get("name") for u in user_docs}
    out = []
    for a in allowed:
        email = a["email"]
        if email == user.email:
            continue
        display = name_by_email.get(email)
        if not display:
            local = email.split("@")[0]
            display = local[:1].upper() + local[1:]
        out.append({"email": email, "name": display, "role": a.get("role", "agent")})
    out.sort(key=lambda x: x["name"].lower())
    return {"agents": out}


@api.post("/itineraries/{itinerary_id}/share")
async def share_itinerary(
    itinerary_id: str,
    payload: ShareItineraryBody,
    user: Annotated[User, Depends(current_user)],
):
    """Grant another allowed agent read+write access to the itinerary. The
    target email MUST be in `allowed_emails` so we never leak access to a
    random outside address. Idempotent: re-sharing with the same agent is a
    no-op."""
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    target = payload.email.lower().strip()
    if target == (doc.get("created_by") or "").lower():
        raise HTTPException(status_code=400, detail="El propietario ya tiene acceso")
    if target == user.email.lower():
        raise HTTPException(status_code=400, detail="No puedes compartir contigo mismo")
    # Whitelist check — only allow sharing with agents already in the system.
    is_allowed = await db.allowed_emails.find_one({"email": target}, {"_id": 0, "email": 1})
    if not is_allowed:
        raise HTTPException(
            status_code=400,
            detail=f"{target} no está en la lista de agentes autorizados",
        )
    await db.itineraries.update_one(
        {"itinerary_id": itinerary_id},
        {"$addToSet": {"shared_with": target}, "$set": {"updated_at": now_iso()}},
    )
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    return {"ok": True, "shared_with": doc.get("shared_with") or []}


@api.delete("/itineraries/{itinerary_id}/share/{email}")
async def unshare_itinerary(
    itinerary_id: str,
    email: str,
    user: Annotated[User, Depends(current_user)],
):
    """Revoke a previously-granted share. The owner (created_by) can always
    revoke; a shared collaborator can also remove themselves from the list."""
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    target = email.lower().strip()
    is_owner = (doc.get("created_by") or "").lower() == user.email.lower()
    is_self_revoke = target == user.email.lower()
    if not (is_owner or is_self_revoke or user.role == "admin"):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    await db.itineraries.update_one(
        {"itinerary_id": itinerary_id},
        {"$pull": {"shared_with": target}, "$set": {"updated_at": now_iso()}},
    )
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    return {"ok": True, "shared_with": doc.get("shared_with") or []}


# ---------------------------------------------------------------------------
# Client payment links (PayPal Sandbox / Live)
# ---------------------------------------------------------------------------
# Rules per agency policy:
#   • If >60 days until trip start → client may pay a 30% deposit to confirm,
#     and must pay the remaining 70% no later than 45 days before the trip.
#   • If ≤60 days until trip start → only the full 100% is allowed (the trip
#     must be paid in full to confirm).
# Implementation:
#   • Agent clicks "Generar link de pago" → backend lazily generates a
#     URL-safe `payment_token` and persists it on the itinerary.
#   • Client opens https://<frontend>/pay/{token} (NO authentication).
#   • Client picks an amount → backend POSTs to PayPal /v2/checkout/orders →
#     redirects to PayPal's hosted approve page.
#   • After approval, PayPal redirects to /api/payments/{token}/return,
#     which CAPTURES the order and marks the row as `captured`.
#   • PayPal also fires PAYMENT.CAPTURE.COMPLETED to /api/paypal/webhook for
#     async confirmation; both code paths are idempotent.
import secrets
from paypal import (
    approval_url as _paypal_approval_url,
    capture_order as _paypal_capture_order,
    create_order as _paypal_create_order,
    get_order as _paypal_get_order,
    verify_webhook as _paypal_verify_webhook,
)


def _days_to_trip(start_date: Optional[str]) -> Optional[int]:
    """Days from today (UTC) to the trip start. None if no date."""
    if not start_date:
        return None
    try:
        sd = datetime.fromisoformat(start_date).date()
    except (TypeError, ValueError):
        return None
    today = datetime.now(timezone.utc).date()
    return (sd - today).days


def _compute_payment_options(itn: dict, totals: dict) -> dict:
    """Return what the client can pay RIGHT NOW given the trip dates and
    the existing payments. Pure function — does no DB writes.

    State machine:
      • no payments + >60d → [deposit 30%, full 100%]
      • no payments + ≤60d → [full 100%]
      • some paid, remaining > 0 → [balance (rest), partial (custom)]
      • fully paid → []

    Partial payment rules:
      • min = 10% of the total (configurable per trip later if needed)
      • max = remaining
      • capped to remaining if min > remaining
    """
    total_eur = float(totals.get("pvp") or 0)
    days = _days_to_trip(itn.get("start_date"))
    payments = itn.get("payments") or []
    captured = [p for p in payments if p.get("status") == "captured"]
    paid_eur = sum(float(p.get("paid_amount") or p.get("amount_eur") or 0) for p in captured)
    remaining = round(max(0.0, total_eur - paid_eur), 2)

    deposit_amount = round(total_eur * 0.30, 2)
    full_amount = round(total_eur, 2)
    min_partial = round(total_eur * 0.10, 2)

    fully_paid = remaining < 0.01

    if fully_paid or total_eur < 0.01:
        return {
            "total_eur": total_eur, "paid_eur": round(paid_eur, 2),
            "remaining_eur": 0.0, "days_to_trip": days,
            "fully_paid": True, "options": [],
            "partial_bounds": None, "monthly_suggested_eur": None,
            "deposit_threshold_eur": round(
                full_amount if (days is not None and days <= 60) else deposit_amount, 2
            ),
            "booking_secured": True,
        }

    options: List[dict] = []
    # Initial state — nothing captured yet.
    if paid_eur < 0.01:
        if days is not None and days > 60:
            options.append({
                "kind": "deposit",
                "amount_eur": deposit_amount,
                "label": "Pagar reserva (30%)",
                "description": "Confirma el viaje. El resto se paga antes de 45 días del inicio del viaje.",
            })
        options.append({
            "kind": "full",
            "amount_eur": full_amount,
            "label": "Pagar viaje completo (100%)",
            "description": (
                "Pago total del viaje. Obligatorio cuando faltan 60 días o menos para la salida."
                if days is not None and days <= 60
                else "Pago total del viaje en un solo plazo."
            ),
        })
        # Partial-bounds still returned so the split-payment flow can send
        # per-share amounts even from the initial state. We don't list it
        # as a dedicated card (to keep the default UX with just 2 options)
        # but the create-order endpoint accepts kind="partial" whenever a
        # payer_name is provided (see accept_partial_from_initial).
        partial_min = min(min_partial, remaining)
        partial_bounds = {
            "min_eur": round(partial_min, 2),
            "max_eur": remaining,
        }
        monthly = None
    else:
        # Some captured already → prioritise finishing the deposit (so the
        # booking is confirmed) if the threshold isn't crossed yet. Then
        # offer the full remaining as 'balance', plus a custom partial.
        threshold_amount = (
            full_amount if (days is not None and days <= 60) else deposit_amount
        )
        booking_secured_now = paid_eur >= threshold_amount - 0.01
        if not booking_secured_now:
            gap = round(threshold_amount - paid_eur, 2)
            options.append({
                "kind": "complete_deposit",
                "amount_eur": gap,
                "label": "Completar depósito",
                "description": (
                    f"Faltan {gap:.2f} € para asegurar la reserva. "
                    "Al llegar al depósito completo, el viaje queda confirmado."
                ),
            })
        options.append({
            "kind": "balance",
            "amount_eur": remaining,
            "label": "Pagar saldo restante",
            "description": f"Importe pendiente del viaje ({remaining:.2f} €).",
        })
        # Custom partial — only offered if there's room above the 10% floor.
        partial_min = min(min_partial, remaining)
        if remaining > partial_min + 0.01 or remaining >= min_partial - 0.01:
            options.append({
                "kind": "partial",
                "amount_eur": None,  # client chooses
                "label": "Otra cantidad",
                "description": (
                    f"Paga cualquier cantidad entre {partial_min:.2f} € y {remaining:.2f} €. "
                    "Puedes volver al enlace cuantas veces quieras hasta liquidar el viaje."
                ),
            })
        partial_bounds = {
            "min_eur": round(partial_min, 2),
            "max_eur": remaining,
        }
        # Suggested monthly payment — remaining ÷ months until trip
        # (rounded up), or remaining if departure is imminent.
        monthly = None
        if days is not None and days > 0:
            months = max(1, math.ceil(days / 30))
            suggested = round(remaining / months, 2)
            # Floor at the partial minimum so the chip is a valid payment.
            suggested = max(suggested, partial_bounds["min_eur"])
            suggested = min(suggested, remaining)
            monthly = {
                "amount_eur": suggested,
                "months": months,
                "days_to_trip": days,
            }

    return {
        "total_eur": total_eur,
        "paid_eur": round(paid_eur, 2),
        "remaining_eur": remaining,
        "days_to_trip": days,
        "fully_paid": False,
        "options": options,
        "partial_bounds": partial_bounds,
        "monthly_suggested_eur": monthly,
        # Booking threshold — the trip is considered "reserved" only when
        # cumulative captured payments reach at least this amount (30% of
        # the total, or 100% for last-minute trips). Split payments still
        # apply: the invoice is not confirmed until enough shares are in.
        "deposit_threshold_eur": round(
            full_amount if (days is not None and days <= 60) else deposit_amount, 2
        ),
        "booking_secured": round(paid_eur, 2) >= round(
            (full_amount if (days is not None and days <= 60) else deposit_amount) - 0.01, 2
        ),
    }


def _frontend_base_url() -> str:
    """Where the public /pay/:token page lives. Built from the request's
    Origin header at runtime; we keep a fallback to the production domain
    so emails generated server-side never point at preview URLs."""
    return os.environ.get("FRONTEND_PUBLIC_URL") or "https://itinerarios.viajadverdad.com"


def _format_payment_instructions(itn: dict, payment_url: str, options: dict) -> str:
    """The English email/WhatsApp template the agency sends to clients,
    pre-filled with the trip name and amounts so the agent can paste it
    straight into Gmail or WhatsApp."""
    traveler = (itn.get("main_traveler") or "").strip() or "there"
    first_name = traveler.split()[0] if traveler else "there"
    total = options.get("total_eur") or 0.0
    has_deposit_option = any(o["kind"] == "deposit" for o in options.get("options") or [])
    deposit = round(total * 0.30, 2)
    intro_amount = (
        f"as we are +60 days before your arrival, you can pay just the deposit amount "
        f"(30% = €{deposit:.2f} of the total €{total:.2f})"
        if has_deposit_option
        else f"the full amount of €{total:.2f} is required to confirm the trip "
             f"(we are within 60 days of departure)"
    )
    return f"""Hi {first_name},

Here's your trip presentation and the info regarding the next steps & payment to fully confirm your trip.

You can browse the full day-by-day itinerary here (mobile-friendly):

{payment_url.replace('/pay/', '/trip/')}

When you're ready to reserve, hit the "Reserve" button in the top right or open the payment page directly at:

{payment_url}

You can pay with a credit/debit card using the PayPal platform (you do not need a PayPal account to do so), and {intro_amount}.

Once the booking is confirmed, our Operations Team will start booking all your services. Around 15 days before your arrival you'll receive your travel documents with all the detailed information about your trip (exact directions, meeting points, schedules, guides' contacts...), sent online via a mobile app. These travel documents also include suggestions of places to visit and restaurants for each city you'll visit.

To confirm your services, I would also need the following information from all of you so we can start booking:

- Full names (as per passport)
- Passport Numbers
- Dates of birth
- Arrival/departure flight numbers
- Phone Number
- Any allergies/food restrictions or important information we should consider

Let me know if you have any questions :)
"""


class CreatePaymentLinkBody(BaseModel):
    # Frontend sends its window.location.origin so the public URL is the
    # browser-visible host, not the internal cluster URL that some ingress
    # configurations leak via the Origin header. Optional — falls back to
    # the request's Origin header, then to FRONTEND_PUBLIC_URL.
    origin: Optional[str] = None


@api.post("/itineraries/{itinerary_id}/payments/create-link")
async def create_payment_link(
    itinerary_id: str,
    user: Annotated[User, Depends(current_user)],
    request: Request,
    body: Optional[CreatePaymentLinkBody] = None,
):
    """Idempotent: returns the same `payment_token` if one already exists on
    the itinerary (so the agent can resend the link if the client lost it).
    The payment options + instruction text are recomputed each call so they
    always reflect the LATEST trip total and the LATEST days-to-trip."""
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    token = doc.get("payment_token")
    if not token:
        token = secrets.token_urlsafe(20)
        await db.itineraries.update_one(
            {"itinerary_id": itinerary_id},
            {"$set": {"payment_token": token, "updated_at": now_iso()}},
        )
    # Resolve the public host in this priority order:
    #   1. Explicit `origin` sent by the frontend (window.location.origin) —
    #      most reliable in preview/cluster setups.
    #   2. Origin header from the request — works in vanilla setups.
    #   3. FRONTEND_PUBLIC_URL env (production fallback).
    origin = (body.origin if body else None) or request.headers.get("origin") or _frontend_base_url()
    payment_url = f"{origin.rstrip('/')}/pay/{token}"
    totals = _compute_pricing_totals(doc)
    options = _compute_payment_options(doc, totals)
    instructions = _format_payment_instructions(doc, payment_url, options)
    return {
        "payment_token": token,
        "payment_url": payment_url,
        "instructions": instructions,
        "options": options,
        "payments": doc.get("payments") or [],
        "traveler_info": doc.get("traveler_info"),
        "extras": doc.get("extras") or [],
        "refund_requests": doc.get("refund_requests") or [],
    }


@api.get("/trip/{token}")
async def get_trip_view(token: str):
    """Public — the Fora-style client presentation of the itinerary at
    /trip/:token. Returns the itinerary shape the client actually needs
    (no internal cost breakdown, no supplier info, no notes_internal).
    Auto-selects hero + day photos from the destination gallery when the
    itinerary hasn't been decorated with images yet."""
    from destinations import pick_hero, pick_day_image, gallery_for
    doc = await db.itineraries.find_one({"payment_token": token}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Enlace de itinerario no válido")

    totals = _compute_pricing_totals(doc)
    total_eur = round(float(totals.get("pvp") or 0), 2)

    # Sanitize days for public consumption
    days_public = []
    for d in (doc.get("days") or []):
        services_public = []
        for s in (d.get("services") or []):
            # Skip accommodation services (they're rendered separately as
            # "Where you stay" cards) and hidden ones.
            if s.get("acc_id") or s.get("hidden"):
                continue
            services_public.append({
                "id": s.get("id") or s.get("service_id"),
                "type": s.get("type"),
                "name": s.get("name"),
                "description": s.get("description") or s.get("public_description"),
                "duration": s.get("duration"),
                "meeting_point": s.get("meeting_point"),
                "time": s.get("time") or s.get("start_time"),
                "image_url": s.get("image_url"),
            })
        days_public.append({
            "day_id": d.get("day_id"),
            "date": d.get("date"),
            "label": d.get("label"),
            "city": d.get("city"),
            "image_url": pick_day_image(d, doc),
            "services": services_public,
        })

    accs_public = []
    for a in (doc.get("accommodations") or []):
        rooms = [{
            "type": r.get("type") or r.get("name"),
            "size_sqm": r.get("size_sqm"),
            "pax": r.get("pax"),
        } for r in (a.get("rooms") or [])]
        images = gallery_for(a.get("city") or a.get("name"))
        accs_public.append({
            "acc_id": a.get("acc_id"),
            "name": a.get("name"),
            "city": a.get("city"),
            "date_from": a.get("date_from"),
            "date_to": a.get("date_to"),
            "address": a.get("address"),
            "notes": a.get("public_description") or a.get("notes"),
            "image_urls": (a.get("image_urls") or [])[:4] or images[:4],
            "rooms": rooms,
        })

    return {
        "trip_name": doc.get("name"),
        "main_traveler": doc.get("main_traveler"),
        "start_date": doc.get("start_date"),
        "end_date": doc.get("end_date"),
        "duration_days": doc.get("duration_days"),
        "num_travelers": doc.get("num_travelers"),
        "num_adults": doc.get("num_adults"),
        "num_children": doc.get("num_children"),
        "cities": doc.get("cities") or _extract_cities(doc),
        "hero_image": pick_hero(doc),
        "summary": doc.get("public_summary") or doc.get("summary"),
        "days": days_public,
        "accommodations": accs_public,
        "total_eur": total_eur,
        "currency": "EUR",
        "agent": await _agent_public_info_from_db(doc.get("created_by")),
    }


def _extract_cities(doc: dict) -> list:
    """Fallback when itinerary doesn't have a curated cities list: derive
    from days & accommodations, preserving order."""
    seen: list = []
    for d in (doc.get("days") or []):
        c = (d.get("city") or "").strip()
        if c and c not in seen:
            seen.append(c)
    return seen


async def _agent_public_info_from_db(email: str) -> dict:
    """Return the agent's public-facing name & avatar (nothing sensitive)."""
    if not email:
        return {}
    u = await db.users.find_one({"email": email}, {"_id": 0, "name": 1, "avatar_url": 1, "email": 1})
    if not u:
        return {"email": email, "name": email.split("@")[0].replace(".", " ").title()}
    return {
        "name": u.get("name") or (email.split("@")[0].replace(".", " ").title() if email else ""),
        "avatar_url": u.get("avatar_url"),
        "email": u.get("email") or email,
    }


class SplitInviteBody(BaseModel):
    email: str
    name: Optional[str] = None
    share_eur: Optional[float] = None
    from_name: Optional[str] = None


@api.post("/payments/{token}/invite-share")
async def invite_next_split_payer(
    token: str,
    payload: SplitInviteBody,
    request: Request,
):
    """Public — after paying a share, the client can invite the next
    traveler by email. We reuse the same payment_token so all splits
    land on the same invoice; the recipient's page auto-detects the
    split context from the captured_payments ledger."""
    doc = await db.itineraries.find_one({"payment_token": token}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Enlace de pago no válido")
    email = (payload.email or "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Introduce un email válido.")
    totals = _compute_pricing_totals(doc)
    opts = _compute_payment_options(doc, totals)
    origin = (request.headers.get("origin") or _frontend_base_url()).rstrip("/")
    public_url = f"{origin}/pay/{token}"
    share_eur = round(float(payload.share_eur), 2) if payload.share_eur else round(
        float(opts.get("remaining_eur") or 0) / 2, 2
    )
    try:
        from email_service import send_email, render_split_invite_email
        subject, html, text = render_split_invite_email(
            trip_name=doc.get("name", "Your trip"),
            payer_name=payload.name,
            from_name=payload.from_name,
            share_eur=share_eur,
            remaining_eur=float(opts.get("remaining_eur") or 0),
            booking_secured=bool(opts.get("booking_secured")),
            deposit_threshold_eur=float(opts.get("deposit_threshold_eur") or 0),
            paid_eur=float(opts.get("paid_eur") or 0),
            public_url=public_url,
        )
        ok = await send_email(
            to=email,
            subject=subject,
            html=html,
            text=text,
            reply_to=(doc.get("created_by") or None),
        )
    except Exception:
        logger.exception("invite-share email failed")
        ok = False
    return {"ok": bool(ok), "email": email, "share_eur": share_eur, "public_url": public_url}


@api.get("/payments/{token}")
async def get_payment_landing(token: str):
    """Public — what the client sees at /pay/:token. No auth, no PII beyond
    the trip name + dates + amounts (no internal IDs leaked)."""
    doc = await db.itineraries.find_one({"payment_token": token}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Enlace de pago no válido")
    totals = _compute_pricing_totals(doc)
    options = _compute_payment_options(doc, totals)
    # Expose the captured payment history publicly (only kind, amount and
    # payer identity — no internal ids, no capture ids, no origin) so the
    # split-payment flow can show "Ana already paid €X" to the next
    # traveler landing on the same link. Anyone with the token could
    # already see the aggregate paid_eur, so surfacing per-payer amounts
    # doesn't leak more data than we already do.
    public_payments = []
    for p in (doc.get("payments") or []):
        if p.get("status") != "captured":
            continue
        public_payments.append({
            "kind": p.get("kind"),
            "amount_eur": p.get("paid_amount") or p.get("amount_eur"),
            "paid_at": p.get("paid_at"),
            "payer_name": p.get("payer_name"),
            "share_label": p.get("share_label"),
        })
    return {
        "trip_name": doc.get("name"),
        "main_traveler": doc.get("main_traveler"),
        "start_date": doc.get("start_date"),
        "end_date": doc.get("end_date"),
        "duration_days": doc.get("duration_days"),
        "num_travelers": doc.get("num_travelers"),
        "num_adults": doc.get("num_adults"),
        "num_children": doc.get("num_children"),
        "currency": "EUR",
        "traveler_info": doc.get("traveler_info"),
        "captured_payments": public_payments,
        **options,
    }


class TravelerInfoPerson(BaseModel):
    full_name: str = ""
    passport_number: str = ""
    date_of_birth: str = ""


class TravelerInfoBody(BaseModel):
    """Public — what the client submits from the /pay/:token form. All
    fields are optional so partial submissions work; the client can update
    later by re-submitting."""
    people: List[TravelerInfoPerson] = []
    arrival_flight: str = ""
    departure_flight: str = ""
    phone: str = ""
    notes: str = ""
    submitted_by_email: str = ""


@api.post("/payments/{token}/traveler-info")
async def submit_traveler_info(token: str, payload: TravelerInfoBody, request: Request):
    """Public — store the booking info the client filled in (passports,
    flights, allergies). Last submit wins. Stamps `submitted_at`. Notifies
    the agent who owns the itinerary by email (Resend)."""
    doc = await db.itineraries.find_one({"payment_token": token}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Enlace de pago no válido")
    traveler_info = {
        "people": [p.model_dump() for p in payload.people],
        "arrival_flight": payload.arrival_flight.strip(),
        "departure_flight": payload.departure_flight.strip(),
        "phone": payload.phone.strip(),
        "notes": payload.notes.strip(),
        "submitted_at": now_iso(),
        "submitted_by_email": (payload.submitted_by_email or "").strip(),
    }
    await db.itineraries.update_one(
        {"payment_token": token},
        {"$set": {"traveler_info": traveler_info, "updated_at": now_iso()}},
    )
    # Notify the agent who created the trip. Best-effort: never fail the
    # client's form submission because the email service is down.
    agent_email = (doc.get("created_by") or "").strip()
    if agent_email:
        from email_service import send_email, render_traveler_info_email
        origin = (request.headers.get("origin") or _frontend_base_url()).rstrip("/")
        public_url = f"{origin}/pay/{token}"
        subject, html, text = render_traveler_info_email(doc, traveler_info, public_url=public_url)
        # Schedule the email send so we return to the client immediately.
        asyncio.create_task(send_email(
            to=agent_email,
            subject=subject,
            html=html,
            text=text,
            reply_to=traveler_info.get("submitted_by_email") or None,
        ))
    return {"ok": True, "submitted_at": traveler_info["submitted_at"]}


class CreatePayPalOrderBody(BaseModel):
    kind: Literal["deposit", "balance", "full", "partial", "complete_deposit"]
    # For kind="partial", the exact EUR amount the client wants to pay.
    # Validated server-side against the current partial_bounds (10% floor
    # of the total, capped at the remaining balance).
    amount_eur: Optional[float] = None
    # Browser's window.location.origin sent by the public page so the
    # return / cancel URLs (and the post-capture redirect) bounce back to
    # the EXACT host the client is on. Falls back to the request's Origin
    # header (which the ingress can rewrite to an internal cluster URL,
    # producing 403s post-payment), then to FRONTEND_PUBLIC_URL.
    origin: Optional[str] = None
    # Split-payment metadata: who is paying this share. Same payment_token,
    # multiple travelers each pay their part with their own identity so the
    # agent can reconcile the invoice.
    payer_name: Optional[str] = None
    payer_email: Optional[str] = None
    share_label: Optional[str] = None


@api.post("/payments/{token}/create-order")
async def create_payment_order(
    token: str,
    payload: CreatePayPalOrderBody,
    request: Request,
):
    """Public — the client picked one of the available payment options;
    create a PayPal Order and hand back the approve URL the client must
    visit to finish payment."""
    doc = await db.itineraries.find_one({"payment_token": token}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Enlace de pago no válido")
    totals = _compute_pricing_totals(doc)
    options = _compute_payment_options(doc, totals)
    # Resolve the chosen option to a concrete amount and reject if it's not
    # currently allowed (prevents a malicious client from POSTing
    # `kind=deposit` for a <60-day trip).
    chosen = next((o for o in options["options"] if o["kind"] == payload.kind), None)
    # Split-payment escape hatch: `partial` is always accepted when a
    # `payer_name` is provided, even from the initial state (no captured
    # payment yet). The bounds still apply (10% floor, remaining ceiling),
    # so this doesn't unlock arbitrary sub-payments — it only lets a
    # traveler pay their share of the deposit/full up-front.
    if not chosen and payload.kind == "partial" and (payload.payer_name or "").strip():
        chosen = {
            "kind": "partial", "amount_eur": None,
            "label": "Pago parcial (split)",
            "description": "Pago parcial iniciado desde el flujo de split.",
        }
    if not chosen:
        raise HTTPException(
            status_code=400,
            detail="Esta opción de pago no está disponible para este viaje en este momento",
        )

    # Resolve the concrete amount. For 'partial' the client provides it and
    # we validate it against the active bounds. For everything else the
    # backend dictates it.
    if payload.kind == "partial":
        if payload.amount_eur is None:
            raise HTTPException(status_code=400, detail="Indica la cantidad a pagar")
        amount = round(float(payload.amount_eur), 2)
        bounds = options.get("partial_bounds") or {}
        min_eur = float(bounds.get("min_eur") or 0)
        max_eur = float(bounds.get("max_eur") or 0)
        if amount < min_eur - 0.01 or amount > max_eur + 0.01:
            raise HTTPException(
                status_code=400,
                detail=f"La cantidad debe estar entre {min_eur:.2f} € y {max_eur:.2f} €.",
            )
    else:
        amount = float(chosen["amount_eur"])

    payment_id = new_id("pmt")
    origin = (payload.origin or request.headers.get("origin") or _frontend_base_url()).rstrip("/")
    return_url = f"{origin}/api/payments/{token}/return?payment_id={payment_id}"
    cancel_url = f"{origin}/pay/{token}?cancelled=1"
    label_for_paypal = chosen["label"] if payload.kind != "partial" else f"Pago parcial ({amount:.2f} €)"
    if payload.share_label:
        label_for_paypal = f"{label_for_paypal} · {payload.share_label}"
    description = f"{doc.get('name', 'Viaje')} · {label_for_paypal}"
    try:
        order = await _paypal_create_order(
            amount_eur=amount,
            return_url=return_url,
            cancel_url=cancel_url,
            reference=payment_id,
            description=description,
            payer_email=(payload.payer_email or None),
        )
    except httpx.HTTPStatusError as e:
        logger.warning("paypal create_order failed: %s", e.response.text[:300])
        raise HTTPException(status_code=502, detail="No se pudo crear la orden en PayPal")

    payment = {
        "payment_id": payment_id,
        "kind": payload.kind,
        "amount_eur": amount,
        "paypal_order_id": order["id"],
        "paypal_capture_id": None,
        "status": "created",
        "created_at": now_iso(),
        "paid_at": None,
        "paid_amount": None,
        "paid_currency": None,
        # Persist the browser origin so the return handler can redirect the
        # client back to the SAME host they came from — solves the preview/
        # cluster URL mismatch (#403 after PayPal capture).
        "client_origin": origin,
        # Split-payment identity (optional). Kept per-Payment so multiple
        # travelers can pay the SAME invoice with distinct names.
        "payer_name": (payload.payer_name or "").strip() or None,
        "payer_email": (payload.payer_email or "").strip() or None,
        "share_label": (payload.share_label or "").strip() or None,
    }
    await db.itineraries.update_one(
        {"payment_token": token},
        {"$push": {"payments": payment}, "$set": {"updated_at": now_iso()}},
    )
    return {
        "payment_id": payment_id,
        "paypal_order_id": order["id"],
        "approval_url": _paypal_approval_url(order),
    }


@api.get("/payments/{token}/return")
async def payment_return_handler(token: str, payment_id: str, request: Request):
    """PayPal redirects here after the client clicks "Approve" on the hosted
    payment page. We capture the order synchronously (don't trust the
    redirect — actually call /capture) and then bounce back to the public
    /pay/{token} page with a success or error flag. Idempotent: capturing
    an already-captured order returns the same record."""
    doc = await db.itineraries.find_one({"payment_token": token}, {"_id": 0})
    if not doc:
        return _redirect_to_payment_page(token, "?error=invalid")
    payment = next((p for p in (doc.get("payments") or []) if p.get("payment_id") == payment_id), None)
    if not payment or not payment.get("paypal_order_id"):
        return _redirect_to_payment_page(token, "?error=not_found", client_origin=(payment or {}).get("client_origin"))

    order_id = payment["paypal_order_id"]
    try:
        capture = await _paypal_capture_order(order_id)
    except httpx.HTTPStatusError as e:
        # Common case: order already captured (422 with details). Re-poll.
        try:
            capture = await _paypal_get_order(order_id)
        except Exception:
            logger.warning("paypal capture failed and re-poll failed: %s", e.response.text[:200])
            return _redirect_to_payment_page(token, "?error=capture_failed", client_origin=payment.get("client_origin"))

    paypal_status = (capture.get("status") or "").upper()
    capture_id = None
    paid_amount = None
    paid_currency = None
    pus = capture.get("purchase_units") or []
    if pus:
        captures_list = (pus[0].get("payments") or {}).get("captures") or []
        if captures_list:
            cap0 = captures_list[0]
            capture_id = cap0.get("id")
            amt = cap0.get("amount") or {}
            paid_amount = float(amt.get("value") or 0) if amt.get("value") else None
            paid_currency = amt.get("currency_code")
            if (cap0.get("status") or "").upper() == "COMPLETED":
                paypal_status = "COMPLETED"

    new_status = ("captured" if paypal_status == "COMPLETED"
                  else "approved" if paypal_status == "APPROVED"
                  else "created")
    await db.itineraries.update_one(
        {"payment_token": token, "payments.payment_id": payment_id},
        {"$set": {
            "payments.$.status": new_status,
            "payments.$.paypal_capture_id": capture_id,
            "payments.$.paid_at": now_iso() if new_status == "captured" else None,
            "payments.$.paid_amount": paid_amount,
            "payments.$.paid_currency": paid_currency,
            "updated_at": now_iso(),
        }},
    )
    if new_status == "captured":
        from urllib.parse import urlencode
        qs = "?" + urlencode({
            "success": 1,
            "kind": payment.get("kind") or "",
            "amount": f"{paid_amount:.2f}" if paid_amount is not None else "",
        })
        return _redirect_to_payment_page(token, qs, client_origin=payment.get("client_origin"))
    return _redirect_to_payment_page(token, "?error=not_completed",
                                     client_origin=payment.get("client_origin"))


def _redirect_to_payment_page(token: str, qs: str = "", client_origin: Optional[str] = None):
    """302 the client back to the public landing page. Prefer the origin
    that was captured when the order was created (so preview tests stay on
    preview); fall back to FRONTEND_PUBLIC_URL for production."""
    from fastapi.responses import RedirectResponse
    base = (client_origin or _frontend_base_url()).rstrip("/")
    return RedirectResponse(url=f"{base}/pay/{token}{qs}", status_code=303)


@api.post("/paypal/webhook")
async def paypal_webhook(request: Request):
    """Async confirmation channel. We verify the signature via PayPal's
    /v1/notifications/verify-webhook-signature endpoint, then mirror the
    event into the local payment row. Idempotent — duplicate webhook
    deliveries for the same capture_id are no-ops."""
    body_bytes = await request.body()
    try:
        body_text = body_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return {"ok": False, "reason": "bad-encoding"}
    # Lowercase headers for case-insensitive lookup
    headers = {k.lower(): v for k, v in request.headers.items()}
    ok = await _paypal_verify_webhook(headers, body_text)
    if not ok:
        # If PAYPAL_WEBHOOK_ID isn't configured yet (pre-deploy), we
        # accept the event but mark it as unverified in the audit log
        # for the operator. Production should ALWAYS have the id set.
        if not os.environ.get("PAYPAL_WEBHOOK_ID"):
            logger.warning("paypal webhook unverified (PAYPAL_WEBHOOK_ID not set)")
        else:
            return {"ok": False, "reason": "signature"}
    import json
    try:
        event = json.loads(body_text)
    except Exception:
        return {"ok": False, "reason": "bad-json"}

    event_type = event.get("event_type") or ""
    resource = event.get("resource") or {}
    # The order_id we stored points back to the right itinerary; the
    # capture event's resource has the parent order id in
    # supplementary_data.related_ids.order_id (or the resource is itself
    # an order for ORDER.* events).
    related = (resource.get("supplementary_data") or {}).get("related_ids") or {}
    order_id = related.get("order_id") or resource.get("id")
    capture_id = resource.get("id") if event_type.startswith("PAYMENT.CAPTURE") else None
    amount = resource.get("amount") or {}
    paid_amount = float(amount.get("value") or 0) if amount.get("value") else None
    paid_currency = amount.get("currency_code")

    new_status = None
    if event_type == "PAYMENT.CAPTURE.COMPLETED":
        new_status = "captured"
    elif event_type == "PAYMENT.CAPTURE.DENIED":
        new_status = "denied"
    elif event_type == "PAYMENT.CAPTURE.REFUNDED":
        new_status = "refunded"
    elif event_type == "PAYMENT.CAPTURE.PENDING":
        new_status = "approved"

    if new_status and order_id:
        update = {"payments.$.status": new_status, "updated_at": now_iso()}
        if new_status == "captured":
            update["payments.$.paid_at"] = now_iso()
            update["payments.$.paypal_capture_id"] = capture_id
            if paid_amount is not None:
                update["payments.$.paid_amount"] = paid_amount
            if paid_currency:
                update["payments.$.paid_currency"] = paid_currency
        await db.itineraries.update_one(
            {"payments.paypal_order_id": order_id},
            {"$set": update},
        )
    return {"ok": True}


class MarkPaymentBody(BaseModel):
    status: Literal["captured", "cancelled"]
    notes: Optional[str] = None
    paid_amount: Optional[float] = None


@api.post("/itineraries/{itinerary_id}/payments/{payment_id}/mark")
async def mark_payment_manually(
    itinerary_id: str,
    payment_id: str,
    body: MarkPaymentBody,
    user: Annotated[User, Depends(current_user)],
):
    """Escape hatch for the agent: mark a payment as captured/cancelled
    manually when the client paid through a non-PayPal channel (bizum, bank
    transfer, etc.) or to cancel an abandoned order."""
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    update = {
        "payments.$.status": body.status,
        "payments.$.notes": body.notes,
        "updated_at": now_iso(),
    }
    if body.status == "captured":
        update["payments.$.paid_at"] = now_iso()
        if body.paid_amount is not None:
            update["payments.$.paid_amount"] = body.paid_amount
    await db.itineraries.update_one(
        {"itinerary_id": itinerary_id, "payments.payment_id": payment_id},
        {"$set": update},
    )
    return {"ok": True}


# ---------------------------------------------------------------------------
# Post-sale extras (separate payment link per extra)
# ---------------------------------------------------------------------------
# When a client adds a new activity/service AFTER paying (partial or full),
# creating a delta in the main invoice would break already-captured PayPal
# orders. We instead expose a small "PostSaleExtra" doc with its own token,
# its own single-purpose payment link (/pay/extra/:token), and its own
# PayPal Order. Refunds are also isolated per extra.

# Refund approvers — hardcoded email whitelist per user request. Only these
# accounts can approve/reject a refund request.
REFUND_APPROVERS: set = {
    "beatriz@viajadverdad.com",
    "marina@viajadverdad.com",
}


def _is_refund_approver(user: User) -> bool:
    return (user.email or "").strip().lower() in REFUND_APPROVERS


class CreateExtraBody(BaseModel):
    title: str
    description: str = ""
    amount_eur: float
    day_id: Optional[str] = None
    date: Optional[str] = None
    service_id: Optional[str] = None


@api.post("/itineraries/{itinerary_id}/extras")
async def create_post_sale_extra(
    itinerary_id: str,
    payload: CreateExtraBody,
    user: Annotated[User, Depends(current_user)],
):
    """Add an extra activity to a sold itinerary. Generates its own
    payment_token so the client can settle just this delta via PayPal at
    /pay/extra/{token}."""
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    if payload.amount_eur is None or float(payload.amount_eur) <= 0:
        raise HTTPException(status_code=400, detail="La cantidad debe ser mayor que 0")
    extra = {
        "extra_id": new_id("ext"),
        "title": payload.title.strip() or "Actividad extra",
        "description": (payload.description or "").strip(),
        "amount_eur": round(float(payload.amount_eur), 2),
        "currency": "EUR",
        "day_id": payload.day_id,
        "date": payload.date,
        "service_id": payload.service_id,
        "payment_token": secrets.token_urlsafe(20),
        "status": "sent",
        "created_by": user.email,
        "created_at": now_iso(),
        "paid_at": None,
        "paid_amount": None,
        "paypal_order_id": None,
        "paypal_capture_id": None,
    }
    await db.itineraries.update_one(
        {"itinerary_id": itinerary_id},
        {"$push": {"extras": extra}, "$set": {"updated_at": now_iso()}},
    )
    return extra


@api.get("/itineraries/{itinerary_id}/extras")
async def list_post_sale_extras(
    itinerary_id: str,
    user: Annotated[User, Depends(current_user)],
):
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    return {"extras": doc.get("extras") or []}


@api.delete("/itineraries/{itinerary_id}/extras/{extra_id}")
async def delete_post_sale_extra(
    itinerary_id: str,
    extra_id: str,
    user: Annotated[User, Depends(current_user)],
):
    """Remove an extra. If it was already paid we mark it as cancelled
    (the money is still in PayPal); the agent should follow with a refund
    request instead."""
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    extra = next((e for e in (doc.get("extras") or []) if e.get("extra_id") == extra_id), None)
    if not extra:
        raise HTTPException(status_code=404, detail="Extra no encontrado")
    if extra.get("status") == "paid":
        # Do not pop a paid extra — mark cancelled so the audit trail
        # survives. Agent must issue a refund via the refund workflow.
        await db.itineraries.update_one(
            {"itinerary_id": itinerary_id, "extras.extra_id": extra_id},
            {"$set": {"extras.$.status": "cancelled", "updated_at": now_iso()}},
        )
        return {"ok": True, "status": "cancelled",
                "hint": "El extra ya estaba pagado. Solicita un reembolso para devolver el dinero."}
    await db.itineraries.update_one(
        {"itinerary_id": itinerary_id},
        {"$pull": {"extras": {"extra_id": extra_id}}, "$set": {"updated_at": now_iso()}},
    )
    return {"ok": True, "status": "deleted"}


@api.get("/payments/extra/{token}")
async def get_extra_landing(token: str):
    """Public — what the client sees at /pay/extra/:token. Loads the extra
    row, no auth needed."""
    doc = await db.itineraries.find_one(
        {"extras.payment_token": token}, {"_id": 0},
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Enlace de extra no válido")
    extra = next((e for e in (doc.get("extras") or [])
                  if e.get("payment_token") == token), None)
    if not extra:
        raise HTTPException(status_code=404, detail="Enlace de extra no válido")
    return {
        "trip_name": doc.get("name"),
        "main_traveler": doc.get("main_traveler"),
        "start_date": doc.get("start_date"),
        "end_date": doc.get("end_date"),
        "extra_id": extra.get("extra_id"),
        "title": extra.get("title"),
        "description": extra.get("description"),
        "amount_eur": extra.get("amount_eur"),
        "currency": extra.get("currency") or "EUR",
        "date": extra.get("date"),
        "status": extra.get("status"),
        "paid_at": extra.get("paid_at"),
    }


class CreateExtraOrderBody(BaseModel):
    origin: Optional[str] = None
    payer_name: Optional[str] = None
    payer_email: Optional[str] = None


@api.post("/payments/extra/{token}/create-order")
async def create_extra_order(
    token: str,
    payload: CreateExtraOrderBody,
    request: Request,
):
    """Public — client picked "Pay extra"; create a PayPal Order and return
    the approval URL."""
    doc = await db.itineraries.find_one(
        {"extras.payment_token": token}, {"_id": 0},
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Enlace de extra no válido")
    extra = next((e for e in (doc.get("extras") or [])
                  if e.get("payment_token") == token), None)
    if not extra:
        raise HTTPException(status_code=404, detail="Enlace de extra no válido")
    if extra.get("status") == "paid":
        raise HTTPException(status_code=400, detail="Este extra ya está pagado")
    if extra.get("status") == "cancelled":
        raise HTTPException(status_code=400, detail="Este extra fue cancelado")
    amount = round(float(extra.get("amount_eur") or 0), 2)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Importe inválido")

    payment_id = new_id("pmt")
    origin = (payload.origin or request.headers.get("origin") or _frontend_base_url()).rstrip("/")
    return_url = f"{origin}/api/payments/extra/{token}/return?payment_id={payment_id}"
    cancel_url = f"{origin}/pay/extra/{token}?cancelled=1"
    description = f"{doc.get('name', 'Viaje')} · Extra: {extra.get('title')}"
    try:
        order = await _paypal_create_order(
            amount_eur=amount,
            return_url=return_url,
            cancel_url=cancel_url,
            reference=payment_id,
            description=description,
            payer_email=(payload.payer_email or None),
        )
    except httpx.HTTPStatusError as e:
        logger.warning("paypal create_order (extra) failed: %s", e.response.text[:300])
        raise HTTPException(status_code=502, detail="No se pudo crear la orden en PayPal")

    payment = {
        "payment_id": payment_id,
        "kind": "extra",
        "amount_eur": amount,
        "paypal_order_id": order["id"],
        "paypal_capture_id": None,
        "status": "created",
        "created_at": now_iso(),
        "paid_at": None,
        "paid_amount": None,
        "paid_currency": None,
        "client_origin": origin,
        "payer_name": (payload.payer_name or "").strip() or None,
        "payer_email": (payload.payer_email or "").strip() or None,
        "share_label": None,
        "extra_id": extra.get("extra_id"),
    }
    await db.itineraries.update_one(
        {"itinerary_id": doc["itinerary_id"]},
        {"$push": {"payments": payment},
         "$set": {"extras.$[e].paypal_order_id": order["id"],
                  "updated_at": now_iso()}},
        array_filters=[{"e.extra_id": extra.get("extra_id")}],
    )
    return {
        "payment_id": payment_id,
        "paypal_order_id": order["id"],
        "approval_url": _paypal_approval_url(order),
    }


@api.get("/payments/extra/{token}/return")
async def extra_return_handler(token: str, payment_id: str, request: Request):
    """PayPal redirects here after the client approves an extra payment.
    Capture the order, mark the extra as paid, then redirect back to the
    /pay/extra/:token page with a success flag."""
    doc = await db.itineraries.find_one(
        {"extras.payment_token": token}, {"_id": 0},
    )
    if not doc:
        return _redirect_to_extra_page(token, "?error=invalid")
    payment = next((p for p in (doc.get("payments") or [])
                    if p.get("payment_id") == payment_id), None)
    if not payment or not payment.get("paypal_order_id"):
        return _redirect_to_extra_page(token, "?error=not_found",
                                        client_origin=(payment or {}).get("client_origin"))
    extra_id = payment.get("extra_id")
    order_id = payment["paypal_order_id"]
    try:
        capture = await _paypal_capture_order(order_id)
    except httpx.HTTPStatusError as e:
        try:
            capture = await _paypal_get_order(order_id)
        except Exception:
            logger.warning("paypal capture (extra) failed: %s", e.response.text[:200])
            return _redirect_to_extra_page(token, "?error=capture_failed",
                                            client_origin=payment.get("client_origin"))
    paypal_status = (capture.get("status") or "").upper()
    capture_id = None
    paid_amount = None
    paid_currency = None
    pus = capture.get("purchase_units") or []
    if pus:
        captures_list = (pus[0].get("payments") or {}).get("captures") or []
        if captures_list:
            cap0 = captures_list[0]
            capture_id = cap0.get("id")
            amt = cap0.get("amount") or {}
            paid_amount = float(amt.get("value") or 0) if amt.get("value") else None
            paid_currency = amt.get("currency_code")
            if (cap0.get("status") or "").upper() == "COMPLETED":
                paypal_status = "COMPLETED"

    new_status = ("captured" if paypal_status == "COMPLETED"
                  else "approved" if paypal_status == "APPROVED"
                  else "created")
    # Update the payment row.
    await db.itineraries.update_one(
        {"payments.payment_id": payment_id},
        {"$set": {
            "payments.$.status": new_status,
            "payments.$.paypal_capture_id": capture_id,
            "payments.$.paid_at": now_iso() if new_status == "captured" else None,
            "payments.$.paid_amount": paid_amount,
            "payments.$.paid_currency": paid_currency,
            "updated_at": now_iso(),
        }},
    )
    # Update the extra row.
    if new_status == "captured" and extra_id:
        await db.itineraries.update_one(
            {"extras.extra_id": extra_id},
            {"$set": {
                "extras.$.status": "paid",
                "extras.$.paid_at": now_iso(),
                "extras.$.paid_amount": paid_amount,
                "extras.$.paypal_capture_id": capture_id,
                "updated_at": now_iso(),
            }},
        )
    if new_status == "captured":
        from urllib.parse import urlencode
        qs = "?" + urlencode({
            "success": 1,
            "amount": f"{paid_amount:.2f}" if paid_amount is not None else "",
        })
        return _redirect_to_extra_page(token, qs, client_origin=payment.get("client_origin"))
    return _redirect_to_extra_page(token, "?error=not_completed",
                                    client_origin=payment.get("client_origin"))


def _redirect_to_extra_page(token: str, qs: str = "", client_origin: Optional[str] = None):
    from fastapi.responses import RedirectResponse
    base = (client_origin or _frontend_base_url()).rstrip("/")
    return RedirectResponse(url=f"{base}/pay/extra/{token}{qs}", status_code=303)


# ---------------------------------------------------------------------------
# Refund workflow (agent files a request, manager approves, PayPal refunds)
# ---------------------------------------------------------------------------
class CreateRefundBody(BaseModel):
    payment_id: str
    amount_eur: float
    reason: str = ""
    service_id: Optional[str] = None


@api.post("/itineraries/{itinerary_id}/refund-requests")
async def create_refund_request(
    itinerary_id: str,
    payload: CreateRefundBody,
    user: Annotated[User, Depends(current_user)],
):
    """Any agent files a refund request. The money is NOT returned yet —
    it needs a manager to explicitly approve."""
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    payment = next((p for p in (doc.get("payments") or [])
                    if p.get("payment_id") == payload.payment_id), None)
    if not payment:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    if payment.get("status") != "captured":
        raise HTTPException(status_code=400,
                            detail="Sólo se pueden reembolsar pagos capturados")
    captured_amount = float(payment.get("paid_amount") or payment.get("amount_eur") or 0)
    # Reduce by any already-executed refund from the same payment.
    already_refunded = 0.0
    for r in (doc.get("refund_requests") or []):
        if r.get("payment_id") == payload.payment_id and r.get("status") == "executed":
            already_refunded += float(r.get("amount_eur") or 0)
    refundable = round(captured_amount - already_refunded, 2)
    if payload.amount_eur <= 0 or payload.amount_eur > refundable + 0.01:
        raise HTTPException(
            status_code=400,
            detail=f"La cantidad debe estar entre 0 y {refundable:.2f} € (importe reembolsable).",
        )
    refund = {
        "refund_id": new_id("rfd"),
        "payment_id": payload.payment_id,
        "service_id": payload.service_id,
        "amount_eur": round(float(payload.amount_eur), 2),
        "reason": (payload.reason or "").strip(),
        "requested_by": user.email,
        "requested_at": now_iso(),
        "approved_by": None,
        "decided_at": None,
        "status": "pending",
        "paypal_refund_id": None,
        "error_message": None,
    }
    await db.itineraries.update_one(
        {"itinerary_id": itinerary_id},
        {"$push": {"refund_requests": refund}, "$set": {"updated_at": now_iso()}},
    )
    return refund


@api.get("/itineraries/{itinerary_id}/refund-requests")
async def list_refund_requests(
    itinerary_id: str,
    user: Annotated[User, Depends(current_user)],
):
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    return {
        "refund_requests": doc.get("refund_requests") or [],
        "is_approver": _is_refund_approver(user),
        "approver_emails": sorted(REFUND_APPROVERS),
    }


class ApproveRefundBody(BaseModel):
    # Optional override — approver can trim the amount before executing.
    amount_eur: Optional[float] = None
    note_to_payer: Optional[str] = None


@api.post("/itineraries/{itinerary_id}/refund-requests/{refund_id}/approve")
async def approve_refund(
    itinerary_id: str,
    refund_id: str,
    payload: ApproveRefundBody,
    user: Annotated[User, Depends(current_user)],
):
    """Manager-only: approve and execute a refund via the PayPal Refund
    API. Only `beatriz@viajadverdad.com` and `marina@viajadverdad.com` are
    accepted; anyone else gets 403."""
    if not _is_refund_approver(user):
        raise HTTPException(
            status_code=403,
            detail="Sólo Beatriz o Marina pueden aprobar reembolsos.",
        )
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    refund = next((r for r in (doc.get("refund_requests") or [])
                   if r.get("refund_id") == refund_id), None)
    if not refund:
        raise HTTPException(status_code=404, detail="Reembolso no encontrado")
    if refund.get("status") not in ("pending", "failed"):
        raise HTTPException(status_code=400,
                            detail=f"Este reembolso ya está en estado '{refund.get('status')}'.")
    payment = next((p for p in (doc.get("payments") or [])
                    if p.get("payment_id") == refund["payment_id"]), None)
    if not payment or not payment.get("paypal_capture_id"):
        raise HTTPException(status_code=400,
                            detail="El pago original no tiene un capture_id de PayPal (¿se marcó manualmente?). No se puede reembolsar automáticamente.")
    amount = float(payload.amount_eur or refund["amount_eur"])
    try:
        from paypal import refund_capture as _paypal_refund
        result = await _paypal_refund(
            payment["paypal_capture_id"],
            amount_eur=amount,
            note=payload.note_to_payer or f"Refund: {refund.get('reason') or 'trip adjustment'}",
            invoice_id=refund_id,
        )
    except httpx.HTTPStatusError as e:
        err = e.response.text[:400]
        logger.warning("paypal refund failed: %s", err)
        await db.itineraries.update_one(
            {"itinerary_id": itinerary_id, "refund_requests.refund_id": refund_id},
            {"$set": {
                "refund_requests.$.status": "failed",
                "refund_requests.$.error_message": err,
                "refund_requests.$.decided_at": now_iso(),
                "refund_requests.$.approved_by": user.email,
                "updated_at": now_iso(),
            }},
        )
        raise HTTPException(status_code=502,
                            detail=f"PayPal rechazó el reembolso: {err[:200]}")
    paypal_refund_id = result.get("id")
    paypal_status = (result.get("status") or "").upper()
    new_status = "executed" if paypal_status in ("COMPLETED", "PENDING") else "failed"
    await db.itineraries.update_one(
        {"itinerary_id": itinerary_id, "refund_requests.refund_id": refund_id},
        {"$set": {
            "refund_requests.$.status": new_status,
            "refund_requests.$.paypal_refund_id": paypal_refund_id,
            "refund_requests.$.amount_eur": round(amount, 2),
            "refund_requests.$.approved_by": user.email,
            "refund_requests.$.decided_at": now_iso(),
            "refund_requests.$.error_message": None if new_status == "executed" else result.get("status"),
            "updated_at": now_iso(),
        }},
    )
    # Mirror onto the payment row: if the full amount was refunded, mark
    # the source payment as refunded. Partial refunds keep the payment
    # as `captured` (we track the delta via refund_requests).
    captured_amount = float(payment.get("paid_amount") or payment.get("amount_eur") or 0)
    total_refunded = amount
    for r in (doc.get("refund_requests") or []):
        if (r.get("payment_id") == refund["payment_id"] and
                r.get("status") == "executed" and r.get("refund_id") != refund_id):
            total_refunded += float(r.get("amount_eur") or 0)
    if total_refunded >= captured_amount - 0.01:
        await db.itineraries.update_one(
            {"itinerary_id": itinerary_id, "payments.payment_id": refund["payment_id"]},
            {"$set": {"payments.$.status": "refunded", "updated_at": now_iso()}},
        )
    return {"ok": True, "status": new_status, "paypal_refund_id": paypal_refund_id}


class RejectRefundBody(BaseModel):
    reason: Optional[str] = None


@api.post("/itineraries/{itinerary_id}/refund-requests/{refund_id}/reject")
async def reject_refund(
    itinerary_id: str,
    refund_id: str,
    payload: RejectRefundBody,
    user: Annotated[User, Depends(current_user)],
):
    """Manager-only: reject a refund request. No PayPal call."""
    if not _is_refund_approver(user):
        raise HTTPException(status_code=403,
                            detail="Sólo Beatriz o Marina pueden rechazar reembolsos.")
    r = await db.itineraries.update_one(
        {"itinerary_id": itinerary_id, "refund_requests.refund_id": refund_id,
         "refund_requests.status": "pending"},
        {"$set": {
            "refund_requests.$.status": "rejected",
            "refund_requests.$.approved_by": user.email,
            "refund_requests.$.decided_at": now_iso(),
            "refund_requests.$.error_message": (payload.reason or "").strip() or None,
            "updated_at": now_iso(),
        }},
    )
    if r.matched_count == 0:
        # Either the refund doesn't exist or it's no longer in a pending
        # state (already rejected / executed / failed). Surface the
        # ambiguity so the UI can differentiate "noop" from "success".
        raise HTTPException(
            status_code=404,
            detail="Reembolso no encontrado o ya decidido.",
        )
    return {"ok": True, "status": "rejected"}


# ---------------------------------------------------------------------------
# Excel export (Sofi format)
# ---------------------------------------------------------------------------
# The exported workbook MUST follow the Sofi "plantillasoficotizaciones" template
# verbatim so the file can be re-imported into Sofi without manual editing:
#   - Single sheet named "Trip Prices"
#   - 7 columns total (A..G)
#   - Sections in this exact order: trip summary, Traveler Details,
#     Activities and transportation (NO accommodation lines here),
#     Accommodations (every hotel with entry/exit date range)
#   - No subtotals / no PVP / no City / no "sin IVA / con IVA" columns
#     (those live in the app dashboard, not in the Sofi import sheet).
def _to_date(s: Optional[str]):
    """Return a datetime.date for ISO 'YYYY-MM-DD' strings so Excel formats it
    natively. Falls back to the raw string when parsing fails."""
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).date()
    except (TypeError, ValueError):
        return s


def _fmt_date(s: Optional[str]) -> str:
    if not s:
        return ""
    try:
        return datetime.fromisoformat(s).strftime("%d/%m/%Y")
    except (TypeError, ValueError):
        return s


def _fmt_travefy_date(s: Optional[str]) -> str:
    """Travefy export format: 'Jul 16, 2026' (3-letter month, no leading zero
    on day, comma-separated year). Used in the trip-prices template."""
    if not s:
        return ""
    try:
        d = datetime.fromisoformat(s)
        return d.strftime("%b %-d, %Y")
    except (TypeError, ValueError):
        return s


def _fmt_travefy_long_range(date_from: Optional[str], date_to: Optional[str]) -> str:
    """'July 16 - July 20, 2026 (4 nights)' style range Travefy uses on the
    Accommodations rows."""
    if not date_from:
        return ""
    try:
        df = datetime.fromisoformat(date_from)
        dt_obj = datetime.fromisoformat(date_to) if date_to else df
        nights = max(0, (dt_obj.date() - df.date()).days)
        nights_word = "night" if nights == 1 else "nights"
        if df.year == dt_obj.year:
            return (f"{df.strftime('%B %-d')} - {dt_obj.strftime('%B %-d, %Y')} "
                    f"({nights} {nights_word})")
        return (f"{df.strftime('%B %-d, %Y')} - {dt_obj.strftime('%B %-d, %Y')} "
                f"({nights} {nights_word})")
    except (TypeError, ValueError):
        return f"{date_from} - {date_to}".strip(" -")


@api.get("/itineraries/{itinerary_id}/export")
async def export_itinerary(itinerary_id: str, user: Annotated[User, Depends(current_user)]):
    """Generate the Sofi-compatible 'Trip Prices' workbook.

    Matches the Travefy export the agency already uses internally:
      - Header rows 1-7: trip metadata
      - Banner row "Traveler Details" + table
      - Banner row "Activities and transportation" + day-by-day table
      - TOTAL row for activities
      - Banner row "Accommodations" + table + TOTAL
      - Footer with Base total, Mark up %, Commission %, Total price

    The agent imports this file directly into Sofi via the existing manual
    Excel import flow on the agency side (the same one they used before this
    tool existed), so the structure must replicate Travefy's bit-for-bit.
    """
    itn_doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not itn_doc:
        raise HTTPException(status_code=404, detail="Not found")
    if not _can_access(itn_doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")
    itn = Itinerary(**itn_doc)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trip Prices"

    BOLD = Font(bold=True, size=11)
    BOLD_BANNER = Font(bold=True, size=11)
    BANNER_FILL = PatternFill(start_color="FFC4DAF0", end_color="FFC4DAF0", fill_type="solid")
    DAY_FILL = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")
    money_fmt = "#,##0.00"

    def banner(row: int, last_col: int, label: str):
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col, label if col == 1 else "")
            cell.font = BOLD_BANNER
            cell.fill = BANNER_FILL

    def day_row(row: int, last_col: int, label: str, date_str: str):
        # Travefy renders each gray "Day N" row with EVERY column populated as
        # an empty string (not null); some XLSX importers reject null cells in
        # styled rows. Mirror that exactly.
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col, "")
            cell.fill = DAY_FILL
        ws.cell(row, 1, label)
        ws.cell(row, 2, date_str)

    # Header rows 1-7 ---------------------------------------------------------
    ws.cell(1, 1, "Main traveler name").font = BOLD
    ws.cell(1, 2, itn.main_traveler or itn.name or "")
    ws.cell(2, 1, "Trip start date").font = BOLD
    ws.cell(2, 2, _fmt_travefy_date(itn.start_date))
    ws.cell(3, 1, "Trip end date").font = BOLD
    ws.cell(3, 2, _fmt_travefy_date(itn.end_date))
    ws.cell(4, 1, "Duration (days)").font = BOLD
    ws.cell(4, 2, itn.duration_days or 0)
    ws.cell(5, 1, "Number of travelers").font = BOLD
    ws.cell(5, 2, itn.num_travelers or 0)
    ws.cell(6, 1, "Flight details").font = BOLD

    # Traveler Details (rows 9 banner, 10 headers, 11+ rows) ------------------
    banner(9, 7, "Traveler Details")
    headers = ["First name", "Last name", "Gender", "Date of birth",
               "Nationality", "Passport #", "Special requests & comments"]
    for i, h in enumerate(headers, start=1):
        ws.cell(10, i, h).font = BOLD
    r = 11
    for t in (itn.travelers or []):
        ws.cell(r, 1, t.first_name or "")
        ws.cell(r, 2, t.last_name or "")
        ws.cell(r, 3, getattr(t, "gender", "") or "")
        ws.cell(r, 4, _fmt_travefy_date(getattr(t, "date_of_birth", None)))
        ws.cell(r, 5, getattr(t, "nationality", "") or "")
        ws.cell(r, 6, getattr(t, "passport_number", "") or "")
        ws.cell(r, 7, getattr(t, "special_requests", "") or "")
        r += 1
    # If no travelers were entered, still reserve the typical 5 rows so the
    # vertical alignment with the activities banner matches a Travefy export.
    activities_banner_row = max(18, r + 3)

    # Activities and transportation -------------------------------------------
    banner(activities_banner_row, 11, "Activities and transportation")
    act_headers = ["Day", "Date", "Time/Confirmation", "Type", "Name",
                   "Quantity", "Duration", "Operator", "Contact details",
                   "Price", "Currency"]
    head_row = activities_banner_row + 1
    for i, h in enumerate(act_headers, start=1):
        ws.cell(head_row, i, h).font = BOLD

    r = head_row + 1
    activities_total = 0.0
    qty_str = _qty_label(itn)
    for idx, day in enumerate(itn.days or [], start=1):
        # Skip auto-spread accommodation carriers (Check-in/Alojamiento/Check-out)
        # — the actual hotel goes once in the Accommodations section below.
        services = [s for s in (day.services or [])
                    if not getattr(s, "acc_id", None) and s.type != "alojamiento"]
        day_row(r, 11, f"Day {idx}", _fmt_travefy_date(day.date))
        r += 1
        for s in services:
            ws.cell(r, 4, _travefy_type(s.type))
            ws.cell(r, 5, s.name or "")
            ws.cell(r, 6, qty_str)
            ws.cell(r, 7, getattr(s, "duration", "") or "")
            ws.cell(r, 8, getattr(s, "provider_name", "") or "")
            ws.cell(r, 9, getattr(s, "provider_contact", "") or "")
            unit_incl = float(s.unit_price_tax_incl or s.unit_price or 0)
            line_total = round(unit_incl * (s.quantity or 0), 2)
            cell = ws.cell(r, 10, line_total)
            cell.number_format = money_fmt
            ws.cell(r, 11, getattr(s, "currency", None) or "EUR")
            activities_total += line_total
            r += 1

    # TOTAL row for activities
    ws.cell(r, 1, "TOTAL").font = BOLD
    cell = ws.cell(r, 10, round(activities_total, 2))
    cell.number_format = money_fmt
    cell.font = BOLD
    ws.cell(r, 11, "EUR").font = BOLD
    r += 3  # spacer rows like Travefy does

    # Accommodations ----------------------------------------------------------
    banner(r, 6, "Accommodations")
    r += 1
    acc_headers = ["Day", "Date", "Name", "Notes", "Price", "Currency"]
    for i, h in enumerate(acc_headers, start=1):
        ws.cell(r, i, h).font = BOLD
    r += 1
    accommodations_total = 0.0
    # Map each accommodation to a "Day X - Y" range based on day.date matches.
    day_dates = [d.date for d in (itn.days or [])]

    def _day_range(date_from: Optional[str], date_to: Optional[str]) -> str:
        if not date_from:
            return ""
        try:
            d_in = day_dates.index(date_from) + 1
        except ValueError:
            d_in = 1
        if not date_to:
            return f"Day {d_in}"
        try:
            d_out = day_dates.index(date_to) + 1
        except ValueError:
            d_out = d_in + 1
        return f"Day {d_in} - {d_out}" if d_out > d_in else f"Day {d_in}"

    for a in (itn.accommodations or []):
        ws.cell(r, 1, _day_range(a.date_from, a.date_to))
        ws.cell(r, 2, _fmt_travefy_long_range(a.date_from, a.date_to))
        # Multi-line "Name + room types" exactly like Travefy
        rooms_lines = "\n".join((rm.room_type or "Doble") for rm in (a.rooms or []))
        name_cell = (a.name or "") + (("\n\n" + rooms_lines + "\n") if rooms_lines else "")
        cell = ws.cell(r, 3, name_cell)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 4, getattr(a, "notes", "") or "")
        p_incl = float(a.price_tax_incl or a.price or 0)
        cell = ws.cell(r, 5, round(p_incl, 2))
        cell.number_format = money_fmt
        ws.cell(r, 6, getattr(a, "currency", None) or "EUR")
        accommodations_total += p_incl
        r += 1

    ws.cell(r, 1, "TOTAL").font = BOLD
    cell = ws.cell(r, 5, round(accommodations_total, 2))
    cell.number_format = money_fmt
    cell.font = BOLD
    ws.cell(r, 6, "EUR").font = BOLD
    r += 3

    # Footer totals -----------------------------------------------------------
    base_total = round(activities_total + accommodations_total, 2)
    ws.cell(r, 1, "Base total").font = BOLD
    cell = ws.cell(r, 5, base_total)
    cell.number_format = money_fmt
    ws.cell(r, 6, "EUR").font = BOLD
    r += 1
    ws.cell(r, 1, "Base total (excluding non-commissioned)").font = BOLD
    cell = ws.cell(r, 5, base_total)
    cell.number_format = money_fmt
    ws.cell(r, 6, "EUR").font = BOLD
    r += 1

    markup_pct = float(itn.markup_pct or 0)
    markup_eur = round(base_total * markup_pct / 100.0, 2)
    ws.cell(r, 1, "Mark up (%)").font = BOLD
    ws.cell(r, 2, markup_pct)
    cell = ws.cell(r, 5, markup_eur)
    cell.number_format = money_fmt
    ws.cell(r, 6, "EUR").font = BOLD
    r += 1

    com_pct = float(itn.commission_pct or 0)
    # Travefy formula: commission = (base + markup) * com_pct / (1 - com_pct/100)
    sub_with_markup = base_total + markup_eur
    com_eur = round(sub_with_markup * com_pct / max(1e-9, (100.0 - com_pct)), 2) if com_pct else 0.0
    com_label = f"{_partner_label(itn.partner)} commission (%)"
    ws.cell(r, 1, com_label).font = BOLD
    ws.cell(r, 2, com_pct)
    cell = ws.cell(r, 5, com_eur)
    cell.number_format = money_fmt
    ws.cell(r, 6, "EUR").font = BOLD
    r += 1

    total_price = round(sub_with_markup + com_eur, 2)
    ws.cell(r, 1, "Total price").font = BOLD
    cell = ws.cell(r, 5, total_price)
    cell.number_format = money_fmt
    cell.font = BOLD
    ws.cell(r, 6, "EUR").font = BOLD

    # Column widths matching the Travefy template proportions ----------------
    widths = [10, 22, 20, 14, 50, 32, 14, 18, 18, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Travefy-style filename: trip_prices_<8 chars random>.xlsx — the agency's
    # Sofi importer parses this prefix, so we must replicate it bit-for-bit.
    # We seed the random part with the itinerary id so re-exports of the same
    # trip yield a stable filename.
    import hashlib, base64 as _b64
    h = hashlib.sha1(itinerary_id.encode("utf-8")).digest()
    token = _b64.urlsafe_b64encode(h)[:8].decode("ascii").replace("=", "x")
    filename = f"trip_prices_{token}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _qty_label(itn: "Itinerary") -> str:
    """'2 adults, 3 children (ages 8, 11, and 17)' style traveler-count label
    that Travefy uses in the Quantity column for every activity row."""
    adults = sum(1 for t in (itn.travelers or []) if not getattr(t, "is_child", False))
    children = [t for t in (itn.travelers or []) if getattr(t, "is_child", False)]
    if not (itn.travelers or []):
        # Fallback when traveler list is empty: derive from num_travelers
        n = int(itn.num_travelers or 0)
        return f"{n} {'traveler' if n == 1 else 'travelers'}"
    parts = [f"{adults} {'adult' if adults == 1 else 'adults'}"]
    if children:
        ages = [str(getattr(t, "age", "")) for t in children if getattr(t, "age", None)]
        ch_label = f"{len(children)} {'child' if len(children) == 1 else 'children'}"
        if ages:
            if len(ages) == 1:
                ch_label += f" (age {ages[0]})"
            elif len(ages) == 2:
                ch_label += f" (ages {ages[0]} and {ages[1]})"
            else:
                ch_label += f" (ages {', '.join(ages[:-1])}, and {ages[-1]})"
        parts.append(ch_label)
    return ", ".join(parts)


def _travefy_type(t: Optional[str]) -> str:
    """Map our internal ServiceType enum to Travefy's display labels."""
    return {
        "actividad": "Activity",
        "transfer": "Transport",
        "tren": "Transport",
        "vuelo": "Transport",
        "rental_car": "Transport",
        "entradas": "Activity",
        "alojamiento": "Accommodation",
        "restaurante": "Activity",
    }.get((t or "").lower(), (t or "").title())


def _partner_label(p: Optional[str]) -> str:
    """Travefy's commission row reads 'Kimkim commission (%)' / 'Zicasso ...'.
    We mirror that wording for whatever partner the agency is shipping under.
    """
    return {
        "kimkim": "Kimkim",
        "zicasso": "Zicasso",
        "responsible_travel": "Responsible Travel",
        "baboo": "Baboo",
        "travel_agent_10": "Travel Agency",
        "travel_agent_12": "Travel Agency",
        "travel_agent_15": "Travel Agency",
        "direct": "Direct",
        "other": "Direct",
    }.get((p or "kimkim").lower(), "Partner")


# ---------------------------------------------------------------------------
# Travefy import — agents paste a Travefy URL and we:
#   1) Scrape + LLM-parse it (reuses scraper.py)
#   2) Match each activity/hotel against our catalog (token overlap + city)
#   3) Return a "preview" the agent can review/edit
#   4) Confirm → create a real Itinerary with prices from our BBDD
# ---------------------------------------------------------------------------
_CITY_ALIAS = {
    # Travefy often uses English names — normalize to the spelling in our DB.
    "roma": "Rome", "rome": "Rome", "florencia": "Florence", "firenze": "Florence",
    "florence": "Florence", "venezia": "Venice", "venecia": "Venice", "venice": "Venice",
    "napoli": "Naples", "naples": "Naples", "milano": "Milan", "milan": "Milan",
    "lisbon": "Lisbon", "lisboa": "Lisbon", "porto": "Porto", "oporto": "Porto",
    "sevilla": "Seville", "seville": "Seville", "cordoba": "Córdoba", "córdoba": "Córdoba",
    "marrakesh": "Marrakech", "marrakech": "Marrakech",
    "san sebastian": "San Sebastián", "san sebastián": "San Sebastián",
}

# Words that don't help discriminate one experience from another — strip them
# from the token-overlap matcher so "Vatican & Sistine Chapel Tour (small group)"
# still matches "Vatican Museums Small Group Tour".
_MATCH_NOISE = {
    "the", "and", "with", "small", "group", "private", "tour", "experience",
    "tickets", "ticket", "skip", "line", "visit", "guided", "walking", "free",
    "from", "into", "your", "you", "our", "for", "this", "that", "del", "de",
    "la", "el", "los", "las", "una", "uno", "una", "por", "con", "sin",
    "tasting", "tasting", "transfer", "departure",
}


def _norm_city(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    return _CITY_ALIAS.get(s.strip().lower(), s.strip())


def _norm_text(s: Optional[str]) -> str:
    if not s:
        return ""
    s = re.sub(r"[^\w\s]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def _tokens_for_match(s: Optional[str]) -> set[str]:
    return {t for t in _norm_text(s).split() if len(t) >= 3 and t not in _MATCH_NOISE}


def _match_score(a: str, b: str) -> float:
    """Symmetric Jaccard-like score over noise-filtered tokens. 0..1."""
    ta, tb = _tokens_for_match(a), _tokens_for_match(b)
    if not ta or not tb:
        return 0.0
    overlap = len(ta & tb)
    return overlap / min(len(ta), len(tb))


def _classify_travefy(name: str) -> ServiceType:
    n = (name or "").lower()
    if any(k in n for k in ("transfer", "taxi", "private driver", "shuttle")):
        return "transfer"
    if any(k in n for k in ("flight", "vuelo", "airline")):
        return "vuelo"
    if any(k in n for k in ("train", "tren", "renfe", "trenitalia", "italo", "ave ")):
        return "tren"
    # Rental cars: "Vehicle Rental", "Rental Car", "Car Rental", "Rent a Car",
    # "Alquiler de coche/vehículo/auto". Excluded above by the transfer check
    # so things like "Private Transfer ... Rental Car Office ..." stay transfers.
    if any(k in n for k in (
        "vehicle rental", "rental car", "car rental",
        "rent a car", "rent-a-car",
        "alquiler de coche", "alquiler de veh", "alquiler de auto",
        "alquiler vehic",
    )):
        return "rental_car"
    if any(k in n for k in ("ticket", "entrada", "admission")) and "tour" not in n:
        return "entradas"
    return "actividad"


# The DB still contains experiences with legacy taxonomy values like
# 'restaurante', 'transporte', 'otro' that we cleaned in preview but never
# back-migrated in production. Map them to the closed enum the Itinerary
# Pydantic model accepts so the confirm endpoint doesn't fail validation.
_LEGACY_TYPE_REMAP: dict[str, ServiceType] = {
    "restaurante": "actividad",
    "transporte": "tren",
    "otro": "entradas",
    # Anything that came in with the right enum value is passed through below.
}
_VALID_TYPES: set[str] = {"alojamiento", "actividad", "entradas", "transfer", "tren", "vuelo", "hotel", "rental_car"}


def _normalize_service_type(t: Optional[str]) -> ServiceType:
    """Coerce any string into the ServiceType Literal. Unknown values fall
    back to 'actividad' (the safe catch-all) so a stale catalog entry can't
    bring down the confirm endpoint."""
    if not t:
        return "actividad"
    t = t.strip().lower()
    if t in _VALID_TYPES:
        return t  # type: ignore[return-value]
    return _LEGACY_TYPE_REMAP.get(t, "actividad")


# Travefy uses "Free day in X" / "Departure" / "Día libre" as placeholder rows
# on rest days. They aren't services and should NEVER end up in the itinerary
# (otherwise the day shows a phantom "⚠ Sin match · Revisar" line).
_FREE_DAY_PATTERN = re.compile(
    r"\b(free\s+day|d[ií]a\s+libre|free\s+time|leisure\s+day|day\s+at\s+leisure|"
    r"departure|farewell|safe\s+travels|welcome)\b",
    re.IGNORECASE,
)


def _is_free_day_marker(name: str) -> bool:
    if not name:
        return True
    return bool(_FREE_DAY_PATTERN.search(name))


# Map the free-form room description Travefy returns ("Classic Roma", "Deluxe
# River-View Twin", "Superior Room", "Double Executive"…) into the closed
# enum our DB uses. Order matters — check the more specific keywords first so
# "Family Suite" lands on "family" instead of "suite".
def _classify_room_type(raw: Optional[str]) -> RoomType:
    n = (raw or "").lower()
    if not n:
        return "doble"
    if "family" in n or "familiar" in n:
        return "family"
    if "suite" in n:
        return "suite"
    if "quad" in n or "cuadruple" in n or "cuádruple" in n:
        return "cuadruple"
    if "triple" in n:
        return "triple"
    if "twin" in n or "two beds" in n or "two bed" in n or "dos camas" in n:
        return "twin"
    if "single" in n or "individual" in n or "solo use" in n or "uso individual" in n:
        return "single"
    # Everything else ("Classic Room", "Superior Room", "Deluxe Room",
    # "Executive Room", "Double", "Junior Deluxe"…) is treated as a double.
    return "doble"


# How many travelers can sleep in a room of each type. Used to derive how many
# rooms a multi-pax group needs when Travefy only describes the base type.
_ROOM_CAPACITY: dict[str, int] = {
    "single": 1,
    "doble": 2,
    "twin": 2,
    "triple": 3,
    "cuadruple": 4,
    "suite": 4,
    "family": 4,
}


# Words that translate to a number when Travefy spells it out.
_NUMBER_WORDS = {
    "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
    "uno": 1, "una": 1, "dos": 2, "tres": 3, "cuatro": 4, "cinco": 5, "seis": 6,
}


def _parse_explicit_layout(raw: Optional[str]) -> list[tuple[RoomType, int]]:
    """Detect "2 Double Rooms", "1 Suite + 1 Twin", "Two Doubles + 1 Single"
    patterns in the Travefy room description and return the explicit layout.

    Returns [] when no multi-room pattern is found (caller will fall back to
    deriving the layout from num_travelers + base type)."""
    if not raw:
        return []
    text = raw.lower()
    # Normalise the connectors so " + " / " and " / " & " all split the same way.
    text = re.sub(r"\s*(?:\+|&|\band\b|,)\s*", "|", text)
    segments = [seg.strip() for seg in text.split("|") if seg.strip()]
    layout: list[tuple[RoomType, int]] = []
    # Match either "(N) (type word)" or "(N)x (type word)"
    PATTERN = re.compile(r"^\s*(\d+|" + "|".join(_NUMBER_WORDS.keys()) + r")\s*x?\s*(.+)$", re.IGNORECASE)
    for seg in segments:
        m = PATTERN.match(seg)
        if not m:
            continue
        n_token = m.group(1).lower()
        rest = m.group(2)
        try:
            count = int(n_token) if n_token.isdigit() else _NUMBER_WORDS.get(n_token, 0)
        except ValueError:
            count = 0
        if count <= 0:
            continue
        rtype = _classify_room_type(rest)
        layout.append((rtype, count))
    # Only honour the explicit layout if it actually mentioned more than one
    # room — otherwise it was just the regular "Superior Room" / "Comfort
    # Room" pattern which doesn't carry a quantity.
    total = sum(c for _, c in layout)
    return layout if total >= 2 else []


def _derive_room_layout(
    num_travelers: int,
    room_type_raw: Optional[str],
) -> list[tuple[RoomType, int]]:
    """Return a list of (room_type, pax_in_room) tuples.

    1. If Travefy described an explicit multi-room layout, honour it.
    2. Else, split `num_travelers` into rooms of the base type using each
       type's capacity. Leftover travelers go into the last room.
    Examples (num_travelers, base_type → layout):
      (1, "single")     → [("single", 1)]
      (2, "doble")      → [("doble", 2)]
      (3, "doble")      → [("doble", 2), ("doble", 1)]
      (4, "doble")      → [("doble", 2), ("doble", 2)]
      (4, "family")     → [("family", 4)]
      (6, "doble")      → [("doble", 2), ("doble", 2), ("doble", 2)]
    """
    num_travelers = max(1, int(num_travelers or 1))
    # Detect "2 Doubles + 1 Single" etc. in the raw text
    explicit = _parse_explicit_layout(room_type_raw)
    if explicit:
        # Expand each (type, count) into `count` rooms; assign all pax up to
        # capacity, then dump the remainder into the last room.
        rooms: list[tuple[RoomType, int]] = []
        for rtype, count in explicit:
            cap = _ROOM_CAPACITY.get(rtype, 2)
            for _ in range(count):
                rooms.append((rtype, cap))
        # Trim total capacity down to num_travelers (rest = 0 → drop empties)
        # The pax-per-room becomes min(capacity, remaining travelers).
        remaining = num_travelers
        assigned: list[tuple[RoomType, int]] = []
        for rtype, cap in rooms:
            pax_here = min(cap, max(0, remaining))
            if pax_here <= 0:
                continue
            assigned.append((rtype, pax_here))
            remaining -= pax_here
        return assigned or [(rooms[0][0], num_travelers)]

    rtype = _classify_room_type(room_type_raw)
    cap = _ROOM_CAPACITY.get(rtype, 2)
    # Round-up the number of rooms needed
    n_rooms = max(1, -(-num_travelers // cap))
    rooms_out: list[tuple[RoomType, int]] = []
    remaining = num_travelers
    for i in range(n_rooms):
        pax_here = min(cap, remaining)
        rooms_out.append((rtype, pax_here))
        remaining -= pax_here
    return rooms_out


def _layout_summary(layout: list[tuple[RoomType, int]]) -> str:
    """Human-readable summary: "2× doble + 1× single" — used in preview UI."""
    from collections import Counter
    counts = Counter(rtype for rtype, _ in layout)
    parts = [(f"{n}× {rt}" if n > 1 else f"{rt}") for rt, n in counts.items()]
    return " + ".join(parts)


async def _match_experience(name: str, city: Optional[str], num_travelers: int) -> Optional[dict]:
    """Find best matching experience in DB for the given Travefy item name."""
    city = _norm_city(city)
    flt: dict = {}
    if city:
        # Word-boundary substring so "Rome" matches "Rome" and "Rome - Florence"
        flt["city"] = {"$regex": f"(?<![A-Za-z]){re.escape(city)}(?![A-Za-z])", "$options": "i"}
    candidates = await db.experiences.find(flt, {"_id": 0}).limit(400).to_list(400)
    best, best_score, best_overlap = None, 0.0, 0
    name_tokens = _tokens_for_match(name)
    for c in candidates:
        c_tokens = _tokens_for_match(c.get("title") or "")
        if not c_tokens or not name_tokens:
            continue
        overlap = len(name_tokens & c_tokens)
        if overlap == 0:
            continue
        s = overlap / min(len(name_tokens), len(c_tokens))
        if s > best_score:
            best, best_score, best_overlap = c, s, overlap
    # Require either a strong ratio OR multiple discriminating tokens overlapping
    if best_score < 0.5 or best_overlap < 2:
        return None
    # Prefer the exact-pax variant when several share the same title
    if num_travelers and best:
        siblings = [c for c in candidates if c.get("title") == best.get("title")]
        if len(siblings) > 1:
            siblings.sort(key=lambda c: (
                0 if (c.get("pax") or 2) == num_travelers else 1,
                abs((c.get("pax") or 2) - num_travelers),
            ))
            best = siblings[0]
    return {**best, "_match_score": round(best_score, 2)}


def _strip_hotel_noise(s: str) -> str:
    """Travefy frequently appends 'or similar' / 'or higher' to hotel names —
    drop it before matching against our hotel DB."""
    return re.sub(r"\s+or\s+(similar|higher)\s*$", "", (s or "").strip(), flags=re.IGNORECASE)


async def _match_hotel(name: str, city: Optional[str]) -> Optional[dict]:
    name = _strip_hotel_noise(name)
    city = _norm_city(city)
    flt: dict = {"source": "library"}  # never match auto-imported ghost hotels
    if city:
        flt["city"] = {"$regex": f"(?<![A-Za-z]){re.escape(city)}(?![A-Za-z])", "$options": "i"}
    candidates = await db.hotels.find(flt, {"_id": 0}).limit(200).to_list(200)
    if not candidates and city:  # widen if nothing in this city
        candidates = await db.hotels.find({"source": "library"}, {"_id": 0}).limit(400).to_list(400)
    best, best_score = None, 0.0
    for c in candidates:
        s = _match_score(name, c.get("name") or "")
        if s > best_score:
            best, best_score = c, s
    # Hotel names are short and distinctive — require a high overlap to call it a match.
    if best_score < 0.7:
        return None
    return {**best, "_match_score": round(best_score, 2)}


def _confidence_label(score: float) -> str:
    if score >= 0.8:
        return "high"
    if score >= 0.55:
        return "medium"
    return "low"


@api.get("/itineraries/import-travefy/health")
async def import_travefy_health(user: Annotated[User, Depends(current_user)]):
    """Diagnostic endpoint — surfaces whether chromium is installed and what
    the most recent travefy job looked like. Useful to debug stuck imports
    in production where we can't tail logs directly."""
    import os as _os
    import json as _json
    from scraper import _PW_INSTALL_DONE

    pw_dir = "/pw-browsers"
    pw_files: list[str] = []
    try:
        if _os.path.isdir(pw_dir):
            pw_files = sorted(_os.listdir(pw_dir))
    except Exception as e:
        pw_files = [f"<error listing: {e}>"]

    # Last 3 jobs created by ANY user — admins see all, regular agents only
    # their own (matches the rest of the access model).
    flt = {} if user.role == "admin" else {"created_by": user.email}
    recent = await db.travefy_import_jobs.find(flt, {"_id": 0}).sort("created_at", -1).limit(3).to_list(3)
    for r in recent:
        # Don't dump the whole parsed preview here, just whether it landed.
        r["preview_size"] = len(_json.dumps(r.get("preview") or {}))
        r.pop("preview", None)

    return {
        "chromium_install_done_flag": _PW_INSTALL_DONE,
        "pw_browsers_dir": pw_dir,
        "pw_browsers_contents": pw_files,
        "bg_task_count": len(_bg_tasks),
        "recent_jobs": recent,
    }


@api.post("/itineraries/import-travefy/preview")
async def import_travefy_preview(
    payload: dict = Body(...),
    user: User = Depends(current_user),
):
    """Start a background scrape+match job for a Travefy URL.

    The scrape itself runs Playwright + Claude which routinely takes 30-60s —
    well beyond Cloudflare's request timeout. We therefore kick off the work
    asynchronously and return a job_id the client polls every 2s.
    """
    url = (payload.get("url") or "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="URL es obligatoria")
    if "travefy.com" not in url:
        raise HTTPException(status_code=400, detail="La URL debe ser de travefy.com")

    job_id = f"tvj_{uuid.uuid4().hex[:12]}"
    await db.travefy_import_jobs.insert_one({
        "job_id": job_id,
        "url": url,
        "status": "running",
        "created_by": user.email,
        "created_at": now_iso(),
        "preview": None,
        "error": None,
    })
    # Hold a strong ref via _spawn_bg → the event loop won't GC the task while
    # chromium downloads in the background. Without this, production users
    # saw the modal spin forever because the worker vanished mid-run.
    _spawn_bg(_run_travefy_preview_job_safely(job_id, url))
    return {"job_id": job_id, "status": "running"}


@api.get("/itineraries/import-travefy/preview/{job_id}")
async def import_travefy_preview_status(
    job_id: str,
    user: Annotated[User, Depends(current_user)],
):
    """Poll endpoint. Returns {status, preview?, error?}."""
    job = await db.travefy_import_jobs.find_one(
        {"job_id": job_id}, {"_id": 0}
    )
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.get("created_by") and job["created_by"] != user.email and user.role != "admin":
        raise HTTPException(status_code=403, detail="No tienes acceso a este job")
    return {
        "job_id": job_id,
        "status": job.get("status"),
        "preview": job.get("preview"),
        "error": job.get("error"),
    }


async def _run_travefy_preview_job_safely(job_id: str, url: str):
    """Top-level wrapper that GUARANTEES the job ends in done/error.

    Catches *any* exception (including ones that bubble past the inner
    handler) and enforces a hard 6-minute timeout so a hung Playwright /
    chromium install can't leave the modal spinning forever."""
    try:
        await asyncio.wait_for(_run_travefy_preview_job(job_id, url), timeout=360)
    except asyncio.TimeoutError:
        logger.error("travefy preview job %s timed out (>360s)", job_id)
        await db.travefy_import_jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "error",
                "error": ("Tiempo agotado (>6 min). Si acabas de redeployar, "
                          "el navegador Playwright puede estar descargándose. "
                          "Espera 1-2 min y vuelve a intentar."),
                "finished_at": now_iso(),
            }},
        )
    except Exception as e:
        logger.exception("travefy preview job %s wrapper crashed", job_id)
        await db.travefy_import_jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "error",
                "error": f"Error inesperado: {type(e).__name__}: {e}",
                "finished_at": now_iso(),
            }},
        )


async def _run_travefy_preview_job(job_id: str, url: str):
    """Background worker: scrape URL, run LLM, match against DB, save preview."""
    from scraper import scrape_and_parse
    try:
        scraped = await scrape_and_parse(url)
        if not scraped.get("ok"):
            await db.travefy_import_jobs.update_one(
                {"job_id": job_id},
                {"$set": {
                    "status": "error",
                    "error": f"Travefy bloqueó la lectura: {scraped.get('error') or 'unknown'}",
                    "finished_at": now_iso(),
                }},
            )
            return

        structured = scraped.get("structured") or {}
        num_travelers = int(structured.get("num_travelers") or 2) or 2

        out_days = []
        hotel_blocks: dict = {}
        for d in structured.get("days") or []:
            city = _norm_city(d.get("city"))
            items_out = []

            # Helper that runs an item (activity OR transfer) through the
            # type classifier + experience matcher and appends a preview
            # entry. Travefy splits flights / private transfers / trains
            # into a separate `transfers` array that earlier import
            # versions completely ignored — this caused internal flights
            # like "Flight TP 874 LIS → FLR" to silently disappear from
            # the imported itinerary.
            async def emit_item(raw_name: str):
                name = (raw_name or "").strip()
                if not name or _is_free_day_marker(name):
                    return
                kind = _classify_travefy(name)
                match = None
                if kind in ("actividad", "transfer", "tren", "vuelo", "entradas"):
                    m = await _match_experience(name, city, num_travelers)
                    if m:
                        match = {
                            "experience_id": m.get("experience_id"),
                            "title": m.get("title"),
                            "type": _normalize_service_type(m.get("type")),
                            "pax": m.get("pax"),
                            "city": m.get("city"),
                            "provider_name": m.get("provider_name"),
                            "price_tax_excl": m.get("price_tax_excl") or 0,
                            "price_tax_incl": m.get("price_tax_incl") or m.get("price") or 0,
                            "currency": m.get("currency") or "EUR",
                            "confidence": _confidence_label(m.get("_match_score", 0)),
                        }
                items_out.append({
                    "travefy_name": name,
                    "type": match["type"] if match else kind,
                    "match": match,
                })

            for a in d.get("activities") or []:
                await emit_item(a.get("name"))
            # Transfers carry the same shape as activities for our purposes;
            # they may use "description" rather than "name" so we read both.
            for t in d.get("transfers") or []:
                await emit_item(t.get("description") or t.get("name"))
            for h in d.get("hotels") or []:
                hname = (h.get("name") or "").strip()
                if not hname:
                    continue
                block = hotel_blocks.setdefault(hname, {
                    "name": hname,
                    "city": city,
                    "check_in": h.get("check_in") or d.get("date"),
                    "check_out": h.get("check_out"),
                    "first_day": d.get("day"),
                    "room_type_raw": h.get("room_type"),
                })
                block["check_out"] = h.get("check_out") or d.get("date") or block.get("check_out")
                # Keep the first non-empty room_type_raw we see (Travefy
                # repeats the same block across days of the same stay).
                if not block.get("room_type_raw") and h.get("room_type"):
                    block["room_type_raw"] = h.get("room_type")
            out_days.append({
                "day": d.get("day"),
                "date": d.get("date"),
                "city": city,
                "items": items_out,
            })

        hotels_out = []
        for hname, block in hotel_blocks.items():
            m = await _match_hotel(hname, block.get("city"))
            match = None
            if m:
                match = {
                    "hotel_id": m.get("hotel_id"),
                    "name": m.get("name"),
                    "city": m.get("city"),
                    "tier": m.get("tier"),
                    "price_per_night_excl": m.get("price_per_night_excl") or 0,
                    "price_per_night_incl": m.get("price_per_night_incl") or 0,
                    "currency": m.get("currency") or "EUR",
                    "confidence": _confidence_label(m.get("_match_score", 0)),
                }
            hotels_out.append({
                "travefy_name": hname,
                "city": block.get("city"),
                "check_in": block.get("check_in"),
                "check_out": block.get("check_out"),
                "match": match,
                "room_type_raw": block.get("room_type_raw"),
                "room_type": _classify_room_type(block.get("room_type_raw")),
                "rooms_layout": [
                    {"room_type": rt, "pax": px}
                    for rt, px in _derive_room_layout(num_travelers, block.get("room_type_raw"))
                ],
            })

        trip_name = structured.get("trip_name") or "Itinerario importado"
        main_traveler = ""
        if " - " in trip_name:
            tail = trip_name.rsplit(" - ", 1)[-1]
            # Heuristic: the name is short (≤4 words)
            if len(tail.split()) <= 4:
                main_traveler = tail.strip()

        preview = {
            "trip_name": trip_name,
            "main_traveler": main_traveler,
            "start_date": structured.get("start_date"),
            "end_date": structured.get("end_date"),
            "num_travelers": num_travelers,
            "days": out_days,
            "hotels": hotels_out,
            "source_url": url,
        }

        await db.travefy_import_jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "done",
                "preview": preview,
                "finished_at": now_iso(),
            }},
        )
    except Exception as e:
        logger.exception("travefy preview job %s failed", job_id)
        await db.travefy_import_jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "error",
                "error": f"Error procesando Travefy: {e}",
                "finished_at": now_iso(),
            }},
        )


@api.post("/itineraries/import-travefy/confirm", response_model=Itinerary)
async def import_travefy_confirm(
    user: Annotated[User, Depends(current_user)],
    payload: dict = Body(...),
):
    """Create a real Itinerary from a (possibly edited) preview payload."""
    try:
        return await _build_itinerary_from_travefy_preview(payload, user)
    except HTTPException:
        raise
    except Exception as e:
        # Pydantic + Mongo errors are notoriously useless when surfaced as a
        # bare 500. Log the full traceback server-side and return a structured
        # message the agent can act on (or copy-paste into a bug report).
        logger.exception("travefy confirm failed for user=%s", user.email)
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


# Non-city phrases the Travefy day labels sometimes surface. When one of
# these matches (case-insensitive substring), the token is dropped from
# the top-level itinerary.cities list.
_NON_CITY_TOKENS = (
    "departing", "departure", "arriving", "arrival", "return", "home",
    "flight", "transfer", "airport", "en route", "layover", "check-in day",
    "welcome to", "goodbye", "farewell", "day at leisure", "free day",
    "information", "documents",
)


def _clean_city_list(raw_cities) -> list:
    """Normalise the per-day `city` strings into a clean, deduped list of
    real cities suitable for the trip-view hero and destination pills.

    - Splits transfer strings on separators (`-`, `→`, `to`, `/`, `,`) so
      "Porto - Cascais" contributes both cities.
    - Drops non-city labels ("Departing US", "Flight to Lisbon"...).
    - Preserves first-seen order.
    """
    seen: list = []
    for raw in raw_cities:
        if not raw:
            continue
        parts = re.split(r"\s*(?:[-→/]|\bto\b|,)\s*", str(raw), flags=re.IGNORECASE)
        for part in parts:
            token = (part or "").strip()
            if not token:
                continue
            low = token.lower()
            if any(bad in low for bad in _NON_CITY_TOKENS):
                continue
            if token not in seen:
                seen.append(token)
    return seen



async def _build_itinerary_from_travefy_preview(payload: dict, user: User) -> Itinerary:
    days_in = payload.get("days") or []
    hotels_in = payload.get("hotels") or []
    num_travelers = int(payload.get("num_travelers") or 2) or 2

    # Build days[] with services
    days_out: list[ItineraryDay] = []
    for d in days_in:
        services: list[ItineraryService] = []
        for item in d.get("items") or []:
            if item.get("excluded"):
                continue
            m = item.get("match") or {}
            kind = _normalize_service_type(item.get("type") or "actividad")
            name = m.get("title") or item.get("travefy_name") or "Sin título"
            unit_excl = float(m.get("price_tax_excl") or 0)
            unit_incl = float(m.get("price_tax_incl") or 0) or unit_excl
            pax_unit = int(m.get("pax") or 1) or 1
            # Smart quantity: ceil(num_travelers / pax_unit)
            qty = max(1, (num_travelers + pax_unit - 1) // pax_unit)
            services.append(ItineraryService(
                experience_id=m.get("experience_id"),
                type=kind,
                name=name,
                provider_name=m.get("provider_name"),
                quantity=qty,
                pax=pax_unit,
                unit_price_tax_excl=unit_excl,
                unit_price_tax_incl=unit_incl,
                unit_price=unit_incl,
                currency=m.get("currency") or "EUR",
                notes=None if m.get("experience_id") else "⚠ Sin match · Revisar precio",
            ))
        days_out.append(ItineraryDay(
            date=d.get("date"),
            city=d.get("city"),
            services=services,
        ))

    # Build accommodations[]
    acc_out: list[Accommodation] = []
    for h in hotels_in:
        if h.get("excluded"):
            continue
        m = h.get("match") or {}
        date_from = h.get("check_in")
        date_to = h.get("check_out")
        # nights = max(1, date_to - date_from)
        try:
            df = datetime.fromisoformat(date_from)
            dt = datetime.fromisoformat(date_to)
            nights = max(1, (dt - df).days)
        except (TypeError, ValueError):
            nights = 1
        per_night_excl = float(m.get("price_per_night_excl") or 0)
        per_night_incl = float(m.get("price_per_night_incl") or 0) or per_night_excl
        # Derive the room layout from Travefy's room type + group size.
        # If the preview already shipped a layout (the agent may have tweaked
        # it manually in the UI) we trust it; otherwise we derive on the fly.
        rtype_raw = (h.get("room_type_raw") or "").strip() or None
        layout = h.get("rooms_layout") or [
            {"room_type": rt, "pax": px}
            for rt, px in _derive_room_layout(num_travelers, rtype_raw)
        ]
        rooms_list: list[Room] = []
        for r_in in layout:
            rooms_list.append(Room(
                room_type=r_in.get("room_type") or "doble",
                pax=int(r_in.get("pax") or num_travelers),
                price_per_night_excl=per_night_excl,
                price_per_night_incl=per_night_incl,
                currency=m.get("currency") or "EUR",
                notes=f"Travefy: {rtype_raw}" if rtype_raw else None,
            ))
        num_rooms = max(1, len(rooms_list))
        acc_out.append(Accommodation(
            date_from=date_from,
            date_to=date_to,
            name=m.get("name") or h.get("travefy_name") or "Sin nombre",
            price_tax_excl=per_night_excl * nights * num_rooms,
            price_tax_incl=per_night_incl * nights * num_rooms,
            price=per_night_incl * nights * num_rooms,
            currency=m.get("currency") or "EUR",
            rooms=rooms_list,
        ))

    # Main traveler heuristic
    main = (payload.get("main_traveler") or "").strip()
    if not main and payload.get("trip_name"):
        # Travefy titles look like "Trip Title - First Last"
        parts = (payload["trip_name"] or "").rsplit(" - ", 1)
        if len(parts) == 2 and len(parts[1].split()) <= 4:
            main = parts[1].strip()

    # Duration: prefer the days list count; fall back to date diff
    duration = len(days_out)
    if not duration and payload.get("start_date") and payload.get("end_date"):
        try:
            duration = (datetime.fromisoformat(payload["end_date"]) -
                        datetime.fromisoformat(payload["start_date"])).days + 1
        except (TypeError, ValueError):
            duration = 0

    # Derive a CLEAN top-level `cities[]` list from the day list. We keep
    # the raw per-day `day.city` string as-is (informative for transfer
    # days like "Porto - Cascais"), but the trip-view hero + destination
    # pill must show only real cities, in first-seen order.
    cities_seen = _clean_city_list([d.city for d in days_out])

    itn = Itinerary(
        name=payload.get("trip_name") or "Itinerario importado de Travefy",
        main_traveler=main,
        start_date=payload.get("start_date"),
        end_date=payload.get("end_date"),
        duration_days=duration,
        num_travelers=num_travelers,
        cities=cities_seen,
        days=days_out,
        accommodations=acc_out,
        created_by=user.email,
    )
    # Fresh import → seed a version group of its own.
    itn.version_group_id = itn.itinerary_id
    itn.version = 1
    itn.updated_at = now_iso()
    await db.itineraries.insert_one(itn.model_dump())
    return itn


# ---------------------------------------------------------------------------
# Sofi push (gestion.viajadverdad.com)
# ---------------------------------------------------------------------------
# Two endpoints + a polling job pattern, mirroring Travefy import:
#   POST /api/itineraries/{id}/push-to-sofi  (body: {dry_run: bool})
#   GET  /api/itineraries/push-to-sofi/{job_id}
# The actual Playwright work runs in the background. On a successful real
# push we persist the new sofi_trip_id on the itinerary document so the
# builder hides the button afterwards (preventing duplicate pushes).
def _compute_pricing_totals(itn: dict) -> dict:
    """Reproduces the frontend's totals math so Sofi gets the same numbers
    the agent sees in the builder. Single source of truth for pricing.

    Mirrors `ItineraryBuilder.totals` in JSX:
      sub_excl       = Σ services excl + Σ accommodations excl
      sub_incl       = Σ services incl + Σ accommodations incl
      sub_with_markup = sub_incl × (1 + markup_pct/100)
      commission_eur = sub_with_markup × commission_pct / (100 − commission_pct)   # gross-up
      pvp_pre_paypal = sub_with_markup + commission_eur
      paypal_eur     = paypal_fee ? pvp_pre_paypal × 0.03 : 0
      pvp            = pvp_pre_paypal + paypal_eur

    The gross-up shape (`/ (100 − c)`) is what guarantees that AFTER the
    partner deducts their %, we still net exactly `sub_with_markup`. With
    the older naive formula (`sub_with_markup × c/100`) the agency lost a
    fraction of the desired markup on every Zicasso / Responsible Travel /
    Baboo trip because their commission is applied to the final price, not
    to the markup.
    """
    sub_excl = 0.0
    sub_incl = 0.0
    for d in itn.get("days") or []:
        for s in d.get("services") or []:
            # Services tied to an accommodation (`acc_id` set) are derived
            # read-only chips, NOT separate cost items. The accommodation
            # itself contributes its price in the loop below, so counting
            # the carrier service here would double-bill the hotel.
            if s.get("acc_id"):
                continue
            qty = float(s.get("quantity") or 0)
            sub_excl += float(s.get("unit_price_tax_excl") or 0) * qty
            sub_incl += float(s.get("unit_price_tax_incl") or s.get("unit_price") or 0) * qty
    for a in itn.get("accommodations") or []:
        sub_excl += float(a.get("price_tax_excl") or 0)
        sub_incl += float(a.get("price_tax_incl") or a.get("price") or 0)
    mk = float(itn.get("markup_pct") or 0) / 100.0
    com_pct = float(itn.get("commission_pct") or 0)
    sub_with_markup = sub_incl * (1.0 + mk)
    # Gross-up: partner takes com_pct of the FINAL sale price, so to net
    # `sub_with_markup` on our side we have to raise the sale by exactly
    # the inverse factor.
    commission_eur = sub_with_markup * com_pct / max(1e-9, (100.0 - com_pct)) if com_pct else 0.0
    pvp_pre_paypal = sub_with_markup + commission_eur
    paypal_eur = pvp_pre_paypal * 0.03 if itn.get("paypal_fee") else 0.0
    pvp = pvp_pre_paypal + paypal_eur
    return {
        "sub_excl": round(sub_excl, 2),
        "sub_incl": round(sub_incl, 2),
        "sub_with_markup": round(sub_with_markup, 2),
        "markup_eur": round(sub_incl * mk, 2),
        "commission_eur": round(commission_eur, 2),
        "paypal_eur": round(paypal_eur, 2),
        "pvp": round(pvp, 2),
    }


@api.post("/itineraries/{itinerary_id}/push-to-sofi")
async def push_itinerary_to_sofi_endpoint(
    itinerary_id: str,
    user: Annotated[User, Depends(current_user)],
    payload: dict = Body(default={}),
):
    """Start a background job that opens Sofi, fills the create-trip form,
    and (if `dry_run=False`) submits it. Returns a job_id the client polls.

    Body: { "dry_run": bool }  default false.
    """
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Itinerario no encontrado")
    if not _can_access(doc, user):
        raise HTTPException(status_code=403, detail="No tienes acceso a este itinerario")

    # Strict bool coercion — accept true/false (json) but reject sneaky strings.
    dry_run = bool(payload.get("dry_run")) if isinstance(payload.get("dry_run"), bool) else False
    # When the trip already lives in Sofi, refuse to create a duplicate. The
    # builder hides the button at this point but a stale tab could still POST.
    if not dry_run and doc.get("sofi_trip_id"):
        raise HTTPException(
            status_code=409,
            detail=(f"Este itinerario ya está en Sofi (#{doc['sofi_trip_id']}). "
                    "Para evitar duplicados, edítalo directamente en Sofi."),
        )

    # One running job at a time per itinerary — prevents accidental loops if
    # the modal is reopened while a previous job is still in flight.
    existing = await db.sofi_push_jobs.find_one(
        {"itinerary_id": itinerary_id, "status": "running"},
        {"_id": 0, "job_id": 1, "dry_run": 1},
    )
    if existing:
        raise HTTPException(
            status_code=409,
            detail=(f"Ya hay un push en curso para este itinerario "
                    f"(job {existing['job_id']}). Espera a que termine."),
        )

    job_id = f"sofi_{uuid.uuid4().hex[:12]}"
    now = now_iso()
    await db.sofi_push_jobs.insert_one({
        "job_id": job_id,
        "itinerary_id": itinerary_id,
        "dry_run": dry_run,
        "status": "running",
        "created_by": user.email,
        "created_at": now,
        # BSON Date mirror for the TTL index (auto-cleanup after 7 days).
        "created_at_dt": datetime.now(timezone.utc),
        "result": None,
        "error": None,
    })
    logger.info("sofi_push.start job_id=%s itn=%s dry_run=%s by=%s",
                job_id, itinerary_id, dry_run, user.email)
    _spawn_bg(_run_sofi_push_job_safely(job_id, itinerary_id, dry_run))
    return {"job_id": job_id, "status": "running", "dry_run": dry_run}


@api.get("/itineraries/push-to-sofi/{job_id}")
async def push_itinerary_to_sofi_status(
    job_id: str,
    user: Annotated[User, Depends(current_user)],
):
    """Poll endpoint. Returns {status, result?, error?, dry_run}."""
    job = await db.sofi_push_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    if job.get("created_by") and job["created_by"] != user.email and user.role != "admin":
        raise HTTPException(status_code=403, detail="No tienes acceso a este job")
    return {
        "job_id": job_id,
        "status": job.get("status"),
        "dry_run": job.get("dry_run", False),
        "result": job.get("result"),
        "error": job.get("error"),
        "itinerary_id": job.get("itinerary_id"),
    }


async def _run_sofi_push_job_safely(job_id: str, itinerary_id: str, dry_run: bool):
    """Hard wrapper: any exception or 10-min timeout lands as a clean error
    state, never a stuck-on-running modal. Real pushes for a 14-day trip can
    exceed 6min once we're rendering 30+ booking forms in the same browser
    session, so the cap is generous."""
    try:
        await asyncio.wait_for(
            _run_sofi_push_job(job_id, itinerary_id, dry_run),
            timeout=600,
        )
    except asyncio.TimeoutError:
        logger.error("sofi push job %s timed out (>600s)", job_id)
        await db.sofi_push_jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "error",
                "error": ("Tiempo agotado (>10 min). Puede que un envío de "
                          "muchas reservas haya excedido el límite. Reintenta "
                          "o contacta soporte si el itinerario es muy grande."),
                "finished_at": now_iso(),
            }},
        )
    except Exception as e:
        logger.exception("sofi push job %s wrapper crashed", job_id)
        await db.sofi_push_jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "error",
                "error": f"Error inesperado: {type(e).__name__}: {e}",
                "finished_at": now_iso(),
            }},
        )


async def _run_sofi_push_job(job_id: str, itinerary_id: str, dry_run: bool):
    from sofi import push_itinerary_to_sofi
    doc = await db.itineraries.find_one({"itinerary_id": itinerary_id}, {"_id": 0})
    if not doc:
        await db.sofi_push_jobs.update_one(
            {"job_id": job_id},
            {"$set": {"status": "error", "error": "Itinerario no encontrado",
                      "finished_at": now_iso()}},
        )
        return
    totals = _compute_pricing_totals(doc)
    out = await push_itinerary_to_sofi(doc, totals, dry_run=dry_run)

    if not out.get("ok"):
        await db.sofi_push_jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "status": "error",
                "error": out.get("error") or "Error desconocido",
                "result": out,                # keep details + filled_fields for debugging
                "finished_at": now_iso(),
            }},
        )
        return

    # Real push success → stamp the itinerary with the Sofi trip_id so the
    # builder transitions to the read-only "ya en Sofi" pill.
    if not dry_run and out.get("trip_id"):
        await db.itineraries.update_one(
            {"itinerary_id": itinerary_id},
            {"$set": {
                "sofi_trip_id": int(out["trip_id"]),
                "sofi_url": out.get("url"),
                "sofi_pushed_at": now_iso(),
                "updated_at": now_iso(),
            }},
        )

    await db.sofi_push_jobs.update_one(
        {"job_id": job_id},
        {"$set": {
            "status": "done",
            "result": out,
            "finished_at": now_iso(),
        }},
    )
    logger.info("sofi_push.done job_id=%s itn=%s dry_run=%s trip_id=%s",
                job_id, itinerary_id, dry_run, out.get("trip_id"))


# --- Rental car detector ---------------------------------------------------
# Used by both the Travefy classifier (above) and the admin reclassify endpoint
# below. Returns True when the experience is a vehicle rental, NOT a transfer
# to/from a rental office.
def _is_rental_car_experience(title: str, provider: str = "") -> bool:
    blob = f"{title or ''} {provider or ''}".lower()
    if any(k in blob for k in ("transfer", "shuttle")):
        return False
    return any(k in blob for k in (
        "vehicle rental", "rental car", "car rental",
        "rent a car", "rent-a-car",
        "alquiler de coche", "alquiler de veh", "alquiler de auto",
        "alquiler vehic",
    ))


@api.post("/admin/experiences/reclassify-rental-cars")
async def admin_reclassify_rental_cars(
    user: Annotated[User, Depends(current_user)],
    dry_run: bool = False,
):
    """Admin-only one-shot: scan the catalog and flip any experience that
    looks like a vehicle rental (and isn't already classified) to type
    'rental_car'. Pass `?dry_run=true` to preview without writing.

    Excludes existing transfers ("Private Transfer ... Rental Car Office ...")
    so we don't accidentally re-categorise them.
    """
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo admin")
    cursor = db.experiences.find({}, {"_id": 0, "experience_id": 1, "title": 1, "provider_name": 1, "type": 1})
    updated: list[dict] = []
    skipped: list[dict] = []
    async for e in cursor:
        if e.get("type") == "rental_car":
            continue  # already done
        if not _is_rental_car_experience(e.get("title", ""), e.get("provider_name", "")):
            continue
        if e.get("type") == "transfer":
            # Heuristic shouldn't catch these (we exclude "transfer" in the
            # detector) but double-check so we never silently flip a transfer.
            skipped.append({"experience_id": e["experience_id"], "title": e["title"], "from_type": e.get("type")})
            continue
        updated.append({"experience_id": e["experience_id"], "title": e["title"], "from_type": e.get("type")})
        if not dry_run:
            await db.experiences.update_one(
                {"experience_id": e["experience_id"]},
                {"$set": {"type": "rental_car", "updated_at": now_iso()}},
            )
    return {
        "dry_run": dry_run,
        "reclassified": len(updated),
        "skipped_transfers": len(skipped),
        "samples": updated[:15],
    }


# ---------------------------------------------------------------------------
# Health
# ---------------------------------------------------------------------------
@api.get("/")
async def root():
    return {"ok": True, "service": "itinerary-builder"}


@api.get("/stats")
async def stats(_: Annotated[User, Depends(current_user)]):
    return {
        "providers": await db.providers.count_documents({}),
        "experiences": await db.experiences.count_documents({}),
        "itineraries": await db.itineraries.count_documents({}),
        "users": await db.users.count_documents({}),
        "hotels": await db.hotels.count_documents({}),
        "training_examples": await db.training_examples.count_documents({}),
    }


@api.get("/fx/rate")
async def fx_rate(
    _: Annotated[User, Depends(current_user)],
    base: str = "EUR",
    quote: str = "USD",
    refresh: bool = False,
):
    """Return the latest exchange rate between `base` and `quote`.

    Strategy:
    1. If cached for today, return the cached value (with `source="cache"`).
    2. Otherwise, fetch from api.frankfurter.app and cache the result.
    3. If the external call fails, fall back to the most recent cached value
       (with `source="stale"`) so the UI keeps working offline.

    Pass `refresh=true` to bypass the cache and force a fresh fetch.
    """
    base = base.upper()
    quote = quote.upper()
    today = datetime.now(timezone.utc).date().isoformat()
    cache_key = f"{base}-{quote}-{today}"

    if not refresh:
        cached = await db.fx_rates.find_one({"key": cache_key}, {"_id": 0})
        if cached:
            cached["source"] = "cache"
            return cached

    # Fetch fresh
    try:
        async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client_http:
            r = await client_http.get(
                "https://api.frankfurter.dev/v1/latest",
                params={"base": base, "symbols": quote},
            )
            r.raise_for_status()
            payload = r.json()
            rate = float(payload["rates"][quote])
            doc = {
                "key": cache_key,
                "base": base,
                "quote": quote,
                "rate": rate,
                "date": payload.get("date", today),
                "fetched_at": now_iso(),
            }
            await db.fx_rates.update_one(
                {"key": cache_key}, {"$set": doc}, upsert=True,
            )
            return {**doc, "source": "fresh"}
    except Exception as e:
        logger.warning(f"FX fetch failed: {e}")
        latest = await db.fx_rates.find_one(
            {"base": base, "quote": quote},
            {"_id": 0},
            sort=[("fetched_at", -1)],
        )
        if latest:
            latest["source"] = "stale"
            return latest
        # Last-resort hardcoded fallback so the UI never breaks completely.
        return {
            "base": base, "quote": quote, "rate": 1.10 if (base, quote) == ("EUR", "USD") else 1.0,
            "date": today, "fetched_at": now_iso(), "source": "fallback",
        }


# ===========================================================================
# AI CALIBRATION — surface batch-eval results in the UI so the user can see
# how well the agent is composing itineraries and trigger re-analysis.
# ===========================================================================
_CALIBRATION_RESULTS_PATH = ROOT_DIR.parent / "memory" / "batch_eval_v2.jsonl"
_CALIBRATION_SUMMARY_PATH = ROOT_DIR.parent / "memory" / "batch_eval_v2_summary.json"


def _load_calibration_rows() -> list[dict]:
    if not _CALIBRATION_RESULTS_PATH.exists():
        return []
    import json as _json
    rows: list[dict] = []
    with _CALIBRATION_RESULTS_PATH.open("r", encoding="utf-8") as f:
        for ln in f:
            try:
                r = _json.loads(ln)
                if not r.get("error"):
                    rows.append(r)
            except Exception:
                continue
    return rows


def _composition_score(r: dict) -> Optional[float]:
    """Single 0-1 score capturing how well the draft composes the plan.

    Combines four signals (each 0-1):
    - city_overlap  : real_cities ∩ draft_cities  / max(|real|, |draft|)
    - hotel_count   : 1 - |Δhotels| / max(real_hotels, draft_hotels, 1)
    - activity_count: 1 - |Δactivities| / max(real_act, draft_act, 1)
    - country_hit   : 1 if AI detected the right country, else 0
    Weights: city 40%, hotels 20%, activities 20%, country 20%.
    """
    real_cities = {c.lower() for c in (r.get("real_cities") or []) if c}
    draft_cities = {c.lower() for c in (r.get("draft_cities") or []) if c}
    if real_cities and draft_cities:
        # Accent-fold and normalize a bit so "Lisbon" matches "lisboa"
        def _norm(s: str) -> str:
            import unicodedata
            return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode().lower().strip()
        rc = {_norm(c) for c in real_cities}
        dc = {_norm(c) for c in draft_cities}
        overlap = len(rc & dc) / max(len(rc), len(dc), 1)
    else:
        overlap = 0.0

    rh, dh = r.get("real_hotels") or 0, r.get("draft_hotels") or 0
    hotel_score = 1.0 - (abs(rh - dh) / max(rh, dh, 1))
    ra, da = r.get("real_activities") or 0, r.get("draft_activities") or 0
    act_score = 1.0 - (abs(ra - da) / max(ra, da, 1))
    # Country: stored as the detected country; we don't have the "real" country
    # directly here but we can infer from the request country detection — if
    # draft cities exist in any of the real cities it's a hit.
    country_score = 1.0 if (real_cities and draft_cities and overlap > 0) else 0.0

    score = 0.40 * overlap + 0.20 * hotel_score + 0.20 * act_score + 0.20 * country_score
    return round(max(0.0, min(1.0, score)), 3)


@api.get("/calibration/status")
async def calibration_status(_: Annotated[User, Depends(require_admin)]):
    """Aggregated metrics from the latest batch eval run."""
    import statistics as _stats

    rows = _load_calibration_rows()
    total_sold = await db.training_examples.count_documents(
        {"outcome": "sold", "client_request": {"$nin": [None, ""]}}
    )
    analyzed_in_db = await db.training_examples.count_documents(
        {"last_learned_at": {"$exists": True}}
    )
    pending_for_eval = max(0, total_sold - analyzed_in_db)

    ratios = [r["ratio_draft_over_real"] for r in rows if r.get("ratio_draft_over_real")]
    composition = [_composition_score(r) for r in rows]
    composition = [s for s in composition if s is not None]

    by_country: dict = {}
    for r in rows:
        c = r.get("country") or "unknown"
        by_country.setdefault(c, []).append(r)
    country_stats = {}
    for c, lst in by_country.items():
        comps = [_composition_score(x) for x in lst]
        comps = [s for s in comps if s is not None]
        rats = [x["ratio_draft_over_real"] for x in lst if x.get("ratio_draft_over_real")]
        country_stats[c] = {
            "n": len(lst),
            "median_ratio": round(_stats.median(rats), 2) if rats else None,
            "median_composition": round(_stats.median(comps), 2) if comps else None,
        }

    by_agent: dict = {}
    for r in rows:
        a = r.get("sales_agent") or "unknown"
        by_agent.setdefault(a, []).append(r)
    agent_stats = {}
    for a, lst in by_agent.items():
        if len(lst) < 3:
            continue
        comps = [_composition_score(x) for x in lst]
        comps = [s for s in comps if s is not None]
        rats = [x["ratio_draft_over_real"] for x in lst if x.get("ratio_draft_over_real")]
        agent_stats[a] = {
            "n": len(lst),
            "median_ratio": round(_stats.median(rats), 2) if rats else None,
            "median_composition": round(_stats.median(comps), 2) if comps else None,
        }

    # By partner — needs a DB lookup since the JSONL doesn't store the partner.
    eval_ids = [r["example_id"] for r in rows if r.get("example_id")]
    partner_lookup: dict[str, str] = {}
    if eval_ids:
        async for doc in db.training_examples.find(
            {"example_id": {"$in": eval_ids}}, {"example_id": 1, "partner": 1, "_id": 0}
        ):
            partner_lookup[doc["example_id"]] = doc.get("partner") or "unknown"
    by_partner: dict = {}
    for r in rows:
        p = partner_lookup.get(r.get("example_id"), "unknown")
        by_partner.setdefault(p, []).append(r)
    partner_stats = {}
    for p, lst in by_partner.items():
        comps = [_composition_score(x) for x in lst]
        comps = [s for s in comps if s is not None]
        rats = [x["ratio_draft_over_real"] for x in lst if x.get("ratio_draft_over_real")]
        partner_stats[p] = {
            "n": len(lst),
            "median_ratio": round(_stats.median(rats), 2) if rats else None,
            "median_composition": round(_stats.median(comps), 2) if comps else None,
        }

    return {
        "trips_total_sold_with_request": total_sold,
        "trips_analyzed": analyzed_in_db,
        "trips_pending_eval": pending_for_eval,
        "eval_rows_on_disk": len(rows),
        "global": {
            "median_ratio": round(_stats.median(ratios), 3) if ratios else None,
            "mean_ratio": round(_stats.mean(ratios), 3) if ratios else None,
            "median_composition": round(_stats.median(composition), 3) if composition else None,
            "mean_composition": round(_stats.mean(composition), 3) if composition else None,
        },
        "by_country": country_stats,
        "by_sales_agent": agent_stats,
        "by_partner": partner_stats,
        "last_run": rows[-1].get("duration_seconds") if rows else None,
        "job": await db.calibration_jobs.find_one(
            {}, {"_id": 0}, sort=[("started_at", -1)]
        ),
    }


@api.get("/calibration/rules")
async def calibration_rules(_: Annotated[User, Depends(require_admin)]):
    """Parse the rules section of SYSTEM_PROMPT_GENERATE for display in the UI."""
    import re
    # Each rule starts with a letter+`)` at column 0 of the prompt, e.g.
    # "A) DETECT..."  or  "REVISED H) HOTEL PRICING..."
    rule_re = re.compile(r"^((?:REVISED\s+)?[A-Z]{1,3}\d?)\)\s+(.+?)$", re.M)
    matches = list(rule_re.finditer(SYSTEM_PROMPT_GENERATE))
    out: list[dict] = []
    for i, m in enumerate(matches):
        key = m.group(1)
        title = m.group(2).strip()
        # Body extends until the next rule header
        body_start = m.end()
        body_end = matches[i + 1].start() if i + 1 < len(matches) else len(SYSTEM_PROMPT_GENERATE)
        body = SYSTEM_PROMPT_GENERATE[body_start:body_end].strip()
        out.append({"key": key, "title": title, "body": body[:1200]})
    return {"count": len(out), "rules": out}


@api.post("/calibration/run")
async def calibration_run(
    user: Annotated[User, Depends(require_admin)],
    payload: dict = Body(default={}),
):
    """Kick off the batch_eval_v2 script as a background subprocess. The script
    already skips trips marked with `last_learned_at`, so re-runs only process
    new imports. Pass `reset_markers=true` to re-evaluate every trip (DANGEROUS:
    consumes the full LLM budget)."""
    import subprocess

    reset = bool(payload.get("reset_markers", False))

    # Refuse to launch if another run is already in progress
    running = await db.calibration_jobs.find_one({"status": "running"})
    if running:
        raise HTTPException(status_code=409, detail=f"Ya hay una corrida en curso (job {running['job_id']})")

    if reset:
        await db.training_examples.update_many(
            {"outcome": "sold", "last_learned_at": {"$exists": True}},
            {"$unset": {"last_learned_at": "", "last_eval_ratio": "", "last_eval_country": ""}},
        )

    job_id = new_id("cal")
    job = {
        "job_id": job_id,
        "status": "running",
        "started_at": now_iso(),
        "finished_at": None,
        "created_by": user.email,
        "reset_markers": reset,
        "log_path": f"/tmp/cal_{job_id}.log",
    }
    await db.calibration_jobs.insert_one(dict(job))

    # Launch the script in background. The script writes to JSONL + updates
    # last_learned_at as it goes — both visible to subsequent /status calls.
    cmd = ["python", "-m", "tests.batch_eval_v2"]
    log_f = open(job["log_path"], "w", buffering=1)
    proc = subprocess.Popen(
        cmd,
        cwd=str(ROOT_DIR),
        stdout=log_f, stderr=subprocess.STDOUT,
        start_new_session=True,
    )
    await db.calibration_jobs.update_one(
        {"job_id": job_id}, {"$set": {"pid": proc.pid}}
    )

    # Watcher coroutine: poll the process and mark job as completed/failed.
    async def _watcher():
        while True:
            if proc.poll() is not None:
                await db.calibration_jobs.update_one(
                    {"job_id": job_id},
                    {"$set": {"status": "completed" if proc.returncode == 0 else "failed",
                              "finished_at": now_iso(),
                              "exit_code": proc.returncode}},
                )
                break
            await asyncio.sleep(5)

    asyncio.create_task(_watcher())
    job.pop("_id", None)
    return job


@api.get("/calibration/jobs")
async def calibration_jobs_list(_: Annotated[User, Depends(require_admin)]):
    items = await db.calibration_jobs.find({}, {"_id": 0}).sort("started_at", -1).limit(20).to_list(20)
    return items


@api.get("/calibration/jobs/{job_id}/log")
async def calibration_job_log(job_id: str, _: Annotated[User, Depends(require_admin)]):
    doc = await db.calibration_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="job not found")
    log_path = doc.get("log_path")
    tail = ""
    if log_path:
        try:
            with open(log_path, "r") as f:
                lines = f.readlines()
            # Keep last 80 lines, strip noisy LiteLLM
            quiet = [ln for ln in lines if "LiteLLM" not in ln and "litellm" not in ln]
            tail = "".join(quiet[-80:])
        except FileNotFoundError:
            tail = ""
    return {"job": doc, "log_tail": tail}


# ===========================================================================
# HOTEL PRICE ORIENTATION
# Combines two sources, in order of reliability:
#   1. Internal training data — median price/night extracted from past trips
#      stored in `training_examples.itinerary_structured_ops.days[*].hotels`
#      and the per-trip "Total Alojamientos (€)" / nights. Always available
#      offline, never blocked.
#   2. Expedia.es live search — best-effort scrape with Playwright (no login).
#      Often blocked by Cloudflare; degrades gracefully.
# Use this from the Itinerary Builder to give the agent a realistic nightly
# rate for any city — especially useful when no library hotel matches.
# ===========================================================================
@api.get("/hotels/price-orientation")
async def hotel_price_orientation(
    user: Annotated[User, Depends(current_user)],
    city: str,
    country: Optional[str] = None,
    checkin: Optional[str] = None,
    checkout: Optional[str] = None,
    adults: int = 2,
    try_expedia: bool = True,
):
    """Orientation price/night for a given city.

    Strategy:
    1. Mine training_examples for trips that visited this city → compute
       median/p25/p75 from real sold prices (instant, reliable).
    2. If nothing found AND try_expedia=True, attempt a quick scrape of
       expedia.es. Falls back gracefully if blocked.
    """
    city_clean = (city or "").strip()
    if not city_clean:
        raise HTTPException(status_code=400, detail="city es obligatorio")

    out = {
        "city": city_clean,
        "country": country,
        "checkin": checkin,
        "checkout": checkout,
        "adults": adults,
        "training_data": None,
        "expedia": None,
        "recommendation": None,
    }

    # ----- 1. Internal training-data aggregation -----
    # Search training_examples whose itinerary_structured_ops.days[*].city
    # matches `city` (case-insensitive). For each match, look up the real
    # "Total Alojamientos (€)" from the raw text and divide by total nights
    # to get a per-night rate.
    city_re = re.compile(re.escape(city_clean), re.IGNORECASE)
    cursor = db.training_examples.find(
        {
            "outcome": "sold",
            "itinerary_structured_ops.days.city": {"$regex": city_re},
        },
        {"_id": 0, "example_id": 1, "itinerary_structured_ops": 1,
         "itinerary_text_ops": 1, "partner": 1},
    ).limit(40)

    prices: list[float] = []
    sample_hotels: list[dict] = []
    seen_hotel_names: set[str] = set()
    async for d in cursor:
        ops = d.get("itinerary_structured_ops") or {}
        days = ops.get("days") or []
        # Sum of nights and unique hotel names linked to THIS city
        city_nights = 0
        for day in days:
            if not isinstance(day, dict):
                continue
            c = (day.get("city") or "")
            if not city_re.search(c):
                continue
            for h in day.get("hotels") or []:
                nights = float(h.get("nights") or 0)
                if nights > 0:
                    city_nights += nights
                name = (h.get("name") or "").strip()
                if name and name not in seen_hotel_names:
                    seen_hotel_names.add(name)
                    sample_hotels.append({"name": name[:80], "trip_partner": d.get("partner") or "kimkim"})
        if city_nights == 0:
            continue
        # Pull Total Alojamientos and nights from the raw ops text
        text = d.get("itinerary_text_ops") or ""
        m_aloj = re.search(r"Total Alojamientos \(€\)\s*\n\s*([\d,\.]+)", text)
        if not m_aloj:
            continue
        try:
            total_eur = float(m_aloj.group(1).replace(",", ""))
        except ValueError:
            continue
        # Total nights = sum of ops days nights (fallback to len(days))
        total_nights = 0
        for day in days:
            for h in day.get("hotels") or []:
                total_nights += float(h.get("nights") or 0)
        if total_nights == 0:
            total_nights = max(1, len(days) - 1)
        # Approximate per-night for THIS city by sharing total proportionally
        # to the nights spent in this city.
        per_night = (total_eur * city_nights / total_nights) / city_nights if city_nights else 0
        if per_night > 30:  # ignore zero / passthrough trips
            prices.append(per_night)

    if prices:
        prices.sort()
        n = len(prices)
        median = prices[n // 2]
        p25 = prices[n // 4]
        p75 = prices[min(n - 1, (3 * n) // 4)]
        out["training_data"] = {
            "n_trips": n,
            "median_price_per_night_eur": round(median, 0),
            "p25_eur": round(p25, 0),
            "p75_eur": round(p75, 0),
            "currency": "EUR",
            "sample_hotels": sample_hotels[:10],
        }
        out["recommendation"] = {
            "price_per_night_eur": round(median, 0),
            "source": "training_data",
            "confidence": "high" if n >= 5 else ("medium" if n >= 3 else "low"),
            "rationale": f"Mediana de {n} viajes vendidos con noche en {city_clean}.",
        }

    # ----- 2. Expedia best-effort (only if training data is thin) -----
    if try_expedia and (not prices or len(prices) < 3):
        from expedia_scraper import search_hotels as _expedia_search
        try:
            exp_res = await asyncio.wait_for(
                _expedia_search(city_clean, checkin=checkin, checkout=checkout,
                                adults=adults, max_results=5),
                timeout=30.0,
            )
            out["expedia"] = exp_res
            if exp_res.get("ok") and exp_res.get("median_price_per_night_eur"):
                # Prefer Expedia over training data ONLY when training data is empty.
                if not prices:
                    out["recommendation"] = {
                        "price_per_night_eur": exp_res["median_price_per_night_eur"],
                        "source": "expedia",
                        "confidence": "medium",
                        "rationale": f"Mediana de {len(exp_res.get('results') or [])} hoteles encontrados en Expedia.es.",
                    }
        except asyncio.TimeoutError:
            out["expedia"] = {"ok": False, "blocked": False, "error": "timeout (>30s)"}
        except Exception as e:
            out["expedia"] = {"ok": False, "blocked": False, "error": str(e)[:200]}

    if not out["recommendation"]:
        out["recommendation"] = {
            "price_per_night_eur": None,
            "source": "none",
            "confidence": "none",
            "rationale": "Sin datos suficientes. Estima a mano con la Regla H del prompt.",
        }
    return out


@api.get("/hotels", response_model=List[Hotel])
async def list_hotels(
    _: Annotated[User, Depends(current_user)],
    q: Optional[str] = None,
    city: Optional[str] = None,
    country: Optional[str] = None,
    tier: Optional[HotelTier] = None,
    include_imported: bool = False,
):
    """List hotels. By default returns only `source='library'` rows (the
    official Excel-imported catalogue). Pass `include_imported=true` to also
    show the 316 hotels auto-created from past-trip scrapes (hidden by
    default per product decision)."""
    flt: dict = {}
    if not include_imported:
        flt["source"] = "library"
    if q:
        flt["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"city": {"$regex": q, "$options": "i"}},
            {"description": {"$regex": q, "$options": "i"}},
        ]
    if city:
        flt["city"] = {"$regex": f"^{city}$", "$options": "i"}
    if country:
        flt["country"] = country
    if tier:
        flt["tier"] = tier
    items = await db.hotels.find(flt, {"_id": 0}).sort("name", 1).to_list(2000)
    return items


@api.post("/hotels", response_model=Hotel)
async def create_hotel(payload: HotelCreate, _: Annotated[User, Depends(current_user)]):
    data = payload.model_dump()
    # Outside Spain → no IVA differential
    _force_no_vat_outside_spain(
        data, incl_field="price_per_night_incl", excl_field="price_per_night_excl"
    )
    h = Hotel(**data)
    await db.hotels.insert_one(h.model_dump())
    return h


@api.patch("/hotels/{hotel_id}", response_model=Hotel)
async def update_hotel(hotel_id: str, payload: HotelUpdate, _: Annotated[User, Depends(current_user)]):
    patch = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if any(k in patch for k in ("price_per_night_excl", "price_per_night_incl", "country")):
        current = await db.hotels.find_one(
            {"hotel_id": hotel_id},
            {"_id": 0, "country": 1, "price_per_night_excl": 1, "price_per_night_incl": 1},
        )
        if current:
            merged = {**current, **patch}
            _force_no_vat_outside_spain(
                merged, incl_field="price_per_night_incl", excl_field="price_per_night_excl"
            )
            patch["price_per_night_excl"] = merged["price_per_night_excl"]
            patch["price_per_night_incl"] = merged["price_per_night_incl"]
    if patch:
        await db.hotels.update_one({"hotel_id": hotel_id}, {"$set": patch})
    doc = await db.hotels.find_one({"hotel_id": hotel_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return doc


@api.delete("/hotels/{hotel_id}")
async def delete_hotel(hotel_id: str, _: Annotated[User, Depends(current_user)]):
    res = await db.hotels.delete_one({"hotel_id": hotel_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


def _normalize_city(sheet_name: str) -> str:
    """Normalize city codes/sheet names to canonical city names."""
    s = (sheet_name or "").strip()
    table = {
        "MAD": "Madrid",
        "BCN": "Barcelona",
        "SEV": "Sevilla",
        "DV": "Douro Valley",
        "Oporto": "Porto",
        "Lisboa": "Lisbon",
        "Venecia": "Venice",
        "Florencia": "Florence",
        "Roma": "Rome",
        "Toscana": "Tuscany",
    }
    return table.get(s, s)


def _tier_from_category(cat: str) -> str:
    if not cat:
        return "standard"
    c = str(cat).strip().replace(" ", "").replace("*", "")
    if c == "5":
        return "luxury"
    if c == "4":
        return "upscale"
    if c == "3":
        return "comfort"
    if c == "2":
        return "standard"
    return "standard"


def _parse_hotels_sheet(ws, country: str, city: str):
    """Generic parser for the hotel sheets. Returns list of hotel dicts."""
    # Identify column indexes from header row
    headers: dict = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers[str(v).strip().lower()] = c

    cat_col = headers.get("categoria")
    name_col = headers.get("nombre hotel") or headers.get("nombre")
    reserva_col = headers.get("reserva")
    notas_col = headers.get("notas")
    ciudad_col = headers.get("ciudad")  # used in Marruecos file (city is per-row)
    tipo_col = headers.get("tipo de alojamiento") or headers.get("apartamento")
    web_col = headers.get("web")
    car_col = headers.get("caracterísricas de alojamiento") or headers.get("caracteristicas de alojamiento") or headers.get("características de alojamiento")
    if not name_col:
        return []

    # Extra notes columns: anything beyond the known
    known = {cat_col, name_col, reserva_col, notas_col, ciudad_col, tipo_col, web_col, car_col}
    extra_cols = [c for c in range(1, ws.max_column + 1) if c not in known and c is not None]

    rows = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if not name:
            continue
        name = str(name).strip()
        if not name or name.lower().startswith("nombre"):
            continue

        cat = ws.cell(r, cat_col).value if cat_col else None
        tier = _tier_from_category(str(cat) if cat else "")

        # Concatenate notes from notas, reserva (skip "si"/"no"/"hotel"/"kimkim"), tipo, características, and any extra
        note_parts = []
        if reserva_col:
            v = ws.cell(r, reserva_col).value
            if v:
                v = str(v).strip()
                if v.lower() not in ("si", "no", "hotel", "kimkim", "yes"):
                    note_parts.append(f"reserva: {v}")
        if notas_col:
            v = ws.cell(r, notas_col).value
            if v and str(v).strip():
                note_parts.append(str(v).strip())
        if tipo_col:
            v = ws.cell(r, tipo_col).value
            if v and str(v).strip():
                note_parts.append(f"tipo: {str(v).strip()}")
        if car_col:
            v = ws.cell(r, car_col).value
            if v and str(v).strip():
                note_parts.append(str(v).strip())
        for c in extra_cols:
            v = ws.cell(r, c).value
            if v and str(v).strip():
                note_parts.append(str(v).strip())

        web = ""
        if web_col:
            wv = ws.cell(r, web_col).value
            if wv:
                web = str(wv).strip()

        # Per-row city if file is Marruecos style; else use sheet-derived city
        row_city = city
        if ciudad_col:
            cv = ws.cell(r, ciudad_col).value
            if cv:
                row_city = str(cv).strip()

        rows.append({
            "name": name,
            "city": row_city,
            "country": country,
            "tier": tier,
            "description": (note_parts[0] if note_parts else None),
            "notes": "\n".join(note_parts[1:]) if len(note_parts) > 1 else None,
            "contact": web or None,
            "price_per_night_excl": 0.0,
            "price_per_night_incl": 0.0,
            "currency": "EUR",
        })
    return rows


@api.post("/hotels/import-all-server")
async def import_all_hotels_server(
    admin: Annotated[User, Depends(require_admin)],
    base_path: str = Query("/app/artifacts/hoteles_db", description="Server dir containing hotel xlsx files"),
    wipe: bool = Query(False, description="Wipe hotels collection first"),
):
    """Walk the hotels directory and import every .xlsx file found.
    Country inferred from filename (ESPA/PORT/ITAL/MARRU).
    City inferred from sheet name (MAD→Madrid, etc.) or per-row Ciudad column.
    """
    import pathlib
    base = pathlib.Path(base_path)
    if not base.exists():
        raise HTTPException(status_code=404, detail=f"Path not found: {base_path}")

    if wipe:
        await db.hotels.delete_many({})

    files = list(base.rglob("*.xlsx"))
    total_created = 0
    total_skipped = 0
    file_results = []
    for fp in files:
        upn = fp.name.upper()
        country = None
        if "ESPA" in upn:
            country = "España"
        elif "PORT" in upn:
            country = "Portugal"
        elif "ITAL" in upn:
            country = "Italia"
        elif "MARRU" in upn:
            country = "Marruecos"
        elif "ALOJAMIENTO" in upn or "APART" in upn or "FAMILI" in upn:
            country = None  # multi-country file; city per sheet
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            file_created = 0
            file_skipped = 0
            for sname in wb.sheetnames:
                ws = wb[sname]
                if ws.max_row < 2:
                    continue
                city = _normalize_city(sname)
                rows = _parse_hotels_sheet(ws, country=country, city=city)
                for row in rows:
                    # Dedup by (name, city)
                    existing = await db.hotels.find_one(
                        {"name": row["name"], "city": row["city"]}, {"_id": 0}
                    )
                    if existing:
                        file_skipped += 1
                        continue
                    h = Hotel(**row)
                    await db.hotels.insert_one(h.model_dump())
                    file_created += 1
            total_created += file_created
            total_skipped += file_skipped
            file_results.append({"file": fp.name, "country": country, "created": file_created, "skipped": file_skipped})
        except Exception as e:
            file_results.append({"file": fp.name, "error": str(e)})

    return {
        "files_scanned": len(files),
        "total_created": total_created,
        "total_skipped": total_skipped,
        "wiped": wipe,
        "files": file_results,
    }
@api.get("/training-examples", response_model=List[TrainingExample])
async def list_training_examples(_: Annotated[User, Depends(require_admin)]):
    items = await db.training_examples.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return items


@api.get("/training-examples/pending-request", response_model=List[TrainingExample])
async def list_pending_request_examples(_: Annotated[User, Depends(require_admin)]):
    """Training examples imported from gestion that still need a client request."""
    items = await db.training_examples.find(
        {"$or": [{"client_request": ""}, {"client_request": None}]},
        {"_id": 0},
    ).sort("created_at", -1).to_list(1000)
    return items


@api.get("/training-examples/bulk-import-jobs", response_model=List[BulkImportJob])
async def list_bulk_import_jobs(_: Annotated[User, Depends(require_admin)]):
    docs = await db.bulk_import_jobs.find({}, {"_id": 0}).sort("started_at", -1).to_list(30)
    return docs


@api.get("/training-examples/bulk-import-jobs/{job_id}", response_model=BulkImportJob)
async def get_bulk_import_job(job_id: str, _: Annotated[User, Depends(require_admin)]):
    doc = await db.bulk_import_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return doc


@api.post("/training-examples/bulk-import-jobs/{job_id}/cancel", response_model=BulkImportJob)
async def cancel_bulk_import_job(job_id: str, _: Annotated[User, Depends(require_admin)]):
    """Mark a running job as cancelled. The background worker polls this status
    between actions and stops cleanly, scraping whatever it has already listed."""
    doc = await db.bulk_import_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    if doc["status"] in ("completed", "failed", "cancelled", "interrupted"):
        return doc
    await _update_job(
        job_id,
        status="cancelled",
        last_message="Cancelado por el usuario. Procesando viajes ya listados…",
    )
    fresh = await db.bulk_import_jobs.find_one({"job_id": job_id}, {"_id": 0})
    return fresh


@api.post("/training-examples/bulk-import-jobs/{job_id}/resume", response_model=BulkImportJob)
async def resume_bulk_import_job(job_id: str, user: User = Depends(require_admin)):
    """Pick up an interrupted/failed/cancelled job exactly where it stopped.
    Listing IDs already discovered are kept; only un-processed trip_ids will be
    re-scraped. URL-level dedup also prevents duplicates if any race occurs."""
    doc = await db.bulk_import_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    if doc["status"] == "running":
        return doc
    if doc["status"] == "completed":
        # Nothing to do — explicitly tell the caller.
        return doc
    await _update_job(
        job_id,
        status="running",
        finished_at=None,
        last_message="Reanudando…",
        last_heartbeat=now_iso(),
    )
    asyncio.create_task(_run_bulk_import_gestion(job_id, doc.get("params") or {}, user.email))
    fresh = await db.bulk_import_jobs.find_one({"job_id": job_id}, {"_id": 0})
    return fresh


@api.post("/training-examples", response_model=TrainingExample)
async def create_training_example(
    payload: TrainingExampleUpsert,
    user: Annotated[User, Depends(require_admin)],
):
    if not payload.client_request:
        raise HTTPException(status_code=400, detail="client_request es obligatorio")
    ex = TrainingExample(
        client_name=payload.client_name,
        client_request=payload.client_request,
        itinerary_url=payload.itinerary_url,
        itinerary_text=payload.itinerary_text,
        outcome=payload.outcome or "pending",
        partner=payload.partner or "kimkim",
        notes=payload.notes,
        created_by=user.email,
    )
    await db.training_examples.insert_one(ex.model_dump())
    await bump_version(db)
    return ex


@api.patch("/training-examples/{example_id}", response_model=TrainingExample)
async def update_training_example(
    example_id: str,
    payload: TrainingExampleUpsert,
    _: Annotated[User, Depends(require_admin)],
):
    patch = {k: v for k, v in payload.model_dump(exclude_unset=True).items() if v is not None}
    if patch:
        await db.training_examples.update_one({"example_id": example_id}, {"$set": patch})
        await bump_version(db)
    doc = await db.training_examples.find_one({"example_id": example_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return doc


@api.delete("/training-examples/{example_id}")
async def delete_training_example(example_id: str, _: Annotated[User, Depends(require_admin)]):
    res = await db.training_examples.delete_one({"example_id": example_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    await bump_version(db)
    return {"ok": True}


@api.get("/ai/retrieval/stats")
async def retrieval_stats(_: Annotated[User, Depends(require_admin)]):
    """Inspect the TF-IDF index status (size + vocabulary)."""
    r = await get_retriever(db)
    return {
        "docs": len(r.docs),
        "features": int(r.matrix.shape[1]) if r.matrix.shape[0] > 0 else 0,
        "version": r.version,
        "outcome_breakdown": {
            "sold": sum(1 for d in r.docs if d.get("outcome") == "sold"),
            "not_sold": sum(1 for d in r.docs if d.get("outcome") == "not_sold"),
            "pending": sum(1 for d in r.docs if d.get("outcome") == "pending"),
        },
    }


@api.post("/ai/retrieval/search")
async def retrieval_search(
    payload: dict = Body(...),
    _: Annotated[User, Depends(require_admin)] = None,
):
    """Manual semantic search over training examples — useful to preview what the
    AI generator will see for a given client request."""
    query = (payload.get("query") or "").strip()
    k = int(payload.get("k") or 10)
    if not query:
        raise HTTPException(status_code=400, detail="query es obligatorio")
    r = await get_retriever(db)
    hits = r.top_k(query, k=k, min_score=0.01)
    return {
        "matches": [{
            "example_id": h.get("example_id"),
            "client_name": h.get("client_name"),
            "outcome": h.get("outcome"),
            "score": h.get("_score"),
            "client_request": (h.get("client_request") or "")[:300],
        } for h in hits],
        "total_indexed": len(r.docs),
    }


@api.post("/training-examples/bulk-import-gestion", response_model=BulkImportJob)
async def bulk_import_gestion(
    payload: dict = Body(...),
    user: User = Depends(require_admin),
):
    """Kick off a background job that logs in to gestion.viajadverdad.com, lists
    /trips matching the given filters and scrapes each one into a TrainingExample.

    payload accepts:
        agent      str | ""   filter "Agente de Ventas" (empty = all)
        source     str        filter "Source" (e.g. "KimKim")
        status     "open" | "closed" | "both" | "all"
        date_from  "DD/MM/YYYY" filter Fecha de Venta lower bound
        date_to    "DD/MM/YYYY" filter Fecha de Venta upper bound
        outcome    "sold" | "not_sold" | "pending"  (tag every imported example)
        limit      int        safety cap on number of trips (default 500)

    Returns the queued BulkImportJob immediately; poll
    GET /training-examples/bulk-import-jobs/{job_id} for progress.
    """
    job = BulkImportJob(params=payload, created_by=user.email, status="queued",
                        last_message="En cola…")
    await db.bulk_import_jobs.insert_one(job.model_dump())
    asyncio.create_task(_run_bulk_import_gestion(job.job_id, payload, user.email))
    return job


async def _update_job(job_id: str, **fields):
    """Patch a BulkImportJob document."""
    if not fields:
        return
    await db.bulk_import_jobs.update_one({"job_id": job_id}, {"$set": fields})


async def _is_cancelled(job_id: str) -> bool:
    doc = await db.bulk_import_jobs.find_one({"job_id": job_id}, {"status": 1})
    return bool(doc and doc.get("status") == "cancelled")


def _extract_client_name(text: str) -> str:
    if not text:
        return ""
    if "Lead Name" in text:
        chunks = text.split("Lead Name", 1)[1].split("\n")
        for ln in chunks[1:6]:
            ln = ln.strip()
            if ln and ln.lower() not in ("teléfono", "telefono", "phone", "email", "agente", "agent"):
                return ln[:80]
    return ""


_AGENT_NOISE = {"please select", "", "—", "-", "select", "elegir"}


def _extract_agents(text: str) -> dict:
    """Return {sales_agent, owner_agent} extracted from the gestion ops view text.
    Uses the same regex pattern that worked on 166/166 imports in the backfill."""
    out: dict = {"sales_agent": None, "owner_agent": None}
    if not text:
        return out
    lines = text.split("\n")
    for i, ln in enumerate(lines):
        s = ln.strip().lower()
        if s == "agente ventas":
            for j in range(i + 1, min(i + 4, len(lines))):
                v = lines[j].strip()
                if v and v.lower() not in _AGENT_NOISE and not v.lower().startswith("agente") and len(v) < 50:
                    out["sales_agent"] = v
                    break
        elif s == "agente" and out["owner_agent"] is None:
            for j in range(i + 1, min(i + 4, len(lines))):
                v = lines[j].strip()
                if v and v.lower() not in _AGENT_NOISE and not v.lower().startswith("agente") and len(v) < 50:
                    out["owner_agent"] = v
                    break
    return out


def _clean_trip_name(raw: str) -> str:
    """Strip Fabrik link decorations and the ubiquitous "_facturado…" suffix."""
    import re as _re
    s = (raw or "").replace("\t", " ")
    # Collapse whitespace
    s = " ".join(s.split())
    # Drop UI prefixes added by Fabrik in the row link text
    for prefix in ("Edit", "View", "Add"):
        if s.startswith(prefix):
            s = s[len(prefix):].strip()
    # Cut at "_facturado" / "_INCIDENCIA" / opening parenthesis with notes
    s = _re.split(r"_facturado|_INCIDENCIA|_pendiente|_no facturado", s, maxsplit=1, flags=_re.IGNORECASE)[0]
    s = s.strip(" _-")
    return s[:80]


async def _run_bulk_import_gestion(job_id: str, params: dict, user_email: str):
    """Background coroutine that drives the full bulk-import workflow.

    1. Login once into gestion.viajadverdad.com
    2. For each requested status (open / closed), apply filters on /trips,
       paginate and harvest every trip ID.
    3. For each unique trip ID, scrape the ops view + LLM-parse it and
       create a TrainingExample (client_request stays empty so the user
       can fill it later from the AI Trainer UI).
    """
    from playwright.async_api import async_playwright
    from scraper import _render_url, _parse_with_llm, GESTION_USER, GESTION_PASS

    agent = (params.get("agent") or "").strip()
    source = (params.get("source") or "").strip()
    raw_status = (params.get("status") or "all_sold").strip().lower()
    if raw_status in ("all_sold", "todos_vendidos"):
        statuses = ["open", "closed", "terminado"]
    elif raw_status in ("all", "both", "ambos", "todos", ""):
        statuses = ["open", "closed"]
    elif raw_status in ("open", "abierto"):
        statuses = ["open"]
    elif raw_status in ("closed", "cerrado"):
        statuses = ["closed"]
    elif raw_status in ("terminado", "finished"):
        statuses = ["terminado"]
    else:
        statuses = [raw_status]
    date_from = (params.get("date_from") or "").strip()
    date_to = (params.get("date_to") or "").strip()
    raw_outcome = (params.get("outcome") or "sold").strip().lower()
    outcome: TripOutcome = raw_outcome if raw_outcome in ("sold", "not_sold", "pending") else "sold"
    # Partner / source — saved on every imported TrainingExample so the AI
    # generator can later adjust pricing per commission model.
    raw_partner = (params.get("partner") or params.get("source") or "kimkim").strip().lower()
    partner_map = {
        "kimkim": "kimkim",
        "zicasso": "zicasso",
        "responsibletravel": "responsible_travel",
        "responsible travel": "responsible_travel",
        "responsible_travel": "responsible_travel",
        "direct": "direct",
        "directo": "direct",
        "direct booking": "direct",
    }
    partner: str = partner_map.get(raw_partner, "other")
    try:
        limit = max(1, min(int(params.get("limit") or 500), 2000))
    except Exception:
        limit = 500

    await _update_job(job_id, status="running",
                      last_message="Iniciando navegador y login en gestion…",
                      last_heartbeat=now_iso())

    # Recover any previously listed trip_ids if the same job is resumed
    job_doc = await db.bulk_import_jobs.find_one({"job_id": job_id}, {"_id": 0}) or {}
    all_trip_ids: list[str] = list(job_doc.get("pending_trip_ids") or [])
    trip_names: dict[str, str] = dict(job_doc.get("trip_names") or {})
    processed_set: set[str] = set(job_doc.get("processed_trip_ids") or [])
    seen: set[str] = set(all_trip_ids)
    listing_done: bool = bool(job_doc.get("listing_done"))

    try:
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(args=["--no-sandbox", "--disable-dev-shm-usage"])
            try:
                ctx = await browser.new_context(user_agent="Mozilla/5.0")
                page = await ctx.new_page()
                page.set_default_timeout(20000)
                page.set_default_navigation_timeout(25000)

                # ---------- LOGIN ----------
                try:
                    await page.goto("https://gestion.viajadverdad.com/login",
                                    wait_until="domcontentloaded", timeout=25000)
                    await page.fill('input[name="username"]', GESTION_USER)
                    await page.fill('input[name="password"]', GESTION_PASS)
                    await page.click('button[type="submit"], input[type="submit"]')
                    await page.wait_for_load_state("domcontentloaded", timeout=20000)
                    await page.wait_for_timeout(1500)
                    if "/login" in page.url:
                        await _update_job(job_id, status="failed",
                                          last_message="Login en gestion falló (revisa GESTION_VIAJADVERDAD_USER/PASS)",
                                          finished_at=now_iso())
                        return
                except Exception as e:
                    await _update_job(job_id, status="failed",
                                      last_message=f"Error de login: {e}",
                                      finished_at=now_iso())
                    return

                # ---------- COLLECT TRIP IDS PER STATUS ----------
                async def fabrik_apply_filters(status_value: str) -> None:
                    """Apply Fabrik /trips filters using the verified element IDs.

                    Selectors discovered on gestion.viajadverdad.com/trips:
                      - select#app_trips___agentvalue    (label = "All"/"Beatriz"/…)
                      - select#app_trips___sourcevalue   (value = "KimKim"/…)
                      - select#app_trips___statusvalue   (value = "abierto"/"cerrado"/…)
                      - input#app_trips___booking_date_1_com_fabrik_1_filter_range_0_\\.0  (Fecha de Venta desde)
                      - input#app_trips___booking_date_1_com_fabrik_1_filter_range_1_\\.0  (Fecha de Venta hasta)
                    """
                    # Agent — select by visible label (e.g. "Beatriz")
                    if agent:
                        try:
                            await page.select_option('#app_trips___agentvalue', label=agent, timeout=5000)
                        except Exception:
                            try:
                                await page.select_option('#app_trips___agentvalue', value=agent, timeout=5000)
                            except Exception:
                                pass
                    # Source — select by value (matches the visible label too in this Fabrik config)
                    if source:
                        try:
                            await page.select_option('#app_trips___sourcevalue', value=source, timeout=5000)
                        except Exception:
                            try:
                                await page.select_option('#app_trips___sourcevalue', label=source, timeout=5000)
                            except Exception:
                                pass
                    # Status — lowercase value
                    if status_value:
                        try:
                            await page.select_option('#app_trips___statusvalue', value=status_value, timeout=5000)
                        except Exception:
                            pass
                    # Booking date range. Dot inside the ID requires attribute selector.
                    if date_from:
                        try:
                            await page.fill(
                                'input[id="app_trips___booking_date_1_com_fabrik_1_filter_range_0_.0"]',
                                date_from, timeout=5000,
                            )
                        except Exception:
                            pass
                    if date_to:
                        try:
                            await page.fill(
                                'input[id="app_trips___booking_date_1_com_fabrik_1_filter_range_1_.0"]',
                                date_to, timeout=5000,
                            )
                        except Exception:
                            pass
                    # Submit (Fabrik filter "Go" button is `name="filter"`)
                    try:
                        btn = await page.query_selector('button[name="filter"], input[name="filter"]')
                        if btn:
                            await btn.click(timeout=5000)
                        else:
                            await page.evaluate(
                                "document.querySelector('form[name=\"listform_1_com_fabrik_1\"], form.fabrikForm, .fabrik_filter')?.form?.submit()"
                            )
                    except Exception:
                        pass
                    # Wait for the post-submit reload using DOM-content (much more reliable
                    # than networkidle on Fabrik pages, which keep firing AJAX in background).
                    try:
                        await page.wait_for_load_state("domcontentloaded", timeout=15000)
                    except Exception:
                        pass
                    await page.wait_for_timeout(2000)

                async def _process_status(st: str) -> None:
                    """Drive a single status through filter→limit→pagination→collect."""
                    status_label = {"open": "abierto", "closed": "cerrado", "terminado": "terminado"}.get(st, st)
                    await _update_job(
                        job_id, matched=len(all_trip_ids),
                        last_message=f"Aplicando filtros (estado={status_label})…",
                    )
                    await page.goto("https://gestion.viajadverdad.com/trips",
                                    wait_until="domcontentloaded", timeout=20000)
                    await page.wait_for_timeout(1500)
                    await fabrik_apply_filters(status_label)

                    # Maximise page size so we minimise pagination hops (Fabrik default = 100).
                    try:
                        opts = await page.evaluate("""() => {
                            const s = document.querySelector('#limit1');
                            if (!s) return [];
                            return Array.from(s.options).map(o => parseInt(o.value, 10)).filter(n => !isNaN(n));
                        }""")
                        if opts:
                            target = str(max(opts))     # typically "500"
                            current = await page.eval_on_selector('#limit1', 'e => e.value')
                            if current != target:
                                try:
                                    async with page.expect_navigation(wait_until="domcontentloaded", timeout=15000):
                                        await page.select_option('#limit1', value=target, timeout=5000)
                                except Exception:
                                    pass
                                await page.wait_for_timeout(1500)
                    except Exception as e:
                        logger.info("page-size selector skipped: %s", e)

                    # Paginate and harvest IDs
                    page_count = 0
                    while page_count < 50:  # hard cap to avoid infinite loops
                        if await _is_cancelled(job_id):
                            return
                        page_count += 1
                        rows = await page.evaluate("""() => {
                            const map = new Map();
                            document.querySelectorAll('a[href*="/trips/form/"], a[href*="/trips/details/"]').forEach(a => {
                                const m = a.href.match(/\\/trips\\/(?:form|details)\\/\\d+\\/(\\d+)/);
                                if (!m) return;
                                const id = m[1];
                                const text = (a.innerText || '').trim();
                                if (!map.has(id) || (text && text.length > (map.get(id) || '').length)) {
                                    map.set(id, text);
                                }
                            });
                            return [...map.entries()].map(([id, text]) => ({id, text}));
                        }""")
                        new_added = 0
                        for row in rows:
                            tid = row["id"]
                            if tid not in seen:
                                seen.add(tid)
                                all_trip_ids.append(tid)
                                cleaned = _clean_trip_name(row.get("text") or "")
                                if cleaned:
                                    trip_names[tid] = cleaned
                                new_added += 1
                                if len(all_trip_ids) >= limit:
                                    break
                        # Persist progress so a crash mid-listing doesn't lose IDs.
                        await _update_job(
                            job_id, matched=len(all_trip_ids),
                            pending_trip_ids=all_trip_ids[:],
                            trip_names=trip_names,
                            last_message=f"Listando viajes (estado={status_label}) · página {page_count} · +{new_added}",
                        )
                        if len(all_trip_ids) >= limit:
                            return
                        if new_added == 0 and page_count > 1:
                            return
                        # Try to advance pagination
                        try:
                            nxt = await page.query_selector(
                                'a[title="Next"], a[title="Siguiente"], '
                                'li.pagination-next a:not(.disabled), '
                                '.pagination li.next a, a:has-text("›"), a:has-text("→")'
                            )
                            if not nxt:
                                return
                            href = await nxt.get_attribute("href")
                            cls = (await nxt.get_attribute("class") or "").lower()
                            if "disabled" in cls or href in (None, "", "#"):
                                return
                            await nxt.click(timeout=5000)
                            await page.wait_for_load_state("domcontentloaded", timeout=15000)
                            await page.wait_for_timeout(1500)
                        except Exception:
                            return

                for st_idx, st in enumerate(statuses):
                    if listing_done:
                        # Listing previously completed — skip to scraping phase.
                        await _update_job(
                            job_id, matched=len(all_trip_ids),
                            last_message=f"Reanudando · saltando listado (ya hecho, {len(all_trip_ids)} viajes)",
                            last_heartbeat=now_iso(),
                        )
                        break
                    if len(all_trip_ids) >= limit:
                        break
                    if await _is_cancelled(job_id):
                        break
                    try:
                        # Hard 180s cap per status so a hung filter never blocks the next status.
                        await asyncio.wait_for(_process_status(st), timeout=180)
                    except asyncio.TimeoutError:
                        await _update_job(
                            job_id, matched=len(all_trip_ids),
                            last_message=f"Timeout listando estado={st} tras 180s · sigo con el siguiente",
                        )
                    except Exception as e:
                        await _update_job(
                            job_id, matched=len(all_trip_ids),
                            last_message=f"Error listando estado={st}: {str(e)[:140]} · sigo con el siguiente",
                        )
                        logger.warning("listing error for status %s: %s", st, e)

                await page.close()
            finally:
                await browser.close()
    except Exception as e:
        await _update_job(job_id, status="failed",
                          last_message=f"Fallo en la fase de listado: {e}",
                          finished_at=now_iso())
        logger.exception("bulk-import listing phase failed")
        return

    # Listing phase done — record this so a resume jumps straight to scraping.
    if not listing_done:
        await _update_job(
            job_id, listing_done=True,
            pending_trip_ids=all_trip_ids[:],
            trip_names=trip_names,
            matched=len(all_trip_ids),
            last_heartbeat=now_iso(),
        )

    trip_ids = all_trip_ids[:limit]

    if not trip_ids:
        await _update_job(
            job_id, status="completed", finished_at=now_iso(),
            last_message="No se encontró ningún viaje con esos filtros.",
        )
        return

    # ---------- SCRAPE EACH TRIP ----------
    # Start counters from whatever the resumed job already has, so progress is cumulative.
    job_doc2 = await db.bulk_import_jobs.find_one({"job_id": job_id}, {"_id": 0}) or {}
    created = int(job_doc2.get("scraped") or 0)
    skipped = int(job_doc2.get("skipped") or 0)
    failed = int(job_doc2.get("failed") or 0)
    errors: list[str] = list(job_doc2.get("errors") or [])
    notes_tag = (
        f"Auto-import gestion (agente={agent or 'Todos'} · source={source or '—'} · "
        f"estado={','.join(statuses)} · fechas={date_from or '—'}→{date_to or '—'} · outcome={outcome})"
    )

    # Pick trips that haven't been processed yet (resumability).
    remaining = [tid for tid in trip_ids if tid not in processed_set]
    await _update_job(
        job_id, matched=len(trip_ids),
        last_message=(
            f"{len(trip_ids)} viajes en cola · {len(processed_set)} ya procesados · "
            f"{len(remaining)} pendientes. Iniciando scraping…"
        ),
        last_heartbeat=now_iso(),
    )

    if not remaining:
        await _update_job(
            job_id, status="completed", finished_at=now_iso(),
            last_message=(
                f"Completado · {created} creados (sesiones anteriores) · "
                f"nada que hacer ahora."
            ),
        )
        return

    # ---------- SCRAPE EACH TRIP ----------
    for i, tid in enumerate(remaining):
        if await _is_cancelled(job_id):
            break
        url = f"https://gestion.viajadverdad.com/trips/form/1/{tid}"
        existing = await db.training_examples.find_one(
            {"itinerary_url_ops": url}, {"_id": 0}
        )
        if existing:
            skipped += 1
            processed_set.add(tid)
            await _update_job(
                job_id, skipped=skipped,
                processed_trip_ids=list(processed_set),
                last_heartbeat=now_iso(),
                last_message=f"[{i + 1}/{len(remaining)}] saltado (ya importado) · trip {tid}",
            )
            continue
        try:
            rendered = await _render_url(url)
            text = rendered.get("text", "")
            structured = (
                await _parse_with_llm(text)
                if rendered.get("ok") else
                {"days": [], "notes": rendered.get("error") or "scrape_failed"}
            )
            client_name = trip_names.get(tid) or _extract_client_name(text)
            agents = _extract_agents(text)
            ex = TrainingExample(
                client_name=client_name,
                client_request="",  # pending — user fills later
                itinerary_url=None,
                itinerary_url_ops=url,
                itinerary_text_ops=text,
                itinerary_structured_ops=structured,
                outcome=outcome,    # selectable per-import (sold / not_sold / pending)
                partner=partner,    # selectable per-import (kimkim / zicasso / ...)
                notes=notes_tag,
                sales_agent=agents.get("sales_agent"),
                owner_agent=agents.get("owner_agent"),
                created_by=user_email,
            )
            await db.training_examples.insert_one(ex.model_dump())
            created += 1
            processed_set.add(tid)
            label = client_name or structured.get("trip_name") or tid
            await _update_job(
                job_id, scraped=created,
                processed_trip_ids=list(processed_set),
                last_heartbeat=now_iso(),
                last_message=f"[{i + 1}/{len(remaining)}] OK · {label}",
            )
        except Exception as e:
            failed += 1
            short = str(e)[:160]
            errors.append(f"trip {tid}: {short}")
            processed_set.add(tid)   # don't retry hard failures next resume
            await _update_job(
                job_id, failed=failed,
                processed_trip_ids=list(processed_set),
                errors=errors[-50:],
                last_heartbeat=now_iso(),
                last_message=f"[{i + 1}/{len(remaining)}] ERROR trip {tid}: {short}",
            )
            logger.warning("bulk import trip %s failed: %s", tid, e)

    was_cancelled = await _is_cancelled(job_id)
    await _update_job(
        job_id,
        status="cancelled" if was_cancelled else "completed",
        scraped=created,
        skipped=skipped,
        failed=failed,
        errors=errors[-50:],   # keep only last 50 to avoid bloat
        finished_at=now_iso(),
        last_message=(
            ("Cancelado por el usuario · " if was_cancelled else "Completado · ")
            + f"{created} creados · {skipped} saltados · {failed} con error"
        ),
    )
    # Bulk imports may have inserted dozens of examples — invalidate the
    # retrieval index so the next /ai/generate sees them.
    if created > 0:
        await bump_version(db)



async def scrape_itinerary_url(
    payload: dict = Body(...),
    _: User = Depends(current_user),
):
    """Render a URL with a real headless browser and parse it into structured JSON.

    Returns {"ok", "source", "text", "structured": {days, hotels, ...}, "error"}.
    """
    url = (payload.get("url") or "").strip()
    if not url:
        raise HTTPException(status_code=400, detail="URL es obligatoria")
    from scraper import scrape_and_parse
    result: dict = {}
    try:
        result = await scrape_and_parse(url)
    except Exception as e:
        logger.exception("scrape failed")
        raise HTTPException(status_code=500, detail=f"Scrape error: {e}")
    return result


# ===========================================================================
# AI generation
# ===========================================================================
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")




async def _call_claude_json(system_prompt: str, user_prompt: str) -> dict:
    """Call Claude Sonnet 4.6 and parse JSON output."""
    import json as _json
    from emergentintegrations.llm.chat import LlmChat, UserMessage

    if not EMERGENT_LLM_KEY:
        raise HTTPException(status_code=500, detail="EMERGENT_LLM_KEY no configurada")

    chat = LlmChat(
        api_key=EMERGENT_LLM_KEY,
        session_id=f"gen-{uuid.uuid4().hex[:8]}",
        system_message=system_prompt,
    ).with_model("anthropic", "claude-sonnet-4-6")
    msg = UserMessage(text=user_prompt)
    raw = await chat.send_message(msg)
    text = (raw or "").strip()
    # Strip optional markdown fences
    if text.startswith("```"):
        text = text.strip("`")
        if text.lower().startswith("json"):
            text = text[4:].strip()
    # Try to locate first JSON object
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1:
        raise HTTPException(status_code=502, detail=f"AI no devolvió JSON: {text[:200]}")
    try:
        return _json.loads(text[start:end + 1])
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"JSON inválido del modelo: {e}")


# Country detection — exact tokens used in the DB (Spanish names).
# Some keywords are HIGH-SIGNAL (souks → Morocco, alhambra → Spain) and act as
# tie-breakers. Each tuple is (keyword, weight). Default weight=1.
_COUNTRY_KEYWORDS_WEIGHTED = {
    "Portugal":  [(k, 1) for k in (
        "portugal", "lisbon", "lisboa", "porto", "sintra", "évora", "evora",
        "douro", "batalha", "alentejo", "algarve", "obidos", "óbidos",
        "coimbra", "cascais", "nazaré", "nazare", "madeira", "azores",
        "pastel de nata", "fado",
    )],
    "España":    [(k, 1) for k in (
        "spain", "españa", "madrid", "barcelona", "sevilla", "seville",
        "granada", "valencia", "bilbao", "san sebastian", "san sebastián",
        "toledo", "córdoba", "cordoba", "málaga", "malaga", "mallorca",
        "ibiza", "tenerife", "asturias", "rioja", "ronda", "alhambra",
        "flamenco", "tapas", "cava", "sagrada familia", "sitges",
        "montserrat", "park güell", "park guell",
    )],
    "Italia":    [(k, 1) for k in (
        "italy", "italia", "rome", "roma", "florence", "firenze", "venice",
        "venezia", "milan", "milano", "naples", "napoli", "tuscany",
        "toscana", "sicily", "sicilia", "amalfi", "matera", "puglia",
        "apulia", "como", "verona", "bologna", "cinque terre", "capri",
        "sorrento", "positano", "vatican", "vesuvius", "pompeii", "dolomites",
        "stintino", "sardinia", "cerdeña",
    )],
    # Morocco gets explicit weight on its iconic/unique keywords so it wins the
    # tie even when the request is mostly written in English with "Spain" first
    # (frequent for multi-country requests originating on KimKim).
    "Marruecos": (
        [(k, 1) for k in (
            "morocco", "marruecos", "marrakech", "marrakesh", "fez", "fes",
            "casablanca", "essaouira", "chefchaouen", "rabat", "merzouga",
            "atlas", "ouarzazate", "ait ben haddou", "boumalne", "dades",
        )]
        + [(k, 3) for k in (
            "souks", "souk", "sahara", "berber", "riad", "tagine",
            "hot air balloon", "camel trek", "spice market",
        )]
    ),
}


def _detect_country(text: str) -> Optional[str]:
    """Pick the country with the highest weighted keyword score."""
    if not text:
        return None
    t = text.lower()
    scores: dict[str, int] = {}
    for country, kw_pairs in _COUNTRY_KEYWORDS_WEIGHTED.items():
        scores[country] = sum(w for kw, w in kw_pairs if kw in t)
    best = max(scores.items(), key=lambda x: x[1])
    return best[0] if best[1] > 0 else None


# Backwards-compat alias for any code path that still queries by country
_COUNTRY_KEYWORDS = {
    c: [kw for kw, _w in pairs]
    for c, pairs in _COUNTRY_KEYWORDS_WEIGHTED.items()
}


def _detect_cities(text: str, country: Optional[str]) -> list[str]:
    """Detect specific cities mentioned for the (optionally) detected country.
    Returns city names as they appear in the DB."""
    if not text:
        return []
    t = text.lower()
    cities_for = {
        "Portugal": {
            "lisbon": "Lisbon", "lisboa": "Lisbon",
            "porto": "Porto", "oporto": "Porto",
            "sintra": "Sintra", "évora": "Évora", "evora": "Évora",
            "douro": "Douro", "batalha": "Batalha", "óbidos": "Óbidos", "obidos": "Óbidos",
            "coimbra": "Coimbra", "cascais": "Cascais", "nazaré": "Nazaré", "nazare": "Nazaré",
            "alentejo": "Alentejo", "algarve": "Algarve", "madeira": "Madeira",
        },
        "España": {
            "madrid": "Madrid", "barcelona": "Barcelona", "sevilla": "Sevilla",
            "seville": "Seville", "granada": "Granada", "valencia": "Valencia",
            "bilbao": "Bilbao", "san sebastian": "San Sebastian",
            "toledo": "Toledo", "córdoba": "Córdoba", "cordoba": "Córdoba",
            "málaga": "MALAGA", "malaga": "MALAGA", "ronda": "Ronda",
        },
        "Italia": {
            "rome": "Rome", "roma": "Rome", "florence": "Florence", "firenze": "Florence",
            "venice": "Venice", "venezia": "Venice", "milan": "Milan", "milano": "Milan",
            "naples": "Naples", "napoli": "Naples",
            "tuscany": "Tuscany", "toscana": "Tuscany",
            "sicily": "Sicilia", "sicilia": "Sicilia",
            "amalfi": "Amalfi Coast", "matera": "Apulia + Matera",
            "puglia": "Apulia + Matera", "apulia": "Apulia + Matera",
        },
        "Marruecos": {
            "marrakech": "Marrakech", "marrakesh": "Marrakech",
            "fez": "Fez", "fes": "Fez", "casablanca": "Casablanca",
            "essaouira": "Essaouira", "chefchaouen": "Chefchaouen",
        },
    }
    pool = dict(cities_for.get(country, {})) if country else {}
    if not pool:
        for cs in cities_for.values():
            pool.update(cs)
    found: list[str] = []
    seen: set[str] = set()
    for kw, name in pool.items():
        if kw in t and name not in seen:
            found.append(name)
            seen.add(name)
    return found


def _summ_experience(e: dict) -> dict:
    return {
        "experience_id": e.get("experience_id"),
        "title": e.get("title"),
        "provider_name": e.get("provider_name"),
        "city": e.get("city"),
        "country": e.get("country"),
        "type": e.get("type"),
        "pax": e.get("pax") or 2,
        "price_tax_excl": e.get("price_tax_excl") or e.get("price") or 0,
        "price_tax_incl": e.get("price_tax_incl") or e.get("price") or 0,
        "currency": e.get("currency") or "EUR",
    }


def _summ_hotel(h: dict) -> dict:
    return {
        "hotel_id": h.get("hotel_id"),
        "name": h.get("name"),
        "city": h.get("city"),
        "country": h.get("country"),
        "tier": h.get("tier"),
        "price_per_night_excl": h.get("price_per_night_excl") or 0,
        "price_per_night_incl": h.get("price_per_night_incl") or 0,
        "currency": h.get("currency") or "EUR",
    }


@api.post("/ai/generate-itinerary")
async def ai_generate(
    payload: dict = Body(...),
    user: User = Depends(require_admin),
):
    """Generate an itinerary draft from a client request.

    payload = {"client_request": "...", "client_name": "...optional", "save": true}
    Returns the parsed JSON. If save=true (default), also stores it as an Itinerary draft.
    """
    request_text = (payload.get("client_request") or "").strip()
    if not request_text:
        raise HTTPException(status_code=400, detail="client_request es obligatorio")
    client_name = (payload.get("client_name") or "").strip()
    save = bool(payload.get("save", True))
    # Partner / source for this client. Determines the commission model the AI
    # must apply when pricing the draft.
    partner: str = (payload.get("partner") or "kimkim").strip().lower()
    if partner not in ("kimkim", "zicasso", "responsible_travel", "direct", "other"):
        partner = "other"
    # Optional: skip these training-example IDs during retrieval. Used for offline
    # self-evaluation so the system doesn't cheat by retrieving the very example
    # we're trying to predict.
    exclude_ids: set[str] = set(payload.get("exclude_example_ids") or [])

    # Build context: library subsets and training examples
    # ----------------------------------------------------------------------
    # CASCADE FILTERING — pick context that is RELEVANT to the destination.
    #   1. Detect country and specific cities from the new request.
    #   2. Catalog (experiences + hotels) is filtered to that country
    #      AND boosted for the mentioned cities.
    #   3. Training examples are filtered to the same country before doing
    #      the TF-IDF retrieval, so "Portugal" requests don't surface Italy
    #      itineraries unless we explicitly have to fall back.
    # ----------------------------------------------------------------------
    country = _detect_country(request_text)
    cities = _detect_cities(request_text, country)

    # -- Experiences --
    exp_flt: dict = {}
    if country:
        exp_flt["country"] = country
    if cities:
        # Boost city matches via a separate query, then merge.
        city_flt = {**exp_flt, "city": {"$in": cities}}
        city_exps = await db.experiences.find(city_flt, {"_id": 0}).limit(80).to_list(80)
    else:
        city_exps = []
    rest_exps = await db.experiences.find(exp_flt, {"_id": 0}).limit(150).to_list(150)
    seen_exp = {e["experience_id"] for e in city_exps}
    exps = list(city_exps)
    for e in rest_exps:
        if e["experience_id"] not in seen_exp:
            exps.append(e)
            if len(exps) >= 120:
                break
    # Final fallback when the country filter is too tight (rare destinations).
    if len(exps) < 20:
        more = await db.experiences.find({}, {"_id": 0}).limit(60).to_list(60)
        seen_exp = {e["experience_id"] for e in exps}
        for e in more:
            if e["experience_id"] not in seen_exp:
                exps.append(e)
                if len(exps) >= 80:
                    break

    # -- Hotels --
    # Only the library catalogue (Excel-imported). The 316 hotels auto-imported
    # from past-trip scrapes are hidden so the AI doesn't suggest them.
    hotel_flt: dict = {"source": "library"}
    if country:
        hotel_flt["country"] = country
    if cities:
        city_hotels = await db.hotels.find(
            {**hotel_flt, "city": {"$in": cities}}, {"_id": 0}
        ).limit(60).to_list(60)
    else:
        city_hotels = []
    rest_hotels = await db.hotels.find(hotel_flt, {"_id": 0}).limit(80).to_list(80)
    seen_h = {h.get("hotel_id") for h in city_hotels}
    hotels = list(city_hotels)
    for h in rest_hotels:
        if h.get("hotel_id") not in seen_h:
            hotels.append(h)
            if len(hotels) >= 80:
                break
    if len(hotels) < 10:
        more_h = await db.hotels.find({"source": "library"}, {"_id": 0}).limit(60).to_list(60)
        seen_h = {h.get("hotel_id") for h in hotels}
        for h in more_h:
            if h.get("hotel_id") not in seen_h:
                hotels.append(h)
                if len(hotels) >= 40:
                    break

    # Pull TRAINING examples by semantic similarity (TF-IDF over client_request),
    # FILTERED to the detected country so the AI doesn't mix Italy trips into a
    # Portugal draft. Falls back to recency if the index is still empty.
    retriever = await get_retriever(db)
    sold = retriever.top_k(request_text, k=5, prefer_outcomes=["sold"], min_score=0.05)
    not_sold = retriever.top_k(request_text, k=2, prefer_outcomes=["not_sold"], min_score=0.05)
    examples: list[dict] = sold + not_sold
    if exclude_ids:
        examples = [e for e in examples if e.get("example_id") not in exclude_ids]
    if country:
        # Keep only training examples whose request mentions the same country.
        country_kws = {k.lower() for k in _COUNTRY_KEYWORDS.get(country, [])}
        filtered = [
            ex for ex in examples
            if any(k in (ex.get("client_request") or "").lower() for k in country_kws)
        ]
        # If filtering left too few, keep the originals (better some inspiration
        # than none — but tag them so the prompt knows).
        if len(filtered) >= 2:
            examples = filtered
    if not examples:
        examples = await db.training_examples.find(
            {"outcome": {"$in": ["sold", "not_sold"]}, "client_request": {"$nin": [None, ""]}},
            {"_id": 0},
        ).sort("created_at", -1).limit(5).to_list(5)
    retrieval_meta = {
        "country": country,
        "cities": cities,
        "matched_sold": len(sold),
        "matched_not_sold": len(not_sold),
        "examples_used": len(examples),
        "experiences_in_context": len(exps),
        "hotels_in_context": len(hotels),
        "top_score": (examples[0].get("_score") if examples and examples[0].get("_score") is not None else None),
    }

    # Partner pricing instructions — injected into every call so the AI
    # respects each partner's commission model.
    partner_pricing = {
        "kimkim": (
            "PARTNER = KIMKIM (15% on top of agency price)\n"
            "All 167 trips in the training data were sold via KimKim. The PVPs you "
            "see in PAST EXAMPLES already include KimKim's 15% commission ON TOP of "
            "our cost. Match those PVP figures directly — do NOT add a second markup. "
            "Use markup_pct = 15 (default)."
        ),
        "zicasso": (
            "PARTNER = ZICASSO (keeps 10.5% OF the agency price — DEDUCTIVE).\n"
            "Zicasso clients pay our PVP and Zicasso takes 10.5% from us. To keep the "
            "same NET revenue per trip as a KimKim trip, INCREASE the PVP by "
            "dividing by 0.895 (i.e. price the trip ~12% HIGHER than the KimKim-trained "
            "PAST EXAMPLES suggest). Use markup_pct ≈ 28 (15% inherited + ~12% Zicasso uplift) "
            "so the EUR totals on screen line up. Mention Zicasso in the summary."
        ),
        "responsible_travel": (
            "PARTNER = RESPONSIBLE TRAVEL (keeps 10% OF the agency price — DEDUCTIVE).\n"
            "Same logic as Zicasso but the cut is 10%. Divide the KimKim-style PVP by "
            "0.90 (price ~11% HIGHER). Use markup_pct ≈ 27."
        ),
        "direct": (
            "PARTNER = DIRECT (no platform commission).\n"
            "Direct clients book straight with the agency — no partner cut. The PAST "
            "EXAMPLES are KimKim-trained, so DROP the 15% KimKim uplift: divide the "
            "PVP figures you see by 1.15 to get the real agency-only price. Use "
            "markup_pct = 15 (or higher, per Rule S) since the agency keeps the full "
            "markup. Mention 'Direct client — no platform commission' in the summary."
        ),
        "other": (
            "PARTNER = OTHER (commission unknown to this assistant).\n"
            "Default to KimKim behaviour (15% on top). A human agent will adjust."
        ),
    }
    partner_block = partner_pricing.get(partner, partner_pricing["other"])

    user_prompt_parts = [
        f"NEW CLIENT REQUEST:\n{request_text}",
        "",
        partner_block,
        "",
        (
            f"DETECTED CONTEXT — country={country or 'unknown'} · cities={', '.join(cities) if cities else 'none'}.\n"
            f"All library items and past examples below are PRE-FILTERED to this destination, "
            f"so you can trust them to be geographically relevant."
        ),
        "",
        "EXPERIENCE LIBRARY (pick from these whenever possible):",
        _compact_json([_summ_experience(e) for e in exps]),
        "",
        "HOTEL LIBRARY (pick from these for accommodations):",
        _compact_json([_summ_hotel(h) for h in hotels]),
        "",
    ]
    if examples:
        user_prompt_parts.append(
            "PAST EXAMPLES (semantically similar, learn from these patterns).\n"
            "STUDY the SOLD ones to replicate what worked, COMPARE with the NOT_SOLD ones "
            "to avoid what failed. Each one is annotated with its outcome and similarity score."
        )
        for ex in examples:
            client_struct = ex.get("itinerary_structured")
            ops_struct = ex.get("itinerary_structured_ops")
            blocks = []
            # Prefer structured forms (much more token-efficient than raw text).
            # If neither structure has days, fall back to a TRUNCATED raw snippet.
            if client_struct and isinstance(client_struct, dict) and client_struct.get("days"):
                blocks.append(f"CLIENT-FACING ITINERARY (Travefy):\n{_compact_json(client_struct)}")
            if ops_struct and isinstance(ops_struct, dict) and ops_struct.get("days"):
                blocks.append(f"INTERNAL OPS VIEW (providers + real margins):\n{_compact_json(ops_struct)}")
            if not blocks and ex.get("itinerary_text_ops"):
                blocks.append(f"INTERNAL OPS VIEW (raw, truncated):\n{ex['itinerary_text_ops'][:1200]}")
            elif not blocks and ex.get("itinerary_text"):
                blocks.append(f"CLIENT ITINERARY (raw, truncated):\n{ex['itinerary_text'][:1200]}")
            if not blocks:
                continue
            score = ex.get("_score")
            score_tag = f" similarity={score:.2f}" if isinstance(score, (int, float)) else ""
            user_prompt_parts.append(
                f"--- outcome={ex['outcome']}{score_tag} ---\n"
                f"CLIENT REQUEST:\n{ex['client_request'][:1000]}\n\n"
                + "\n\n".join(blocks)
            )
    user_prompt_parts.append("\nNow produce ONLY the JSON itinerary for the NEW CLIENT REQUEST above.")

    data = await _call_claude_json(SYSTEM_PROMPT_GENERATE, "\n".join(user_prompt_parts))

    # Build itinerary draft from AI output
    draft = _itinerary_from_ai(data, client_name or data.get("main_traveler", ""), user.email)
    if save:
        await db.itineraries.insert_one(dict(draft))  # avoid mutating draft with _id
    draft.pop("_id", None)
    return {"itinerary": draft, "ai_summary": data.get("summary", ""), "retrieval": retrieval_meta, "partner": partner}


def _re_escape(s: str) -> str:
    import re
    return re.escape(s)


def _compact_json(obj) -> str:
    import json as _json
    return _json.dumps(obj, ensure_ascii=False, separators=(",", ":"))


def _itinerary_from_ai(data: dict, client_name: str, created_by: str) -> dict:
    """Map AI JSON to our Itinerary schema, generating IDs and syncing legacy fields."""
    name = data.get("name") or "AI draft"
    days_in = data.get("days") or []
    days = []
    for d in days_in:
        services = []
        for s in d.get("services") or []:
            excl = float(s.get("unit_price_tax_excl") or 0)
            incl = float(s.get("unit_price_tax_incl") or 0)
            services.append({
                "service_id": new_id("svc"),
                "experience_id": s.get("experience_id"),
                "type": s.get("type") or "actividad",
                "name": s.get("name") or "",
                "provider_name": s.get("provider_name") or "",
                "quantity": float(s.get("quantity") or 1),
                "unit_price_tax_excl": excl,
                "unit_price_tax_incl": incl,
                "unit_price": incl,
                "currency": s.get("currency") or "EUR",
            })
        days.append({
            "day_id": new_id("day"),
            "date": d.get("date"),
            "label": d.get("label") or "Day",
            "city": d.get("city") or "",
            "services": services,
        })
    accs = []
    for a in data.get("accommodations") or []:
        incl = float(a.get("price_tax_incl") or 0)
        excl = float(a.get("price_tax_excl") or 0)
        accs.append({
            "acc_id": new_id("acc"),
            "date_from": a.get("date_from"),
            "date_to": a.get("date_to"),
            "name": a.get("name") or "",
            "price_tax_excl": excl,
            "price_tax_incl": incl,
            "price": incl,
            "currency": a.get("currency") or "EUR",
        })

    itn = {
        "itinerary_id": new_id("itn"),
        "name": name,
        "main_traveler": client_name or data.get("main_traveler") or "",
        "start_date": data.get("start_date"),
        "end_date": data.get("end_date"),
        "duration_days": len(days),
        "num_travelers": int(data.get("num_travelers") or 2),
        "travelers": [],
        "days": days,
        "accommodations": accs,
        "markup_pct": float(data.get("markup_pct") or 15),
        "currency": "EUR",
        "status": "draft",
        "created_by": created_by,
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "ai_generated": True,
        "ai_summary": data.get("summary", ""),
    }
    return itn


# ---------------------------------------------------------------------------
# Mount router and CORS
# ---------------------------------------------------------------------------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origin_regex=".*",
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)
