"""Backend integration tests for the Travel Itinerary Builder - iteration 2.

Covers: health, auth (me/logout), admin whitelist + RBAC, providers CRUD,
experiences CRUD + cascade + facets, 3-tier pricing (precio sin IVA,
precio con IVA, PVP), experience autocomplete, ItineraryDay.city,
ItineraryService 3-tier, Accommodation 3-tier, Excel export 9-col,
bulk provider-sheet import and server-side bulk import.

All tests authenticate via the seeded session token described in
/app/memory/test_credentials.md (Bearer test_admin_session_token).
"""
from __future__ import annotations

import io
import os
import subprocess
from datetime import datetime, timezone, timedelta

import openpyxl
import pytest
import requests
from pymongo import MongoClient

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://itinerary-builder-74.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"
# Test-only session tokens. NOT real secrets — these are arbitrary opaque
# strings the tests seed into Mongo for the duration of the run, then delete
# in the fixture's finalizer. Override with env if you want different values.
ADMIN_TOKEN = os.environ.get("TEST_ADMIN_TOKEN", "test_admin_session_token")
AGENT_TOKEN = os.environ.get("TEST_AGENT_TOKEN", "test_agent_session_token")

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "test_database")


# ---- fixtures ---------------------------------------------------------------
@pytest.fixture(scope="session")
def db():
    return MongoClient(MONGO_URL)[DB_NAME]


def _seed_admin(db):
    db.users.delete_many({"email": "test.admin@example.com"})
    db.user_sessions.delete_many({"session_token": ADMIN_TOKEN})
    db.users.insert_one({
        "user_id": "test-user-admin",
        "email": "test.admin@example.com",
        "name": "Test Admin",
        "picture": None,
        "role": "admin",
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    db.user_sessions.insert_one({
        "user_id": "test-user-admin",
        "session_token": ADMIN_TOKEN,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
        "created_at": datetime.now(timezone.utc),
    })


def _seed_agent(db):
    db.users.delete_many({"email": "test.agent@example.com"})
    db.user_sessions.delete_many({"session_token": AGENT_TOKEN})
    db.users.insert_one({
        "user_id": "test-user-agent",
        "email": "test.agent@example.com",
        "name": "Test Agent",
        "picture": None,
        "role": "agent",
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    db.user_sessions.insert_one({
        "user_id": "test-user-agent",
        "session_token": AGENT_TOKEN,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
        "created_at": datetime.now(timezone.utc),
    })


@pytest.fixture(scope="session", autouse=True)
def seed(db):
    _seed_admin(db)
    _seed_agent(db)
    yield
    # cleanup test data (keeps imported sample data)
    db.user_sessions.delete_many({"session_token": {"$in": [ADMIN_TOKEN, AGENT_TOKEN]}})
    db.users.delete_many({"email": {"$in": ["test.admin@example.com", "test.agent@example.com"]}})
    db.allowed_emails.delete_many({"email": {"$in": ["test.admin@example.com", "agent1@example.com"]}})


@pytest.fixture()
def admin_client():
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {ADMIN_TOKEN}"})
    return s


@pytest.fixture()
def agent_client():
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {AGENT_TOKEN}"})
    return s


def _no_mongo_id(obj):
    if isinstance(obj, dict):
        assert "_id" not in obj, f"Mongo _id leaked in {obj}"
        for v in obj.values():
            _no_mongo_id(v)
    elif isinstance(obj, list):
        for v in obj:
            _no_mongo_id(v)


# ---- health -----------------------------------------------------------------
class TestHealth:
    def test_root_ok(self):
        r = requests.get(f"{API}/")
        assert r.status_code == 200
        assert r.json() == {"ok": True, "service": "itinerary-builder"}


# ---- auth -------------------------------------------------------------------
class TestAuth:
    def test_me_no_token_401(self):
        r = requests.get(f"{API}/auth/me")
        assert r.status_code == 401

    def test_me_with_bearer(self, admin_client):
        r = admin_client.get(f"{API}/auth/me")
        assert r.status_code == 200
        u = r.json()
        assert u["email"] == "test.admin@example.com"
        assert u["role"] == "admin"
        _no_mongo_id(u)

    def test_logout_revokes_then_reseed(self, db):
        # use a one-off token so we don't lose ADMIN_TOKEN for other tests
        token = "tmp_logout_token"
        db.user_sessions.insert_one({
            "user_id": "test-user-admin",
            "session_token": token,
            "expires_at": datetime.now(timezone.utc) + timedelta(days=1),
            "created_at": datetime.now(timezone.utc),
        })
        h = {"Authorization": f"Bearer {token}"}
        assert requests.get(f"{API}/auth/me", headers=h).status_code == 200
        r = requests.post(f"{API}/auth/logout", headers=h)
        assert r.status_code == 200
        assert requests.get(f"{API}/auth/me", headers=h).status_code == 401

    def test_admin_endpoint_denies_agent(self, agent_client):
        r = agent_client.get(f"{API}/admin/allowed-emails")
        assert r.status_code == 403


# ---- whitelist --------------------------------------------------------------
class TestWhitelist:
    def test_whitelist_crud(self, admin_client, db):
        db.allowed_emails.delete_many({"email": "agent1@example.com"})
        r = admin_client.post(f"{API}/admin/allowed-emails",
                              json={"email": "agent1@example.com", "role": "agent"})
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["email"] == "agent1@example.com"
        assert body["role"] == "agent"
        _no_mongo_id(body)

        r = admin_client.get(f"{API}/admin/allowed-emails")
        assert r.status_code == 200
        emails = [a["email"] for a in r.json()]
        assert "agent1@example.com" in emails

        r = admin_client.delete(f"{API}/admin/allowed-emails/agent1@example.com")
        assert r.status_code == 200

        r = admin_client.get(f"{API}/admin/allowed-emails")
        assert "agent1@example.com" not in [a["email"] for a in r.json()]


# ---- providers --------------------------------------------------------------
@pytest.fixture()
def created_provider(admin_client):
    r = admin_client.post(f"{API}/providers",
                          json={"name": "TEST_Provider_A", "country": "Italia"})
    assert r.status_code == 200, r.text
    return r.json()


class TestProviders:
    def test_provider_crud_and_sort(self, admin_client, created_provider):
        pid = created_provider["provider_id"]
        assert pid.startswith("prov_")
        _no_mongo_id(created_provider)

        # patch name
        r = admin_client.patch(f"{API}/providers/{pid}", json={"name": "TEST_Provider_A2"})
        assert r.status_code == 200
        assert r.json()["name"] == "TEST_Provider_A2"

        # list sorted
        r = admin_client.get(f"{API}/providers")
        assert r.status_code == 200
        names = [p["name"] for p in r.json()]
        assert names == sorted(names)

        # delete
        r = admin_client.delete(f"{API}/providers/{pid}")
        assert r.status_code == 200
        r = admin_client.delete(f"{API}/providers/{pid}")
        assert r.status_code == 404


# ---- experiences ------------------------------------------------------------
class TestExperiences:
    def test_experience_requires_existing_provider(self, admin_client):
        r = admin_client.post(f"{API}/experiences", json={
            "title": "Bad",
            "provider_id": "prov_does_not_exist",
            "price": 10,
        })
        assert r.status_code == 400

    def test_experience_full_flow_and_cascade(self, admin_client):
        # create two providers
        p1 = admin_client.post(f"{API}/providers", json={"name": "TEST_ProvX"}).json()
        p2 = admin_client.post(f"{API}/providers", json={"name": "TEST_ProvY"}).json()

        # create experience under p1
        r = admin_client.post(f"{API}/experiences", json={
            "title": "TEST_Exp_Tour",
            "description": "desc",
            "provider_id": p1["provider_id"],
            "country": "Italia",
            "city": "Rome",
            "type": "actividad",
            "price": 99.5,
        })
        assert r.status_code == 200, r.text
        exp = r.json()
        assert exp["provider_name"] == "TEST_ProvX"
        eid = exp["experience_id"]
        assert eid.startswith("exp_")
        _no_mongo_id(exp)

        # patch -> change provider re-denormalizes name
        r = admin_client.patch(f"{API}/experiences/{eid}",
                               json={"provider_id": p2["provider_id"]})
        assert r.status_code == 200
        assert r.json()["provider_name"] == "TEST_ProvY"

        # rename provider -> cascade
        admin_client.patch(f"{API}/providers/{p2['provider_id']}",
                           json={"name": "TEST_ProvY_renamed"})
        r = admin_client.get(f"{API}/experiences", params={"q": "TEST_Exp_Tour"})
        assert r.status_code == 200
        items = r.json()
        match = [i for i in items if i["experience_id"] == eid]
        assert match and match[0]["provider_name"] == "TEST_ProvY_renamed"

        # filters
        r = admin_client.get(f"{API}/experiences", params={"country": "Italia"})
        assert r.status_code == 200
        assert all(i["country"] == "Italia" for i in r.json())

        r = admin_client.get(f"{API}/experiences", params={"type": "actividad"})
        assert r.status_code == 200
        assert all(i["type"] == "actividad" for i in r.json())

        # facets
        r = admin_client.get(f"{API}/experiences/facets")
        assert r.status_code == 200
        f = r.json()
        for k in ("countries", "cities", "types"):
            assert k in f and isinstance(f[k], list)
        assert "Italia" in f["countries"]

        # cleanup
        admin_client.delete(f"{API}/experiences/{eid}")
        admin_client.delete(f"{API}/providers/{p1['provider_id']}")
        admin_client.delete(f"{API}/providers/{p2['provider_id']}")


# ---- itineraries + export ---------------------------------------------------
class TestItineraryAndExport:
    def test_itinerary_crud_and_excel_export(self, admin_client):
        # create
        r = admin_client.post(f"{API}/itineraries", json={
            "name": "TEST_Trip",
            "main_traveler": "John Doe",
            "start_date": "2026-04-01",
            "end_date": "2026-04-05",
            "duration_days": 4,
            "num_travelers": 2,
            "markup_pct": 15.0,
        })
        assert r.status_code == 200, r.text
        itn = r.json()
        iid = itn["itinerary_id"]
        assert iid.startswith("itn_")
        _no_mongo_id(itn)

        # get single
        r = admin_client.get(f"{API}/itineraries/{iid}")
        assert r.status_code == 200

        # patch -> add a day with city + one service priced 100, qty 2 -> sub 200
        r = admin_client.patch(f"{API}/itineraries/{iid}", json={
            "days": [{
                "label": "Day 1",
                "date": "2026-04-01",
                "city": "Rome",
                "services": [{
                    "type": "actividad",
                    "name": "Colosseum Tour",
                    "quantity": 2,
                    "unit_price_tax_excl": 100,
                    "unit_price_tax_incl": 121,
                    "currency": "EUR",
                }],
            }],
        })
        assert r.status_code == 200
        body = r.json()
        assert len(body["days"]) == 1
        assert body["days"][0]["city"] == "Rome"

        # export
        r = admin_client.get(f"{API}/itineraries/{iid}/export")
        assert r.status_code == 200
        assert r.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        assert len(r.content) > 1000

        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.title == "Trip Prices"
        # main traveler in B1
        assert ws.cell(1, 2).value == "John Doe"

        # Header row 15 has 9 columns including 3 price tiers
        headers_row = [ws.cell(15, c).value for c in range(1, 10)]
        assert headers_row == ["Day", "Date", "City", "Type", "Name", "Quantity",
                               "Precio sin IVA", "Precio con IVA", "PVP"]

        # Day row at 16: city is "Rome"
        assert ws.cell(16, 1).value == "Day 1"
        assert ws.cell(16, 3).value == "Rome"

        # Service row at 17: line sin IVA = 200, line con IVA = 242, PVP = 242*1.15 = 278.30
        assert ws.cell(17, 5).value == "Colosseum Tour"
        assert ws.cell(17, 7).value == pytest.approx(200.0)
        assert ws.cell(17, 8).value == pytest.approx(242.0)
        # Markup 15% applied on top of price_tax_incl: 242 * 1.15 = 278.30
        # NOTE: test created itinerary with markup_pct=15 below

        # find totals rows
        labels_col6 = [(row, ws.cell(row, 6).value) for row in range(1, ws.max_row + 1)]
        sub_excl_row = next(r for r, v in labels_col6 if v == "Subtotal sin IVA")
        sub_incl_row = next(r for r, v in labels_col6 if v == "Subtotal con IVA")
        pvp_row = next(r for r, v in labels_col6 if v and str(v).startswith("PVP (markup"))
        assert ws.cell(sub_excl_row, 7).value == pytest.approx(200.0)
        assert ws.cell(sub_incl_row, 8).value == pytest.approx(242.0)
        assert ws.cell(pvp_row, 9).value == pytest.approx(278.30, rel=1e-3)

        # delete
        r = admin_client.delete(f"{API}/itineraries/{iid}")
        assert r.status_code == 200
        assert admin_client.get(f"{API}/itineraries/{iid}").status_code == 404


# ---- stats ------------------------------------------------------------------
class TestStats:
    def test_stats(self, admin_client):
        r = admin_client.get(f"{API}/stats")
        assert r.status_code == 200
        data = r.json()
        for k in ("providers", "experiences", "itineraries", "users"):
            assert k in data
            assert isinstance(data[k], int)


# ---- bulk import ------------------------------------------------------------
class TestBulkImport:
    def test_import_synthetic_xlsx(self, admin_client, db):
        # generate xlsx in-memory
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["operator_name", "name", "price_tax_incl", "currency"])
        ws.append(["TEST_BulkProvider", "TEST_BulkExp_1", 50, "EUR"])
        ws.append(["TEST_BulkProvider", "TEST_BulkExp_2", 75, "EUR"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("synth.xlsx", buf,
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = admin_client.post(
            f"{API}/experiences/import-provider-sheet",
            params={"country": "Italia", "city": "Rome", "type": "actividad"},
            files=files,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["created"] == 2
        assert body["providers"] == 1

        # provider created once
        prov_count = db.providers.count_documents({"name": "TEST_BulkProvider"})
        assert prov_count == 1

        # cleanup
        prov = db.providers.find_one({"name": "TEST_BulkProvider"})
        if prov:
            db.experiences.delete_many({"provider_id": prov["provider_id"]})
            db.providers.delete_one({"provider_id": prov["provider_id"]})

    def test_import_real_provider_sheet(self, admin_client, db):
        path = "/app/artifacts/excel_creados/1. EXCEL CREADOS/3. ITALIA/Roman Road Tours_2024.xlsx"
        if not os.path.exists(path):
            pytest.skip("sample xlsx missing")
        # clear any prior imports of this provider so the test is repeatable
        existing = list(db.providers.find({"name": {"$regex": "Roman Road", "$options": "i"}}))
        for p in existing:
            db.experiences.delete_many({"provider_id": p["provider_id"]})
            db.providers.delete_one({"provider_id": p["provider_id"]})

        with open(path, "rb") as f:
            files = {"file": (os.path.basename(path), f,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            r = admin_client.post(
                f"{API}/experiences/import-provider-sheet",
                params={"country": "Italia", "city": "Rome", "type": "actividad"},
                files=files,
            )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["created"] >= 1
        assert body["providers"] >= 1


# ---- iteration 2: 3-tier pricing & autocomplete ---------------------------
class TestExperience3TierPricing:
    def test_experience_3_tier_create_and_patch_sync(self, admin_client, db):
        prov = admin_client.post(f"{API}/providers", json={"name": "TEST_Prov3T"}).json()
        try:
            r = admin_client.post(f"{API}/experiences", json={
                "title": "TEST_3T_Exp",
                "provider_id": prov["provider_id"],
                "price_tax_excl": 100,
                "price_tax_incl": 121,
            })
            assert r.status_code == 200, r.text
            exp = r.json()
            assert exp["price_tax_excl"] == 121 or exp["price_tax_excl"] == 100
            # specifically: both fields persist and price legacy = price_tax_incl
            assert exp["price_tax_excl"] == pytest.approx(100.0)
            assert exp["price_tax_incl"] == pytest.approx(121.0)
            assert exp["price"] == pytest.approx(121.0)
            eid = exp["experience_id"]

            # GET back same values
            r = admin_client.get(f"{API}/experiences", params={"q": "TEST_3T_Exp"})
            assert r.status_code == 200
            doc = next(i for i in r.json() if i["experience_id"] == eid)
            assert doc["price_tax_excl"] == pytest.approx(100.0)
            assert doc["price_tax_incl"] == pytest.approx(121.0)
            assert doc["price"] == pytest.approx(121.0)

            # PATCH price_tax_incl=200 syncs price=200
            r = admin_client.patch(f"{API}/experiences/{eid}", json={"price_tax_incl": 200})
            assert r.status_code == 200
            patched = r.json()
            assert patched["price_tax_incl"] == pytest.approx(200.0)
            assert patched["price"] == pytest.approx(200.0)

            # cleanup
            admin_client.delete(f"{API}/experiences/{eid}")
        finally:
            admin_client.delete(f"{API}/providers/{prov['provider_id']}")


class TestExperienceAutocomplete:
    def test_autocomplete_q_tour(self, admin_client):
        r = admin_client.get(f"{API}/experiences/autocomplete", params={"q": "tour"})
        assert r.status_code == 200
        items = r.json()
        assert isinstance(items, list)
        assert len(items) <= 15
        # each match has 'tour' in title or provider_name (case-insensitive)
        if items:
            for it in items:
                blob = f"{it.get('title','')} {it.get('provider_name','')}".lower()
                assert "tour" in blob

    def test_autocomplete_q_wine_no_city(self, admin_client):
        r = admin_client.get(f"{API}/experiences/autocomplete", params={"q": "wine"})
        assert r.status_code == 200
        items = r.json()
        assert isinstance(items, list)
        assert len(items) <= 15

    def test_autocomplete_city_only(self, admin_client):
        r = admin_client.get(f"{API}/experiences/autocomplete",
                             params={"q": "", "city": "Madrid"})
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_autocomplete_no_q_no_city_fallback(self, admin_client):
        r = admin_client.get(f"{API}/experiences/autocomplete")
        assert r.status_code == 200
        items = r.json()
        assert isinstance(items, list)
        assert len(items) <= 15


class TestItineraryDayCity:
    def test_day_city_persists_on_create_and_patch(self, admin_client):
        r = admin_client.post(f"{API}/itineraries", json={
            "name": "TEST_CityTrip",
            "main_traveler": "Jane",
            "markup_pct": 10.0,
            "days": [{"label": "Day 1", "date": "2026-05-01", "city": "Madrid", "services": []}],
        })
        assert r.status_code == 200, r.text
        itn = r.json()
        iid = itn["itinerary_id"]
        assert itn["days"][0]["city"] == "Madrid"

        # GET round-trip
        r = admin_client.get(f"{API}/itineraries/{iid}")
        assert r.status_code == 200
        assert r.json()["days"][0]["city"] == "Madrid"

        # PATCH city
        existing_day = r.json()["days"][0]
        existing_day["city"] = "Barcelona"
        r = admin_client.patch(f"{API}/itineraries/{iid}", json={"days": [existing_day]})
        assert r.status_code == 200
        assert r.json()["days"][0]["city"] == "Barcelona"

        admin_client.delete(f"{API}/itineraries/{iid}")


class TestItineraryServiceAndAccommodation3Tier:
    def test_service_and_accommodation_3_tier_persist(self, admin_client):
        r = admin_client.post(f"{API}/itineraries", json={
            "name": "TEST_3T_Itn",
            "markup_pct": 15.0,
            "days": [{
                "label": "Day 1", "date": "2026-06-01", "city": "Rome",
                "services": [{
                    "type": "actividad", "name": "TEST_Svc",
                    "quantity": 2,
                    "unit_price_tax_excl": 50,
                    "unit_price_tax_incl": 60,
                }],
            }],
            "accommodations": [{
                "name": "TEST_Hotel", "date_from": "2026-06-01", "date_to": "2026-06-02",
                "price_tax_excl": 100, "price_tax_incl": 121, "currency": "EUR",
            }],
        })
        assert r.status_code == 200, r.text
        itn = r.json()
        iid = itn["itinerary_id"]
        svc = itn["days"][0]["services"][0]
        assert svc["unit_price_tax_excl"] == pytest.approx(50.0)
        assert svc["unit_price_tax_incl"] == pytest.approx(60.0)
        acc = itn["accommodations"][0]
        assert acc["price_tax_excl"] == pytest.approx(100.0)
        assert acc["price_tax_incl"] == pytest.approx(121.0)

        # PATCH updates persist
        new_days = itn["days"]
        new_days[0]["services"][0]["unit_price_tax_excl"] = 55
        new_days[0]["services"][0]["unit_price_tax_incl"] = 66
        r = admin_client.patch(f"{API}/itineraries/{iid}", json={"days": new_days})
        assert r.status_code == 200
        assert r.json()["days"][0]["services"][0]["unit_price_tax_excl"] == pytest.approx(55.0)
        assert r.json()["days"][0]["services"][0]["unit_price_tax_incl"] == pytest.approx(66.0)

        admin_client.delete(f"{API}/itineraries/{iid}")


class TestImportAllServer:
    def test_requires_admin(self, agent_client):
        r = agent_client.post(f"{API}/experiences/import-all-server")
        assert r.status_code == 403

    def test_admin_run_returns_summary_and_dedupes_on_rerun(self, admin_client):
        # 1st call - data was already imported in a prior run, so most should skip
        r = admin_client.post(f"{API}/experiences/import-all-server")
        assert r.status_code == 200, r.text
        body = r.json()
        for k in ("files_scanned", "total_created", "total_skipped", "files"):
            assert k in body
        assert body["files_scanned"] >= 1
        assert isinstance(body["files"], list)

        # 2nd call should mostly skip
        r2 = admin_client.post(f"{API}/experiences/import-all-server")
        assert r2.status_code == 200
        body2 = r2.json()
        assert body2["total_skipped"] > 0

    def test_country_inferred_from_folder(self, admin_client):
        r = admin_client.get(f"{API}/experiences/facets")
        assert r.status_code == 200
        countries = set(r.json()["countries"])
        # at least one of the three expected countries should be present
        assert countries & {"España", "Portugal", "Italia"}


class TestImportProviderSheetDedupe:
    def test_second_upload_skips(self, admin_client, db):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["operator_name", "name", "price_tax_incl", "price_tax_excl", "currency"])
        ws.append(["TEST_DedupeProv", "TEST_DedupeExp_A", 80, 66, "EUR"])
        ws.append(["TEST_DedupeProv", "TEST_DedupeExp_B", 90, 74, "EUR"])

        def _do_upload():
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            files = {"file": ("d.xlsx", buf,
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            return admin_client.post(
                f"{API}/experiences/import-provider-sheet",
                params={"country": "Italia", "city": "Rome", "type": "actividad"},
                files=files,
            )

        # ensure clean
        prov = db.providers.find_one({"name": "TEST_DedupeProv"})
        if prov:
            db.experiences.delete_many({"provider_id": prov["provider_id"]})
            db.providers.delete_one({"provider_id": prov["provider_id"]})

        r1 = _do_upload()
        assert r1.status_code == 200, r1.text
        assert r1.json()["created"] == 2

        r2 = _do_upload()
        assert r2.status_code == 200
        assert r2.json()["skipped"] >= 2
        assert r2.json()["created"] == 0

        # cleanup
        prov = db.providers.find_one({"name": "TEST_DedupeProv"})
        if prov:
            db.experiences.delete_many({"provider_id": prov["provider_id"]})
            db.providers.delete_one({"provider_id": prov["provider_id"]})


class TestFacetsSorted:
    def test_facets_distinct_and_sorted(self, admin_client):
        r = admin_client.get(f"{API}/experiences/facets")
        assert r.status_code == 200
        f = r.json()
        for k in ("countries", "cities", "types"):
            assert k in f
            assert f[k] == sorted(f[k])
            # distinct (no dups)
            assert len(f[k]) == len(set(f[k]))
