import openpyxl, sys, os, json

def dump(path, max_rows=30, max_cols=15):
    print(f"\n{'='*80}\nFILE: {path}\n{'='*80}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print("ERROR:", e); return
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n-- Sheet: {sname} ({ws.max_row} rows x {ws.max_column} cols) --")
        for r in range(1, min(ws.max_row, max_rows) + 1):
            row = []
            for c in range(1, min(ws.max_column, max_cols) + 1):
                v = ws.cell(r, c).value
                if v is None: v = ""
                v = str(v)[:40]
                row.append(v)
            if any(x.strip() for x in row):
                print(f"  R{r}: " + " | ".join(row))

# Itinerary samples
for f in ["helensmith.xlsx", "terrylin.xlsx", "viajes_directos.xlsx"]:
    dump(f"/app/artifacts/{f}", max_rows=60, max_cols=12)

# Provider samples
provs = [
    "/app/artifacts/excel_creados/1. EXCEL CREADOS/3. ITALIA/Roman Road Tours_2024.xlsx",
    "/app/artifacts/excel_creados/1. EXCEL CREADOS/3. ITALIA/Eyes of Rome - Studio Associato _Eyes of Rome_.xlsx",
]
for p in provs:
    dump(p, max_rows=30, max_cols=10)
